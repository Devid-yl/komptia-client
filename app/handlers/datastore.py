"""Handlers pour le gestionnaire de fichiers / données utilisateur.

Conventions équipe sénior (aligné session A ``dashboard.py`` + session B
``contacts.py``) :

* **Thin handler** : parse → validation → ``StorageManager`` → response.
  La logique FS bas-niveau est dans les helpers du module (testables sans
  HTTP) ; la logique métier quota/DB est dans ``StorageManager``.
* **Validation de chemin fail-closed** : ``_safe_path`` et
  ``FileValidator.validate_path`` utilisent ``Path.is_relative_to`` (Python
  3.9+) plutôt que ``str.startswith`` — ce dernier laissait passer
  ``/data/user11`` quand ``user_dir = /data/user1`` (CWE-22).
* **Content-Disposition sûr** : ``_content_disposition`` strip les
  caractères de contrôle (CR/LF/NUL) avant de composer l'en-tête (CWE-93
  HTTP header injection). Le ``filename*`` RFC 5987 est UTF-8-quoté, le
  ``filename`` legacy ASCII-only est déduit d'une variante safe.
* **Aperçu Excel robuste** : ``_preview_excel`` préserve les zéros
  (``"" if c is None else str(c)`` — le précédent ``str(c) if c else ""``
  perdait ``0``, ``False``, ``""``) et borne le nombre de lignes itérées
  pour éviter un zip-bomb xlsx (ratio compression / RAM).
* **Rate-limit** sur les endpoints coûteux (upload, SQL exec, save-search,
  preview) via ``_check_rate_limit(limiter, user_id, *quota)`` — pattern
  aligné sur ``app/handlers/contacts.py`` itér 1.
* **Role decorators eager** : tous les endpoints mutateurs utilisent
  ``@require_role("admin", "user")`` ; ``@authenticated`` est réservé
  aux lectures pures (list, preview, download, folders, context-files).
  Les ``reader`` n'écrivent jamais — plus de ``MoveAPIHandler`` silently
  ouvert à ``@authenticated``.
* **Logging ``extra=``** : ``user_id``, ``request_id`` et clés métier
  passent par ``logger.xxx(..., extra=...)`` plutôt que dans le message
  — homogène avec les autres handlers et parseable.
* **Imports top-level** : ``urllib.parse.quote``, ``base64``, ``openpyxl``
  (lazy guard autour de l'import) en tête de module plutôt qu'en corps
  de fonction — évite les coûts répétés et rend les dépendances
  auditables par les outils statiques.
"""

from __future__ import annotations

import asyncio
import base64
import csv
import io
import json
import mimetypes
import os
import re
import shutil
import uuid
from pathlib import Path
from typing import Any, Final, Optional
from urllib.parse import quote

import tornado.web
from sqlalchemy import select
from sqlalchemy.exc import OperationalError, SQLAlchemyError

from app.config import DATA_DIR, PYODBC_AVAILABLE
from app.core import clock
from app.constants import FILE_CHUNK_BYTES
from app.core.database import get_session
from app.core.db_retry import retry_on_locked
from app.handlers.base import BaseHandler, authenticated, require_role
from app.models.audit import AuditAction, AuditLog
from app.models.search_history import SearchHistory
from app.models.user_storage import FileMetadata
from app.services.export.csv_export import to_csv_bytes
from app.services.storage_manager import (
    StorageManager,
    calculate_hash_from_bytes,
)
from app.utils.gzip_safe import gunzip_first_member
from app.utils.http_streaming import stream_file_to_handler
from app.utils.logger import get_logger
from app.utils.rate_limiter import RateLimiter

if PYODBC_AVAILABLE:
    from app.services.database.query_executor import QueryExecutor


logger = get_logger(__name__)


# ── Racine datastore ─────────────────────────────────────────────
DATASTORE_DIR: Final[Path] = DATA_DIR / "datastore"


# ── Extensions autorisées ────────────────────────────────────────
ALLOWED_EXTENSIONS: Final[frozenset[str]] = frozenset(
    {
        ".csv",
        ".xlsx",
        ".xls",
        ".json",
        ".xml",
        ".txt",
        ".md",
        ".log",
        ".pdf",
        ".doc",
        ".docx",
        ".png",
        ".jpg",
        ".jpeg",
        ".sql",
    }
)


# ── Magic bytes (détection réelle de format) ─────────────────────
# On conserve un ``dict`` mutable-looking mais on ne le mute jamais côté
# code : le dict reste ``MAGIC_BYTES[prefix] = [ext, ...]`` car openpyxl
# et doc OOXML partagent le même préfixe ``PK`` (zip archive).
MAGIC_BYTES: Final[dict[bytes, list[str]]] = {
    b"%PDF": [".pdf"],
    b"\x89PNG": [".png"],
    b"\xff\xd8\xff": [".jpg", ".jpeg"],
    b"PK": [".xlsx", ".docx", ".zip"],
    b"\xd0\xcf\x11\xe0": [".xls", ".doc"],
}


# ── Caps & budgets (partagés entre endpoints) ────────────────────
#: Budget partagé des payloads SQL (execute ET save) — un seul endroit
#: pour éviter la dérive si on bump la limite.
MAX_SQL_PAYLOAD_BYTES: Final[int] = 256 * 1024
MAX_SQL_FILENAME_LEN: Final[int] = 200
#: Cap commun sur les noms de fichiers user-facing (upload, rename,
#: save-search, mkdir). 200 caractères : en deçà du cap NTFS/ext4 (255)
#: pour laisser une marge aux suffixes ``_1.sql``, ``_2.csv``.
MAX_FILENAME_LEN: Final[int] = 200
#: Cap preview texte — au-delà, on tronque avec une mention explicite.
MAX_PREVIEW_TEXT_BYTES: Final[int] = 100_000
#: Cap lecture texte en RAM (évite DoS si user upload un 50 Mo CSV puis
#: hammer ``/preview`` — le fichier est lu ENTIER avant troncation).
MAX_PREVIEW_FILE_BYTES: Final[int] = 5 * 1024 * 1024
#: Cap image preview (base64 explose x4/3 : 5 Mo → ~7 Mo JSON).
MAX_PREVIEW_IMAGE_BYTES: Final[int] = 5 * 1024 * 1024
#: Preview xlsx : traité comme un classeur normal (intent user 2026-05-14).
#: Sentinelle "no cap caller" — la VRAIE protection contre zip-bombs est
#: ``MAX_PREVIEW_FILE_BYTES`` (5 MiB) qui borne la taille du FICHIER avant
#: parsing openpyxl. Pas besoin d'un cap row count distinct.
MAX_PREVIEW_XLSX_ROWS: Final[int] = 1_000_000_000
MAX_PREVIEW_XLSX_COLS: Final[int] = 200
#: ⚠️ Plus AUCUN cap de décompression .afz.json hardcodé ici. La SOURCE UNIQUE
#: de toutes les limites de classeur (disque ET taille décompressée en RAM, sur
#: TOUS les chemins : upload / download / download anonymisé / lecture) est le
#: quota de stockage par user défini par l'admin (/admin/performance →
#: STORAGE_QUOTA_PER_USER_BYTES), résolu au runtime via
#: ``storage_manager.get_storage_quota_bytes()``. Un seul bouton, une seule
#: responsabilité. ⚠️ Ce quota borne donc aussi le pic RAM de décompression :
#: l'admin doit le dimensionner vs la RAM du conteneur (cf. docstring de
#: ``get_storage_quota_bytes`` ; le défaut 500 Mio est sûr).
#: Cap récursion ``DatastoreFoldersAPIHandler.scan_folders`` — défense
#: contre symlink cycle et arborescence très profonde.
MAX_FOLDER_DEPTH: Final[int] = 20
MAX_FOLDERS_LISTED: Final[int] = 1000


# ── Rate-limit quotas (max_requests, window_seconds) ─────────────
RATE_LIMIT_UPLOAD: Final[tuple[int, int]] = (20, 60)
RATE_LIMIT_SQL_EXECUTE: Final[tuple[int, int]] = (30, 60)
RATE_LIMIT_SAVE_SEARCH: Final[tuple[int, int]] = (20, 60)
RATE_LIMIT_PREVIEW: Final[tuple[int, int]] = (60, 60)
#: Download/mutateurs : pas de cap historique → vecteur DoS interne user
#: hostile (boucle download 4 GiB .afz.json, batch delete 10k paths).
RATE_LIMIT_DOWNLOAD: Final[tuple[int, int]] = (60, 60)
RATE_LIMIT_RENAME: Final[tuple[int, int]] = (60, 60)
RATE_LIMIT_MOVE: Final[tuple[int, int]] = (30, 60)
RATE_LIMIT_DELETE: Final[tuple[int, int]] = (30, 60)
RATE_LIMIT_FOLDERS: Final[tuple[int, int]] = (60, 60)
#: Cap batch DELETE pour éviter DoS quota query.
MAX_DELETE_BATCH: Final[int] = 1000


# ── Rate-limiters instances (module-scope, thread-safe) ───────────
_upload_limiter = RateLimiter()
_sql_exec_limiter = RateLimiter()
_save_search_limiter = RateLimiter()
_preview_limiter = RateLimiter()
_download_limiter = RateLimiter()
_rename_limiter = RateLimiter()
_move_limiter = RateLimiter()
_delete_limiter = RateLimiter()
_folders_limiter = RateLimiter()


# ── Noms réservés Windows (portabilité — datastore peut tourner ───
#   sur SMB/CIFS Windows en prod). ``CON``, ``PRN``, ``AUX``, ``NUL``
#   sont interdits comme nom de fichier ou stem (même avec extension).
_WINDOWS_RESERVED_NAMES: Final[frozenset[str]] = frozenset(
    {"CON", "PRN", "AUX", "NUL"}
    | {f"COM{i}" for i in range(1, 10)}
    | {f"LPT{i}" for i in range(1, 10)}
)

#: Pattern pour détecter les caractères interdits dans un nom de fichier
#: user-fourni : NUL + tous les contrôles < 0x20 + DEL + séparateurs
#: path (``/`` ``\``). Inclut aussi les contrôles Unicode bidi-override
#: (U+202A..U+202E, U+2066..U+2069) et zero-width / line-separator
#: (U+200B..U+200F, U+2028..U+2029, U+FEFF) — vecteur de spoofing visuel
#: documenté NPM/GitHub : ``invoice<U+202E>fdp.exe`` s'affiche
#: ``invoiceexe.pdf`` dans la liste, le shell exécute pourtant ``.exe``.
#: Utilisé par ``_sanitize_user_filename``.
_FORBIDDEN_FILENAME_CHARS: Final[re.Pattern[str]] = re.compile(
    "[\x00-\x1f\x7f/\\\\"
    "\u200b-\u200f"  # zero-width + LRM/RLM
    "\u2028-\u2029"  # line / paragraph separator
    "\u202a-\u202e"  # bidi LRE/RLE/PDF/LRO/RLO
    "\u2066-\u2069"  # bidi LRI/RLI/FSI/PDI
    "\ufeff"  # BOM / ZWNBSP
    "]"
)


# ─── Helpers : rate-limit ─────────────────────────────────────────


def _check_rate_limit(
    limiter: RateLimiter, user_id: int, max_requests: int, window_seconds: int
) -> None:
    """Lève ``HTTPError(429)`` si le rate-limit utilisateur est dépassé.

    Pattern aligné sur ``app/handlers/contacts.py`` — un seul endroit
    où décider du status, du message et du format de la clé.
    """
    key = f"user:{user_id}"
    if not limiter.check(key, max_requests=max_requests, window_seconds=window_seconds):
        raise tornado.web.HTTPError(
            429,
            "Trop de requêtes. Veuillez patienter quelques secondes.",
        )


# ─── Helpers : path safety ────────────────────────────────────────


def _user_dir(user_id: int) -> Path:
    """Retourne le dossier datastore de l'utilisateur, le crée si nécessaire."""
    p = DATASTORE_DIR / str(user_id)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _normalize_rel_path(target: Path, user_dir: Path) -> str:
    """Retourne le chemin de ``target`` relatif à ``user_dir``, en POSIX.

    Single source of truth pour les paths persistés (``FileMetadata.file_path``,
    ``AuditLog.details["path"]``, ``StorageManager.register_*``). Le path
    ainsi obtenu est :
      * sans trailing slash ;
      * normalisé via ``Path.resolve`` côté caller (déjà appelé par
        :func:`_safe_path`) ;
      * en POSIX (``/``) — même sur Windows, pour cohérence DB.

    Caller doit avoir validé ``target`` via :func:`_safe_path` au préalable.
    """
    rel = target.relative_to(user_dir)
    return rel.as_posix()


def _atomic_move_failing_if_exists(src: Path, dst: Path) -> None:
    """Déplace ``src`` vers ``dst`` en échouant atomiquement si ``dst`` existe.

    Race-safe contre le pattern TOCTOU ``dst.exists()`` + ``os.rename``/
    ``shutil.move`` : ces deux derniers **remplacent silencieusement**
    ``dst`` sur POSIX, causant une perte de données silencieuse si deux
    requêtes concurrentes renomment vers la même cible.

    Stratégie :
      * **Sur la même partition** : ``os.link(src, dst)`` + ``os.unlink(src)``.
        ``link`` lève :class:`FileExistsError` atomiquement (au niveau VFS)
        si ``dst`` existe — garantie kernel, pas de race.
      * **Cross-partition** (link refusé : EXDEV) : fallback ``shutil.move``
        avec re-check ``dst.exists()`` AVANT — race subsiste mais bornée
        au cas multi-partition (rare en pratique pour un datastore user).
      * **Dossier** : ``os.link`` ne fonctionne que pour les fichiers
        réguliers ; pour les dossiers on utilise ``os.rename`` qui sur
        POSIX échoue avec :class:`OSError` (ENOTEMPTY) si ``dst`` existe
        non-vide, et sur Windows échoue toujours si ``dst`` existe.

    Lève :class:`FileExistsError` si ``dst`` existe.
    Lève :class:`OSError` pour les autres erreurs (permission, quota, etc.).
    """
    if src.is_dir():
        # POSIX ``rename(2)`` remplace **silencieusement** un dossier vide
        # à la cible (atomique mais déstructif). Pour préserver la
        # promesse fail-if-exists, pré-check explicite ``dst.exists()``
        # avant le rename. Race window résiduelle de quelques µs entre
        # le check et le rename : acceptable pour des dossiers (un
        # adversaire devrait créer dst dans cette fenêtre, et le
        # comportement de fallback est correct : rename succeed = pas
        # de perte de données utilisateur, juste un overwrite consenti).
        # Solution kernel-atomic (Linux ``renameat2`` + ``RENAME_NOREPLACE``)
        # non portable et non exposée par Python stdlib < 3.12.
        if dst.exists():
            raise FileExistsError(str(dst))
        os.rename(src, dst)
        return
    try:
        os.link(src, dst)
    except FileExistsError:
        raise
    except OSError as exc:
        # EXDEV (cross-device link not permitted) → fallback shutil.move
        # avec re-check (race résiduelle bornée au cas multi-partition).
        if getattr(exc, "errno", None) != 18:  # EXDEV
            raise
        if dst.exists():
            raise FileExistsError(str(dst))
        shutil.move(str(src), str(dst))
        return
    os.unlink(src)


# Set global de référence pour empêcher le GC asyncio des tasks audit en
# vol (cf. doctrine fire-and-forget : sans cette ancre, ``create_task``
# peut être collecté avant l'await interne, perdant l'audit).
_AUDIT_TASKS: set[asyncio.Task[Any]] = set()


def _schedule_audit_fire_and_forget(
    *,
    # A8-F3 — ``action`` est une CONSTANTE string de ``AuditAction`` (ex
    # ``AuditAction.FILE_UPLOAD == "file_upload"``), pas un enum. Annoté ``str``
    # (et non ``AuditAction``) pour éviter de re-suggérer un ``.value`` fautif.
    action: str,
    user_id: int,
    details: dict[str, object],
    ip_address: str | None,
    op_label: str,
    entity_id: int | str | None = None,
) -> None:
    """Schedule un audit log en background avec retry sur ``database is locked``.

    Pattern dupliqué jusqu'à présent uniquement dans
    ``_schedule_audit_file_download``. Factorisé ici pour servir aussi
    rename/move/delete (les mutateurs perdaient l'audit log quand le
    rollback de session emportait le ``db.add(AuditLog...)`` final).

    ``entity_id`` (optionnel) : pour les actions liées à une row précise
    (``FILE_UPLOAD`` → ``FileMetadata.id``), sert à la traçabilité du
    rapport admin. Si l'ID n'est pas encore connu au schedule (très
    rare — l'INSERT est fait dans une autre session juste avant),
    laisser à ``None``.

    Best-effort : si le retry persistant échoue, on log un warning et on
    abandonne la ligne — le TTL `cleanup_db_retention_job` garantit que
    l'absence ne crée pas d'invariant cassé long terme.
    """
    # A8-F3 — ``AuditAction`` est une classe de CONSTANTES string
    # (``FILE_UPLOAD = "file_upload"``), PAS un enum : ``action`` est déjà une
    # string, pas d'attribut ``.value``. Avant, ``action.value`` levait
    # ``AttributeError: 'str' object has no attribute 'value'`` → 500 SYNCHRONE
    # sur CHAQUE upload/rename/move/delete (le fichier était déjà écrit → opé
    # faite + 500 = UI en erreur + re-upload doublon). ``log_action`` attend
    # d'ailleurs une ``str`` (audit.py:62).
    op_name = f"audit_logs[{action} user={user_id}] ({op_label})"

    async def _commit_audit() -> None:
        async with get_session() as db:
            db.add(
                AuditLog.log_action(
                    action=action,
                    user_id=user_id,
                    entity_type="file",
                    entity_id=entity_id,
                    details=details,
                    ip_address=ip_address,
                )
            )

    async def _audit_task() -> None:
        try:
            await retry_on_locked(_commit_audit, max_attempts=5, operation_name=op_name)
        except OperationalError as exc:
            logger.warning(
                "Audit log %s : abandon après retries (database is locked) : %s",
                op_name,
                exc,
            )

    task = asyncio.create_task(_audit_task())
    _AUDIT_TASKS.add(task)
    task.add_done_callback(_AUDIT_TASKS.discard)


async def _storage_register_deletion_with_retry(
    user_id: int,
    relative_path: str,
    *,
    max_attempts: int = 5,
) -> bool:
    """Wrapper retry de :meth:`StorageManager.register_deletion`.

    Ouvre une session SQLAlchemy fraîche à chaque tentative — c'est
    OBLIGATOIRE car après une ``OperationalError("database is locked")``,
    la session courante est en état "rollback only" et tout réutilisation
    échoue avec :class:`InvalidRequestError`.

    L'opération sous-jacente est idempotente : SELECT FileMetadata par
    (user_id, file_path) puis UPDATE user_storage (decrement bytes/count)
    + DELETE FileMetadata. Si un retry trouve la row déjà supprimée par
    un commit antérieur, ``register_deletion`` renvoie ``False`` (no-op).

    Returns:
        ``True`` si la row a été effectivement supprimée à cette
        tentative, ``False`` si l'entry n'existait plus en BDD.
    """
    op_name = f"storage.register_deletion[user={user_id} path={relative_path}]"

    async def _coro() -> bool:
        async with get_session() as fresh_db:
            mgr = StorageManager(fresh_db, DATASTORE_DIR)
            return await mgr.register_deletion(user_id, relative_path)

    return await retry_on_locked(_coro, max_attempts=max_attempts, operation_name=op_name)


async def _storage_register_upload_with_retry(
    user_id: int,
    file_path: Path,
    relative_path: str,
    *,
    description: Optional[str] = None,
    file_size: Optional[int] = None,
    file_hash: Optional[str] = None,
    max_attempts: int = 5,
) -> FileMetadata:
    """Wrapper retry de :meth:`StorageManager.register_upload`.

    Idempotent côté BDD : INSERT FileMetadata + UPDATE quota dans la même
    transaction. Si un retry suit un rollback du tour précédent (lock),
    aucune row n'a été persistée et le rejeu est sans effet de bord.

    Le fichier sur disque est déjà écrit AVANT cet appel (atomic via
    ``.tmp + os.replace``) — le retry ne touche pas au disque.
    """
    op_name = f"storage.register_upload[user={user_id} path={relative_path}]"

    async def _coro() -> FileMetadata:
        async with get_session() as fresh_db:
            mgr = StorageManager(fresh_db, DATASTORE_DIR)
            return await mgr.register_upload(
                user_id,
                file_path,
                relative_path,
                description=description,
                file_size=file_size,
                file_hash=file_hash,
            )

    return await retry_on_locked(_coro, max_attempts=max_attempts, operation_name=op_name)


def _safe_path(user_dir: Path, rel_path: str) -> Optional[Path]:
    """Valide et résout un chemin relatif à l'intérieur du dossier utilisateur.

    Empêche :
      * les path traversals ``../../etc`` explicites ;
      * les path traversals par préfixe (``/data/user11`` ne "startswith"
        PLUS ``/data/user1`` — on utilise ``Path.is_relative_to`` 3.9+) ;
      * les null bytes et séparateurs suspects dans ``rel_path``
        (``/evil\0.sql`` tentait de casser ``Path.resolve``).

    Retourne ``None`` si le chemin n'est pas safe — le caller doit renvoyer
    ``400`` ou ``404`` sans exécuter l'opération.
    """
    if not isinstance(rel_path, str):
        return None
    # NUL byte : CWE-158. ``Path`` peut lever ``ValueError`` selon la plate-
    # forme, mais on préfère fail-closed avant même de toucher ``resolve``.
    if "\x00" in rel_path:
        return None
    cleaned = rel_path.strip("/").replace("\\", "/")
    # Rejeter ``..`` segment explicite AVANT résolution (évite d'appeler
    # ``resolve`` sur un chemin hostile qui pourrait déréférencer un
    # symlink racine).
    if ".." in cleaned.split("/"):
        return None
    try:
        resolved = (user_dir / cleaned).resolve()
        user_dir_resolved = user_dir.resolve()
    except (OSError, RuntimeError):
        return None
    if resolved != user_dir_resolved and not resolved.is_relative_to(user_dir_resolved):
        return None
    return resolved


def _sanitize_user_filename_with_reason(
    raw: Any,
    *,
    require_extension: bool = False,
    default_ext: Optional[str] = None,
) -> tuple[Optional[str], Optional[str]]:
    """Variante de :func:`_sanitize_user_filename` qui retourne aussi le motif.

    Retourne ``(name, None)`` sur succès, ``(None, reason)`` sinon. Le
    motif est une chaîne FR prête pour affichage utilisateur — permet
    au handler de surfacer un message d'erreur précis ("Nom de fichier
    réservé Windows" plutôt que "Paramètres invalides" générique).

    Les motifs forment un vocabulaire stable utilisé par les tests de
    régression — ne pas reformuler à la légère.
    """
    if not isinstance(raw, str):
        return None, "Nom invalide"
    name = raw.strip()
    # NTFS strip les espaces/points trailing silencieusement → ``file.csv``
    # et ``file.csv.`` résolvent le même fichier sur Windows mais Python
    # les voit distincts. Rejeter explicitement.
    if name != name.rstrip(". "):
        return None, "Le nom ne peut pas se terminer par un point ou un espace"
    if not name:
        return None, "Nom vide"
    if name in (".", ".."):
        return None, "Nom de fichier interdit"
    if _FORBIDDEN_FILENAME_CHARS.search(name):
        return None, "Caractères interdits dans le nom"
    if name.startswith("."):
        return None, "Les fichiers cachés (commençant par '.') ne sont pas autorisés"
    stem_upper = Path(name).stem.upper()
    if stem_upper in _WINDOWS_RESERVED_NAMES:
        return None, f"Nom réservé par Windows : {stem_upper}"
    if require_extension and default_ext:
        p = Path(name)
        if p.suffix.lower() != default_ext.lower():
            name = f"{name}{default_ext}"
    if len(name) > MAX_FILENAME_LEN:
        return None, f"Nom trop long (max {MAX_FILENAME_LEN} caractères)"
    p = Path(name)
    if not p.stem.strip(".").strip():
        return None, "Stem du nom vide"
    return name, None


def _sanitize_user_filename(
    raw: Any,
    *,
    require_extension: bool = False,
    default_ext: Optional[str] = None,
) -> Optional[str]:
    """Retourne un nom de fichier sûr ou ``None`` si invalide.

    Règles (defense in depth contre CWE-73/22/158/434) :

    * ``raw`` doit être un ``str`` non vide après ``strip``.
    * Rejet des séparateurs de chemin (``/``, ``\\``) et des entrées ``.``
      ``..`` — pas de path-traversal déguisé en nom.
    * Rejet des null bytes et caractères de contrôle (< 0x20 + 0x7F) :
      évite CWE-158 + les ``ValueError`` non-gérés de ``write_text`` sur
      null byte.
    * Rejet des Unicode bidi-override (U+202A..U+202E, U+2066..U+2069) et
      zero-width (U+200B..U+200F, U+FEFF) : vecteur de spoofing visuel
      documenté NPM/GitHub.
    * Rejet des trailing dots/spaces (NTFS strip silencieusement →
      collision).
    * Rejet des dotfiles (préfixe ``.``) : ``DatastoreListAPIHandler``
      skippe ``child.name.startswith(".")`` — un fichier dotfile serait
      invisible dans l'UI après création (donnée "fausse silencieusement").
    * Rejet des noms réservés Windows (``CON``, ``PRN``, ``AUX``, ``NUL``,
      ``COM1..9``, ``LPT1..9``) — portable vers SMB/CIFS.
    * Cap longueur à ``MAX_FILENAME_LEN`` (< 255 NTFS/ext4).
    * Si ``require_extension=True``, force la présence de ``default_ext``
      — utilisé par save-search (``.csv``) et save-sql (``.sql``).

    Caller a besoin du motif d'échec → utiliser
    :func:`_sanitize_user_filename_with_reason`.
    """
    name, _reason = _sanitize_user_filename_with_reason(
        raw, require_extension=require_extension, default_ext=default_ext
    )
    return name


def _content_disposition(filename: str) -> str:
    """Compose un en-tête ``Content-Disposition`` safe.

    Strip les caractères de contrôle (CR/LF/NUL) du filename AVANT de le
    placer dans l'en-tête — prévient CWE-93 (HTTP response header
    injection / splitting). Produit :

    * ``filename="ascii-fallback"`` — guillemets escapés, bytes non-ASCII
      remplacés par ``_`` ;
    * ``filename*=UTF-8''percent-encoded`` — RFC 5987 pour les clients
      modernes, garantit que le nom Unicode est préservé.
    """
    # Strip control chars + NUL (CR / LF = CWE-93 ; autres = log pollution).
    safe = "".join(c for c in filename if ord(c) >= 32 and c != "\x7f")
    # ASCII fallback pour le ``filename`` legacy : non-ASCII → ``_``.
    ascii_fallback = safe.encode("ascii", "replace").decode("ascii").replace("?", "_")
    # Échapper guillemets + backslash pour le token quoted-string.
    ascii_fallback = ascii_fallback.replace("\\", "\\\\").replace('"', '\\"')
    encoded = quote(safe, safe="")
    return f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{encoded}"


# ─── Helpers : introspection FS ───────────────────────────────────


def _file_info(p: Path, user_dir: Path) -> Optional[dict]:
    """Construit le dict d'info d'un fichier/dossier.

    Robuste à la suppression concurrente entre ``iterdir`` et ``stat``
    (``FileNotFoundError``) : retourne ``None`` plutôt que crash 500
    si le fichier disparaît entre-temps. Retourne ``None`` aussi sur
    ``OSError`` (permissions, symlink cassé).
    """
    try:
        stat = p.stat()
    except (FileNotFoundError, OSError):
        return None
    try:
        rel = p.relative_to(user_dir)
    except ValueError:
        return None
    is_file = p.is_file()
    return {
        "name": p.name,
        "path": str(rel),
        "is_dir": p.is_dir(),
        "size": stat.st_size if is_file else 0,
        "size_human": _human_size(stat.st_size) if is_file else "",
        "modified": clock.from_timestamp(stat.st_mtime).isoformat(),
        "modified_human": clock.local_from_timestamp(stat.st_mtime).strftime("%d/%m/%Y %H:%M"),
        "extension": p.suffix.lower() if is_file else "",
        "mime": (mimetypes.guess_type(str(p))[0] or "") if is_file else "",
    }


def _human_size(size: int) -> str:
    """Conversion octets → format lisible.

    Défensif sur les entrées négatives (artéfact de ``quota_remaining``
    qui clampe déjà à 0 mais d'autres callers pourraient passer un
    delta signé) : on clampe à 0 pour ne pas afficher ``-512 o``.
    """
    if size < 0:
        size = 0
    value: float = float(size)
    for unit in ("o", "Ko", "Mo", "Go"):
        if abs(value) < 1024:
            return f"{value:.0f} {unit}" if unit == "o" else f"{value:.1f} {unit}"
        value /= 1024
    return f"{value:.1f} To"


def _dir_stats(user_dir: Path) -> dict:
    """Statistiques globales du dossier utilisateur.

    ``rglob`` + ``stat`` par fichier peut coûter cher sur un dossier
    avec 10k fichiers : les callers ``async`` devraient wrapper cet
    appel dans ``asyncio.to_thread`` (voir ``DatastoreListAPIHandler``).
    """
    total_files = 0
    total_size = 0
    by_ext: dict[str, int] = {}
    try:
        for p in user_dir.rglob("*"):
            try:
                if not p.is_file():
                    continue
                total_files += 1
                total_size += p.stat().st_size
            except (FileNotFoundError, OSError):
                # Concurrent delete → skip cette entrée plutôt que crash.
                continue
            ext = p.suffix.lower() or "(sans ext)"
            by_ext[ext] = by_ext.get(ext, 0) + 1
    except (FileNotFoundError, OSError):
        # ``user_dir`` supprimé pendant l'itération → stats vides.
        pass
    return {
        "total_files": total_files,
        "total_size": total_size,
        "total_size_human": _human_size(total_size),
        "by_extension": dict(sorted(by_ext.items(), key=lambda x: -x[1])[:10]),
    }


def _walk_folders(
    user_dir: Path,
    *,
    max_depth: int = MAX_FOLDER_DEPTH,
    max_count: int = MAX_FOLDERS_LISTED,
) -> list[dict]:
    """Retourne les sous-dossiers (depth-first, limités en profondeur/taille).

    * Ignore les symlinks pointant hors du dossier utilisateur (CWE-59
      symlink traversal + symlink cycle → récursion infinie).
    * Cap de profondeur ``max_depth`` et cap global ``max_count`` pour
      éviter un scan coûteux sur un arborescence explosive.
    """
    folders: list[dict] = []
    user_dir_resolved = user_dir.resolve()

    def _recurse(path: Path, depth: int) -> bool:
        """Retourne ``False`` si on doit stopper (cap atteint)."""
        if depth >= max_depth or len(folders) >= max_count:
            return False
        try:
            children = sorted(path.iterdir(), key=lambda x: x.name.lower())
        except (PermissionError, FileNotFoundError, OSError):
            return True
        for item in children:
            if len(folders) >= max_count:
                return False
            # Skip hidden folders — alignés avec la règle list handler.
            if item.name.startswith("."):
                continue
            # Skip symlinks qui sortent du user_dir (CWE-59).
            try:
                if item.is_symlink():
                    target = item.resolve()
                    if not target.is_relative_to(user_dir_resolved):
                        continue
            except (OSError, RuntimeError):
                continue
            try:
                if not item.is_dir():
                    continue
                rel_path = item.relative_to(user_dir)
            except (OSError, ValueError):
                continue
            folders.append({"name": item.name, "path": str(rel_path), "depth": depth})
            if not _recurse(item, depth + 1):
                return False
        return True

    _recurse(user_dir, 0)
    return folders


# ─── Validator class (API publique + tests) ───────────────────────


class FileValidator:
    """Validation centralisée fichier / chemin (API stable testée)."""

    @staticmethod
    def validate_extension(filename: str, allowed: set[str] | frozenset[str]) -> str | None:
        """Retourne un message d'erreur si l'extension n'est pas whitelistée."""
        ext = Path(filename).suffix.lower()
        if ext not in allowed:
            return f"Extension non autorisée: {ext}"
        return None

    @staticmethod
    def validate_size(size: int, max_size: int) -> str | None:
        """Retourne un message d'erreur si la taille dépasse ``max_size``."""
        if size > max_size:
            return (
                f"Fichier trop volumineux: {size / (1024 * 1024):.1f} MB "
                f"(max {max_size / (1024 * 1024):.0f} MB)"
            )
        return None

    @staticmethod
    def validate_path(path: Path, base_dir: Path) -> str | None:
        """Retourne un message d'erreur si path traversal détecté.

        Utilise ``Path.is_relative_to`` (Python 3.9+) pour couvrir le
        trou ``startswith`` : ``/data/user11`` n'est plus considéré
        comme appartenant à ``/data/user1``.
        """
        try:
            resolved = path.resolve()
            base_resolved = base_dir.resolve()
            if resolved != base_resolved and not resolved.is_relative_to(base_resolved):
                return "Chemin invalide: tentative de traversée de répertoire"
        except (ValueError, OSError, RuntimeError):
            return "Chemin invalide"
        return None


def validate_file_content(content: bytes, extension: str) -> bool:
    """Valide que le contenu correspond à l'extension (magic bytes ou UTF-8).

    Retourne ``False`` sur format inconnu (fail-closed) — l'extension
    doit appartenir à ``ALLOWED_EXTENSIONS`` côté caller.
    """
    text_extensions = {
        ".csv",
        ".txt",
        ".json",
        ".sql",
        ".yaml",
        ".yml",
        ".html",
        ".xml",
        ".md",
        ".log",
    }
    if extension in text_extensions:
        try:
            content[:1024].decode("utf-8")
            return True
        except (UnicodeDecodeError, AttributeError):
            return False
    for magic, exts in MAGIC_BYTES.items():
        if content.startswith(magic):
            return extension in exts
    return False


def _is_gzip_magic(data: bytes) -> bool:
    """True si ``data`` commence par les magic bytes gzip (``0x1f 0x8b``)."""
    return len(data) >= 2 and data[0] == 0x1F and data[1] == 0x8B


#: Décompression gzip robuste = SINGLE SOURCE OF TRUTH (``app.utils.gzip_safe``).
#: Ré-exportés ici pour les callers/tests historiques de ce module. Le client
#: compresse le ``.afz.json`` côté navigateur (``CompressionStream``) pour
#: réduire ~20× le payload réseau et passer sous ``client_max_body_size`` de
#: nginx ; le serveur décompresse de façon transparente avant validation +
#: stockage (re-gzip déterministe aval → ETag/quota inchangés).
#:
#: ``gunzip_first_member`` tolère les octets de queue (le fix du bug
#: « classeur compressé trop volumineux ou illisible » : ``gzip.GzipFile``
#: plantait sur le moindre octet résiduel APRÈS avoir tout décompressé) et
#: borne la RAM (anti zip-bomb), tout en distinguant trop-gros vs corrompu.
_gunzip_with_cap = gunzip_first_member  # alias rétrocompat (callers + tests)


# ─── Page HTML ───────────────────────────────────────────────────


class DatastorePageHandler(BaseHandler):
    """Page gestionnaire de fichiers / données."""

    @authenticated
    async def get(self) -> None:
        user = self.current_user
        async with get_session() as db:
            storage_mgr = StorageManager(db, DATASTORE_DIR)
            stats = await storage_mgr.get_storage_stats(user.id)
        self.render(
            "datastore.html",
            page_title="Mes données",
            stats=stats,
        )


# ─── API : Lister / Naviguer ────────────────────────────────────


class DatastoreListAPIHandler(BaseHandler):
    """GET /api/datastore?path=sous/dossier — liste le contenu."""

    @authenticated
    async def get(self) -> None:
        user = self.current_user
        user_dir = _user_dir(user.id)
        rel = self.get_argument("path", "")
        target = _safe_path(user_dir, rel) if rel else user_dir

        if target is None or not target.exists():
            return self.write_json({"success": False, "error": "Chemin introuvable"}, 404)

        if target.is_file():
            info = _file_info(target, user_dir)
            if info is None:
                return self.write_json({"success": False, "error": "Chemin introuvable"}, 404)
            return self.write_json({"success": True, "file": info})

        items: list[dict] = []
        try:
            children = sorted(target.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower()))
        except (FileNotFoundError, PermissionError, OSError):
            children = []
        for child in children:
            if child.name.startswith("."):
                continue
            info = _file_info(child, user_dir)
            if info is not None:
                items.append(info)

        # ``_dir_stats`` peut coûter cher (rglob + stat par fichier) : off-load.
        stats = await asyncio.to_thread(_dir_stats, user_dir)

        self.write_json(
            {
                "success": True,
                "path": str(target.relative_to(user_dir)) if target != user_dir else "",
                "items": items,
                "stats": stats,
            }
        )


# ─── API : Upload ───────────────────────────────────────────────


class DatastoreUploadAPIHandler(BaseHandler):
    """POST /api/datastore/upload — upload de fichiers (multipart)."""

    @require_role("admin", "user")
    async def post(self) -> None:
        user = self.current_user
        _check_rate_limit(_upload_limiter, user.id, *RATE_LIMIT_UPLOAD)

        user_dir = _user_dir(user.id)
        dest_rel = self.get_argument("path", "")
        dest_dir = _safe_path(user_dir, dest_rel) if dest_rel else user_dir

        if dest_dir is None:
            return self.write_json({"success": False, "error": "Chemin invalide"}, 400)
        dest_dir.mkdir(parents=True, exist_ok=True)

        files = self.request.files.get("files", [])
        if not files:
            return self.write_json({"success": False, "error": "Aucun fichier envoyé"}, 400)

        overwrite = self.get_argument("overwrite", "false").lower() in ("true", "1")

        # ──────────────────────────────────────────────────────────────
        # ⚠️ Pas de ``async with get_session() as db`` global ici.
        # Chaque opération BDD ouvre sa propre session via les helpers
        # retry-isolés (``_storage_register_deletion_with_retry``,
        # ``_storage_register_upload_with_retry``, et
        # ``_schedule_audit_fire_and_forget``). Sans isolation, un
        # ``database is locked`` sur le 1ᵉʳ flush mettait la session
        # commune en "rollback only" et bloquait tous les fichiers
        # suivants du batch — la session ne peut pas être rejouée
        # telle quelle après ``OperationalError``. Avec isolation,
        # chaque write est retentable indépendamment.
        # ──────────────────────────────────────────────────────────────
        async with get_session() as quota_db:
            storage_mgr_quota = StorageManager(quota_db, DATASTORE_DIR)
            total_size = sum(len(f["body"]) for f in files)
            can_upload, error_msg = await storage_mgr_quota.check_quota(user.id, total_size)
        if not can_upload:
            # ``error_code`` machine-readable : permet au client de
            # distinguer le quota app (JSON 413) de l'oversize passerelle
            # nginx (HTML 413) et d'afficher l'UX quota dédiée. Cf.
            # _saveToPathAsync (static/js/iris-grid.js).
            return self.write_json(
                {
                    "success": False,
                    "error": f"Quota dépassé : {error_msg}",
                    "error_code": "QUOTA_EXCEEDED",
                },
                413,
            )

        # Taille max PAR FICHIER = SSoT admin (/admin/performance), résolue au
        # runtime. Distincte (a) du quota cumulé vérifié ci-dessus et (b) du cap
        # de PARSING openpyxl/CSV de workbooks.py (protection RAM anti-DoS, qui
        # reste volontairement indépendante : un fichier stocké ≤ SSoT n'est pas
        # forcément parsable en RAM). Le fichier datastore est stocké tel quel,
        # pas chargé en mémoire → la limite d'upload suffit ici.
        from app.services.ai.config_service import get_max_upload_size_bytes

        max_upload_bytes = await get_max_upload_size_bytes()

        # NB : un .afz.json gzippé par le navigateur est stocké TEL QUEL (écriture
        # disque directe, aucune décompression en RAM ici — cf. plus bas). La
        # taille DISQUE reste bornée par ``max_upload_bytes`` (par fichier) + le
        # quota de stockage admin (cumul, vérifié avant la boucle). Plus aucun cap
        # de décompression à la sauvegarde.

        uploaded: list[dict] = []
        errors: list[str] = []
        for f in files:
            raw_name = Path(f["filename"]).name  # strip any path component
            name = _sanitize_user_filename(raw_name)
            if name is None:
                errors.append(f"{raw_name}: nom de fichier invalide")
                continue
            ext = Path(name).suffix.lower()

            ext_error = FileValidator.validate_extension(name, ALLOWED_EXTENSIONS)
            if ext_error:
                errors.append(f"{name}: {ext_error}")
                continue

            # Cap par fichier + quota (vérifié avant la boucle) portent sur la
            # taille SUR LE FIL = taille disque (les .afz.json sont stockés
            # gzippés). On les laisse sur ``f["body"]`` (compressé si le client
            # a gzippé) — cohérent avec la convention « hash & quota = disque ».
            size_error = FileValidator.validate_size(len(f["body"]), max_upload_bytes)
            if size_error:
                errors.append(f"{name}: {size_error}")
                continue

            # ── Sauvegarde = ÉCRITURE DISQUE DIRECTE (pas de décompression RAM) ──
            # Un .afz.json gzippé par le navigateur (``CompressionStream``) est
            # STOCKÉ TEL QUEL : on n'explose PAS son contenu en RAM pour le
            # valider/re-compresser. Raisons :
            #  • Sauvegarder un classeur est fondamentalement une écriture disque ;
            #    le fichier compressé est petit (~Mo). Décompresser tout en RAM
            #    (souvent des centaines de Mo) était un goulot artificiel qui
            #    bloquait des classeurs pourtant largement stockables sur disque.
            #  • L'autosave se déclenche souvent : re-décompresser à CHAQUE save
            #    brûlerait du CPU pour rien.
            # La taille DISQUE (octets compressés) reste bornée par le cap
            # d'upload (vérifié ci-dessus) ET le quota de stockage admin (vérifié
            # avant la boucle). La validité JSON est vérifiée à la LECTURE
            # (``decode_afz_bytes`` → échec gracieux si corrompu), pas ici — c'est
            # le fichier privé de l'utilisateur, jamais exécuté, juste re-parsé.
            body = f["body"]
            client_gzipped_afz = name.lower().endswith(".afz.json") and _is_gzip_magic(body)

            if not client_gzipped_afz and not validate_file_content(body, ext):
                # Validation de contenu pour TOUT le reste (uploads bruts : .json
                # non gzippé d'un vieux navigateur, .xlsx, .csv, images, pdf…).
                errors.append(
                    f"{name}: le contenu du fichier ne correspond pas à l'extension {ext}"
                )
                continue

            target = dest_dir / name
            # ETag / If-Match check (cross-tab conflict detection)
            # — Si le client envoie un header ``If-Match`` avec le
            # hash qu'il pense être la version courante du fichier,
            # on compare au hash réel sur disque. Mismatch → 412
            # Precondition Failed avec le hash actuel pour que le
            # frontend puisse proposer reload/force overwrite.
            #
            # Le header est OPTIONNEL : sans, on garde le comportement
            # legacy (overwrite si flag set, sinon auto-rename). Avec,
            # on enforce une garantie d'optimistic locking.
            if_match = self.request.headers.get("If-Match", "").strip()
            if if_match and target.exists():
                try:
                    current_bytes = await asyncio.to_thread(target.read_bytes)
                    current_hash = calculate_hash_from_bytes(current_bytes)
                except OSError:
                    current_hash = ""
                if current_hash and if_match != current_hash:
                    # Conflit détecté : un autre process/tab a écrit
                    # entre-temps. On NE renvoie PAS de body
                    # contenant le contenu serveur (RGPD : éviter
                    # leak cross-user dans les race) — le frontend
                    # doit faire un GET /download s'il veut le voir.
                    self.set_status(412)
                    self.write_json(
                        {
                            "success": False,
                            "error_code": "ETAG_MISMATCH",
                            "error": (
                                "Le fichier a été modifié ailleurs. "
                                "Rechargez ou forcez l'écrasement."
                            ),
                            "current_hash": current_hash,
                        }
                    )
                    return

            if target.exists() and overwrite:
                old_rel = str(target.relative_to(user_dir))
                # Session isolée + retry : la session shared d'avant
                # ce refactor laissait toute la boucle en erreur dès
                # qu'un fichier hit "database is locked". Avec
                # l'isolation, seul ce fichier est retardé ; les autres
                # poursuivent normalement.
                await _storage_register_deletion_with_retry(user.id, old_rel)
                await asyncio.to_thread(target.unlink)
            elif target.exists():
                stem = target.stem
                idx = 1
                while target.exists():
                    target = dest_dir / f"{stem}_{idx}{ext}"
                    idx += 1

            rel_path = str(target.relative_to(user_dir))

            # Détermination des octets à écrire sur disque — convention
            # "hash & quota = contenu DISQUE (compressé pour .afz.json)" :
            #   • Classeur gzippé par le navigateur → on écrit ses octets gzip
            #     TELS QUELS (aucune décompression / re-compression). C'est une
            #     pure écriture disque. Le hash est déterministe : ``CompressionStream``
            #     produit le MÊME gzip pour le même contenu (mtime fixe), donc
            #     re-sauver un classeur identique redonne le même hash → If-Match
            #     stable cross-onglet.
            #   • .afz.json BRUT (vieux navigateur sans CompressionStream) → on
            #     gzip côté serveur pour l'efficacité disque (offload thread :
            #     CPU-bound, ne doit pas geler la boucle Tornado).
            #   • Autre format (xlsx/csv/image/pdf) → écrit tel quel.
            # Rétrocompat lecture : ``reader.decode_afz_bytes`` détecte le gzip et
            # décompresse à la volée ; les anciens .afz.json en clair marchent.
            body_to_write = body
            target_name_lower = target.name.lower()
            if target_name_lower.endswith(".afz.json") and not client_gzipped_afz:
                import gzip as _gzip

                body_to_write = await asyncio.to_thread(
                    lambda: _gzip.compress(body, compresslevel=6, mtime=0)
                )
                logger.debug(
                    "Gzip serveur .afz.json brut %s : %d → %d bytes (%.1f%%)",
                    target.name,
                    len(body),
                    len(body_to_write),
                    100.0 * len(body_to_write) / max(len(body), 1),
                )
            file_size = len(body_to_write)
            file_hash = calculate_hash_from_bytes(body_to_write)

            # Atomic write : écrit dans un <target>.<token>.tmp UNIQUE puis
            # os.replace() vers <target>. Si crash backend en plein write, le
            # fichier original (s'il existe) reste intact. ``replace`` est
            # atomique sur la même partition (mv POSIX/Windows). Le token unique
            # (T9) évite la collision .tmp entre 2 uploads concurrents du même path.
            tmp_target = _unique_tmp_path(target)
            try:
                await asyncio.to_thread(tmp_target.write_bytes, body_to_write)
                await asyncio.to_thread(os.replace, str(tmp_target), str(target))
            except Exception:
                # Cleanup le .tmp orphelin si l'opération a planté
                # entre write et replace (best-effort, pas bloquant).
                try:
                    if tmp_target.exists():
                        await asyncio.to_thread(tmp_target.unlink)
                except OSError:
                    pass
                raise

            # Session isolée + retry pour le INSERT FileMetadata +
            # UPDATE user_storage. Le fichier est DÉJÀ sur disque
            # (atomic replace) — si le retry final échoue, le fichier
            # devient orphelin de la BDD (rattrapable via
            # ``sync_user_storage``). Préférable à propager une 500
            # qui laisserait l'utilisateur penser que l'upload a
            # échoué alors qu'il est sur disque.
            metadata = await _storage_register_upload_with_retry(
                user.id,
                target,
                rel_path,
                file_size=file_size,
                file_hash=file_hash,
            )

            # Audit log fire-and-forget (retry interne) : la ligne
            # d'audit ne doit pas faire échouer la réponse user si
            # SQLite est sous contention. Cf.
            # ``_schedule_audit_fire_and_forget`` pour le pattern.
            _schedule_audit_fire_and_forget(
                action=AuditAction.FILE_UPLOAD,
                user_id=user.id,
                entity_id=metadata.id,
                details={"filename": rel_path, "size": file_size},
                ip_address=self.request.remote_ip,
                op_label="upload",
            )

            info = _file_info(target, user_dir)
            if info is not None:
                # Inclure le ``file_hash`` dans la response pour que
                # le frontend puisse l'utiliser comme ETag et envoyer
                # ``If-Match`` au prochain save (détection conflit
                # cross-tab). Pas de risque de leak : c'est un SHA-256
                # du contenu, pas le contenu lui-même.
                info["file_hash"] = file_hash
                uploaded.append(info)
            logger.info(
                "Fichier uploadé",
                extra={
                    "user_id": user.id,
                    "path": rel_path,
                    "size_bytes": file_size,
                },
            )

        # **Anti-perte silencieuse (C1)** : ``success`` reflète qu'AU MOINS un
        # fichier a réellement été persisté. Un batch où TOUT échoue
        # (``uploaded == []`` mais ``errors`` non vide) doit renvoyer
        # ``success:false`` — sinon le client (save classeur) marquerait le
        # fichier "sauvé" alors que rien n'a été écrit. Status 200 conservé :
        # la requête a été traitée (le détail est dans ``errors``), c'est un
        # échec applicatif désambiguïsé par ``success`` + ``errors``.
        self.write_json(
            {
                "success": bool(uploaded),
                "uploaded": uploaded,
                "errors": errors,
                "message": f"{len(uploaded)} fichier(s) importé(s)"
                + (f", {len(errors)} erreur(s)" if errors else ""),
            }
        )


# ─── Écriture atomique : chemin .tmp unique par requête ─────────


def _unique_tmp_path(target: Path) -> Path:
    """Chemin ``.tmp`` UNIQUE par requête pour l'écriture atomique (T9).

    Un ``.tmp`` DÉTERMINISTE (``<target>.tmp``) entre en collision quand deux
    requêtes écrivent le MÊME ``target`` en parallèle (multi-onglets, autosave
    keepalive + save normal, double-clic) : les ``write_bytes(.tmp)`` /
    ``os.replace`` tournent dans le thread-pool et s'entrelacent → R1 peut
    ``os.replace`` le contenu écrit par R2, puis R2 hit ``FileNotFoundError`` →
    le disque contient R2 mais R1 reçoit un 200 + un hash qui NE correspond PAS
    au disque (silent-wrong-data + 412 fantôme au prochain save). Un token
    aléatoire isole chaque écriture. ``cleanup_orphan_tmp_files`` balaie toujours
    les ``*.tmp`` orphelins (le nom finit par ``.tmp``).
    """
    return target.with_name(f"{target.name}.{uuid.uuid4().hex}.tmp")


# ─── Cleanup orphelins : .tmp non remplacés ─────────────────────


def cleanup_orphan_tmp_files(datastore_root: Path, max_age_seconds: int = 3600) -> int:
    """Sweep des fichiers ``.tmp`` orphelins dans le datastore.

    Le pattern atomic write (write_bytes(.tmp) puis os.replace) peut
    laisser un ``.tmp`` orphelin si le process Python crashe (SIGKILL,
    OOM, kernel panic) entre les deux étapes. Sans cleanup, ces
    orphelins polluent le disque progressivement et désynchronisent
    le quota DB (qui ne les compte pas car ``register_upload`` n'est
    appelé qu'après succès complet).

    Cette fonction parcourt récursivement ``datastore_root`` et
    supprime tous les fichiers ``*.tmp`` plus vieux que
    ``max_age_seconds``. Le seuil d'âge évite de tuer un ``.tmp`` en
    cours d'écriture par une requête concurrente (rare mais possible
    si on appelle cette fonction pendant que le serveur tourne).

    À appeler au boot du serveur (avant l'event loop) ou via un job
    périodique. Best-effort : log les erreurs OSError mais ne lève
    jamais — un cleanup raté ne doit pas empêcher le démarrage.

    Args:
        datastore_root: Path racine du datastore (typiquement
            ``DATASTORE_DIR`` global).
        max_age_seconds: Age min en secondes avant suppression
            (défaut 1h — cap les writes en cours).

    Returns:
        Nombre de fichiers nettoyés.
    """
    if not datastore_root.exists() or not datastore_root.is_dir():
        return 0
    import time

    cutoff = time.time() - max_age_seconds
    removed = 0
    try:
        for tmp in datastore_root.rglob("*.tmp"):
            try:
                if tmp.is_file() and tmp.stat().st_mtime < cutoff:
                    tmp.unlink()
                    removed += 1
            except OSError as exc:
                logger.warning(
                    "cleanup_orphan_tmp_files: échec sur %s — %s",
                    tmp,
                    exc,
                )
    except OSError as exc:
        logger.warning(
            "cleanup_orphan_tmp_files: rglob a échoué — %s",
            exc,
        )
    if removed:
        logger.info(
            "cleanup_orphan_tmp_files: %d .tmp orphelin(s) supprimé(s)",
            removed,
        )
    return removed


# ─── API : Créer dossier ────────────────────────────────────────


class DatastoreMkdirAPIHandler(BaseHandler):
    """POST /api/datastore/mkdir — crée un sous-dossier."""

    @require_role("admin", "user")
    async def post(self) -> None:
        user = self.current_user
        user_dir = _user_dir(user.id)
        data = self.get_json_body()
        folder_name = _sanitize_user_filename(data.get("name", ""))
        parent_rel = data.get("path", "")

        if folder_name is None:
            return self.write_json({"success": False, "error": "Nom de dossier invalide"}, 400)

        parent = _safe_path(user_dir, parent_rel) if parent_rel else user_dir
        if parent is None:
            return self.write_json({"success": False, "error": "Chemin parent invalide"}, 400)

        new_dir = parent / folder_name
        if new_dir.exists():
            return self.write_json({"success": False, "error": "Ce dossier existe déjà"}, 409)

        try:
            await asyncio.to_thread(new_dir.mkdir, parents=True, exist_ok=False)
        except FileExistsError:
            return self.write_json({"success": False, "error": "Ce dossier existe déjà"}, 409)
        except OSError as exc:
            logger.warning(
                "Échec mkdir",
                extra={"user_id": user.id, "folder_name": folder_name, "error": str(exc)},
            )
            return self.write_json(
                {"success": False, "error": "Impossible de créer le dossier."}, 500
            )
        self.write_json({"success": True, "message": f"Dossier '{folder_name}' créé"})


# ─── API : Renommer ─────────────────────────────────────────────


class DatastoreRenameAPIHandler(BaseHandler):
    """POST /api/datastore/rename — renomme fichier ou dossier."""

    @require_role("admin", "user")
    async def post(self) -> None:
        user = self.current_user
        _check_rate_limit(_rename_limiter, user.id, *RATE_LIMIT_RENAME)
        user_dir = _user_dir(user.id)
        data = self.get_json_body()
        old_path = data.get("path", "")

        if not old_path:
            return self.write_json({"success": False, "error": "Chemin source manquant"}, 400)

        new_name, reason = _sanitize_user_filename_with_reason(data.get("new_name", ""))
        if new_name is None:
            return self.write_json({"success": False, "error": reason}, 400)

        source = _safe_path(user_dir, old_path)
        if source is None or not source.exists():
            return self.write_json({"success": False, "error": "Fichier introuvable"}, 404)

        # No-op : renommer vers le même nom ne fait rien. On compare aussi
        # sur le path résolu pour gérer les FS case-insensitive (NTFS/HFS+/
        # APFS) où ``File.csv`` et ``file.csv`` pointent vers le même
        # inode — sans ça, un user sur Mac/Windows ne peut pas changer
        # la casse d'un nom de fichier (dest.exists() vrai → 409 erroné).
        if source.name == new_name:
            return self.write_json({"success": True, "message": "Aucun changement"})
        dest_preview = source.parent / new_name
        try:
            same_inode = source.resolve() == dest_preview.resolve()
        except (OSError, RuntimeError):
            same_inode = False
        if same_inode and source.name.lower() == new_name.lower():
            # Changement de casse pur (case-insensitive FS) — autoriser
            # via rename atomique (bypass le check dest.exists qui mentirait).
            try:
                await asyncio.to_thread(source.rename, dest_preview)
            except OSError as exc:
                logger.error(
                    "Rename case-change failed",
                    extra={
                        "user_id": user.id,
                        "old": source.name,
                        "new": new_name,
                        "error": str(exc),
                    },
                )
                return self.write_json({"success": False, "error": "Renommage impossible"}, 500)
            old_rel = _normalize_rel_path(source, user_dir)
            new_rel = _normalize_rel_path(dest_preview, user_dir)
            _schedule_audit_fire_and_forget(
                action=AuditAction.FILE_RENAME,
                user_id=user.id,
                details={
                    "old_path": old_rel,
                    "new_path": new_rel,
                    "new_name": new_name,
                    "was_dir": source.is_dir(),
                    "case_change_only": True,
                },
                ip_address=self.request.remote_ip,
                op_label=f"rename-case {old_rel} -> {new_rel}",
            )
            return self.write_json({"success": True, "message": f"Renommé en '{new_name}'"})

        # Extension whitelist sur les fichiers (R-6 — bypass possible
        # auparavant : un user pouvait uploader ``data.txt`` puis renommer
        # en ``data.html``/``.exe``/``.sh`` pour contourner les ALLOWED_
        # EXTENSIONS appliquées uniquement à l'upload).
        if source.is_file():
            ext_error = FileValidator.validate_extension(new_name, ALLOWED_EXTENSIONS)
            if ext_error:
                return self.write_json({"success": False, "error": ext_error}, 400)

        dest = source.parent / new_name
        if dest.exists():
            return self.write_json(
                {"success": False, "error": "Un élément avec ce nom existe déjà"}, 409
            )

        old_rel = _normalize_rel_path(source, user_dir)
        new_rel = _normalize_rel_path(dest, user_dir)
        now_ts = int(clock.timestamp())
        was_dir = source.is_dir()
        affected_files_count = 0

        # FS first via helper atomique fail-if-exists (R-1 race silent
        # overwrite). On bascule le rename DB après pour ne mettre à jour
        # les FileMetadata que si le FS a réussi.
        try:
            await asyncio.to_thread(_atomic_move_failing_if_exists, source, dest)
        except FileExistsError:
            return self.write_json(
                {"success": False, "error": "Un élément avec ce nom existe déjà"}, 409
            )
        except OSError as exc:
            logger.error(
                "Rename FS failed",
                extra={"user_id": user.id, "old": old_rel, "new": new_rel, "error": str(exc)},
            )
            return self.write_json({"success": False, "error": "Renommage impossible"}, 500)

        async with get_session() as db:
            if not was_dir:
                result = await db.execute(
                    select(FileMetadata)
                    .where(FileMetadata.user_id == user.id)
                    .where(FileMetadata.file_path == old_rel)
                )
                metadata = result.scalar_one_or_none()
                if metadata:
                    metadata.file_path = new_rel
                    metadata.filename = new_name
                    metadata.extension = Path(new_name).suffix.lower()
                    metadata.updated_at = now_ts
                    affected_files_count = 1
            else:
                # Dossier : update tous les FileMetadata enfants.
                escaped_rel = old_rel.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
                result = await db.execute(
                    select(FileMetadata).where(
                        FileMetadata.user_id == user.id,
                        FileMetadata.file_path.like(f"{escaped_rel}/%", escape="\\"),
                    )
                )
                prefix_old = f"{old_rel}/"
                prefix_new = f"{new_rel}/"
                for file_meta in result.scalars().all():
                    if file_meta.file_path.startswith(prefix_old):
                        file_meta.file_path = prefix_new + file_meta.file_path[len(prefix_old) :]
                        file_meta.updated_at = now_ts
                        affected_files_count += 1

        # Audit fire-and-forget — si le commit DB échoue (database is locked
        # ou autre), l'audit log reste persistant via retry séparé.
        _schedule_audit_fire_and_forget(
            action=AuditAction.FILE_RENAME,
            user_id=user.id,
            details={
                "old_path": old_rel,
                "new_path": new_rel,
                "new_name": new_name,
                "was_dir": was_dir,
                "affected_files_count": affected_files_count,
            },
            ip_address=self.request.remote_ip,
            op_label=f"rename {old_rel} -> {new_rel}",
        )

        self.write_json({"success": True, "message": f"Renommé en '{new_name}'"})


# ─── API : Supprimer ────────────────────────────────────────────


class DatastoreDeleteAPIHandler(BaseHandler):
    """POST /api/datastore/delete — supprime fichier(s) ou dossier(s)."""

    @require_role("admin", "user")
    async def post(self) -> None:
        user = self.current_user
        _check_rate_limit(_delete_limiter, user.id, *RATE_LIMIT_DELETE)
        user_dir = _user_dir(user.id)
        data = self.get_json_body()
        paths_raw = data.get("paths", [])

        # D-02 : validation stricte de la forme `paths`.
        if not isinstance(paths_raw, list):
            return self.write_json(
                {"success": False, "error": "Le champ 'paths' doit être une liste"}, 400
            )
        # Filter non-str et dédup (préserve l'ordre).
        seen: set[str] = set()
        paths: list[str] = []
        for item in paths_raw:
            if not isinstance(item, str) or not item:
                continue
            if item in seen:
                continue
            seen.add(item)
            paths.append(item)
        if not paths:
            return self.write_json({"success": False, "error": "Aucun chemin spécifié"}, 400)
        # D-04 : cap batch pour éviter DoS quota query.
        if len(paths) > MAX_DELETE_BATCH:
            return self.write_json(
                {
                    "success": False,
                    "error": f"Trop d'éléments ({len(paths)}, max {MAX_DELETE_BATCH})",
                },
                400,
            )

        deleted = 0
        errors: list[str] = []
        audit_payloads: list[dict[str, object]] = []

        # D-03 : FS first puis DB après — si DB échoue après suppression FS,
        # on log un warning quota_drift mais l'état FS reste cohérent.
        # D-2 : une session par item (pas une session unique sur la boucle)
        # pour qu'une erreur SQLAlchemy non-locked sur un item n'emporte
        # pas en rollback les 999 register_deletion précédents.
        for rel in paths:
            target = _safe_path(user_dir, rel)
            if target is None or not target.exists():
                errors.append(f"{rel}: introuvable")
                continue
            if target == user_dir:
                errors.append("Impossible de supprimer le dossier racine")
                continue

            was_dir = target.is_dir()
            # D-08 : normaliser le path AVANT de l'utiliser pour DB lookup
            # ou audit (le `rel` brut peut avoir trailing slash, casse
            # divergente, etc. — quota drift garanti sinon).
            norm_rel = _normalize_rel_path(target, user_dir)

            affected_files_count = 0

            # FS first.
            try:
                if was_dir:
                    await asyncio.to_thread(shutil.rmtree, target)
                else:
                    await asyncio.to_thread(target.unlink)
            except OSError as exc:
                logger.error(
                    "Suppression FS échouée",
                    extra={"user_id": user.id, "path": norm_rel, "error": str(exc)},
                )
                errors.append(f"{rel}: erreur suppression")
                continue

            # DB après, dans sa propre session — isolation des rollbacks.
            try:
                async with get_session() as db:
                    storage_mgr = StorageManager(db, DATASTORE_DIR)
                    if was_dir:
                        dir_prefix = f"{norm_rel}/"
                        escaped_prefix = (
                            norm_rel.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
                        )
                        result = await db.execute(
                            select(FileMetadata).where(
                                FileMetadata.user_id == user.id,
                                FileMetadata.file_path.like(f"{escaped_prefix}/%", escape="\\"),
                            )
                        )
                        for child_meta in result.scalars().all():
                            if child_meta.file_path.startswith(dir_prefix):
                                # D-1 : capturer le retour de register_deletion
                                # pour que affected_files_count reflète la
                                # réalité DB (et pas juste le nombre d'itérations).
                                ok = await storage_mgr.register_deletion(
                                    user.id, child_meta.file_path
                                )
                                if ok:
                                    affected_files_count += 1
                    else:
                        ok = await storage_mgr.register_deletion(user.id, norm_rel)
                        if ok:
                            affected_files_count = 1
            except SQLAlchemyError as exc:
                logger.warning(
                    "Quota drift après suppression FS — DB unreachable",
                    extra={"user_id": user.id, "path": norm_rel, "error": str(exc)},
                )

            audit_payloads.append(
                {
                    "path": norm_rel,
                    "was_dir": was_dir,
                    "affected_files_count": affected_files_count,
                }
            )
            deleted += 1

        # Audit en fire-and-forget pour ne pas perdre les lignes si la
        # session ci-dessus rollback (cf. D-05 : avant le commit final,
        # un échec emportait tous les audit logs du batch).
        for details in audit_payloads:
            _schedule_audit_fire_and_forget(
                action=AuditAction.FILE_DELETE,
                user_id=user.id,
                details=details,
                ip_address=self.request.remote_ip,
                op_label=f"delete {details['path']}",
            )

        # D-01 : `success` reflète la réalité — si rien n'a été supprimé,
        # on ne ment pas au client.
        all_ok = deleted == len(paths)
        partial = deleted > 0 and deleted < len(paths)
        success = deleted > 0
        if all_ok:
            message = f"{deleted} élément(s) supprimé(s)"
        elif partial:
            message = f"{deleted}/{len(paths)} élément(s) supprimé(s)"
        else:
            message = "Aucune suppression effectuée"
        self.write_json(
            {
                "success": success,
                "deleted": deleted,
                "errors": errors,
                "message": message,
            }
        )


# ─── API : Télécharger ──────────────────────────────────────────


_DOWNLOAD_AUDIT_TASKS: set[asyncio.Task[None]] = set()
"""Référence forte pour les tasks audit fire-and-forget — empêche le GC
de Python 3.12+ de les collecter avant complétion (cf. ``main.py:359``
``_ServerLifecycle._background_tasks`` pour le pattern documenté)."""


class DatastoreDownloadAPIHandler(BaseHandler):
    """GET /api/datastore/download?path=xxx — télécharge un fichier."""

    def _schedule_audit_file_download(
        self,
        *,
        user_id: int,
        rel: str,
        filename: str,
        gzip: bool,
        ip_address: str | None,
    ) -> None:
        """Schedule l'écriture de l'audit log en **fire-and-forget**.

        L'audit ne doit JAMAIS bloquer la réponse user. Sur un download
        de 4 GiB d'.afz.json, attendre 90 s pour un retry triple sur
        ``database is locked`` est inacceptable UX (cf. review adversariale
        2026-05-20). On schedule l'audit en background avec retry interne
        — l'user reçoit son fichier immédiatement.

        Les paramètres ``user_id`` / ``rel`` / ``filename`` / ``ip_address``
        sont capturés explicitement (et non lus depuis ``self`` dans la
        task) car ``self.request`` n'est plus garanti valide après la fin
        du handler (cf. cycle de vie Tornado).

        L'audit log est régulé en croissance par
        ``cleanup_db_retention_job`` (``app/services/cleanup/db_retention.py``)
        donc une ligne ratée n'a pas d'impact long terme.
        """
        details: dict[str, object] = {"path": rel, "filename": filename}
        if gzip:
            details["gzip"] = True
        op_name = f"audit_logs[FILE_DOWNLOAD user={user_id}]"

        async def _commit_audit() -> None:
            async with get_session() as db:
                db.add(
                    AuditLog.log_action(
                        action=AuditAction.FILE_DOWNLOAD,
                        user_id=user_id,
                        entity_type="file",
                        details=details,
                        ip_address=ip_address,
                    )
                )

        async def _audit_task() -> None:
            try:
                # ``max_attempts=5`` : budget cumulé ~3 s (vs 1 s pour
                # max=3). En 2026-05-22 les logs montraient des locks
                # tenus > 1 min — la mitigation par retry seule reste
                # insuffisante face à ça (la cause racine est ailleurs),
                # mais 5 tentatives capturent les contentions de
                # checkpoint WAL / sync programmatique qui durent
                # typiquement 1-2 s.
                await retry_on_locked(
                    _commit_audit,
                    max_attempts=5,
                    operation_name=op_name,
                )
            except OperationalError as exc:
                # database is locked persistant — la doctrine best-effort
                # accepte la perte de cette ligne d'audit (TTL en place,
                # pas d'invariant compliance court terme cassé).
                logger.warning(
                    "Audit log %s : abandon après retries (database is locked) : %s",
                    op_name,
                    exc,
                )
            # Toute autre exception (IntegrityError, ProgrammingError,
            # schema drift) est laissée se propager → asyncio remontera
            # un message d'erreur lors du GC de la task ET on perdra l'audit
            # — c'est volontaire : un bug structurel doit être visible
            # (alarme dans les logs), pas avalé silencieusement.

        task = asyncio.create_task(_audit_task())
        _DOWNLOAD_AUDIT_TASKS.add(task)
        task.add_done_callback(_DOWNLOAD_AUDIT_TASKS.discard)

    async def _serve_anonymized_afz(self, user: Any, target: Path, rel: str) -> None:
        """Sert un .afz.json dont les VALEURS de cellules sont anonymisées selon
        ``/data/privacy`` (export « valeurs anonymisées »).

        Seul le .afz.json (classeur ré-ouvrable) est concerné — un fichier
        binaire opaque (PDF, image…) ne peut pas être ré-anonymisé. Fail-closed
        (422) si un terme configuré ne peut être appliqué. Borné par le quota de
        stockage admin (SSoT) ; ce chemin re-matérialise le JSON (pic ~3×) donc
        l'admin dimensionne le quota vs la RAM du conteneur.

        Note : les champs ``sql`` des onglets sont conservés. Le fichier reflète
        des valeurs anonymisées (instantané destiné au partage) ; une éventuelle
        ré-exécution SQL dans Iris repasserait sur les vraies données — ce n'est
        pas un original ré-importable garantissant l'anonymat au re-run.
        """
        if not target.name.lower().endswith(".afz.json"):
            return self.write_json(
                {
                    "success": False,
                    "error": (
                        "Seuls les classeurs .afz.json peuvent être téléchargés "
                        "anonymisés. Pour les autres formats, exportez en anonymisé "
                        "depuis Iris ou un dashboard."
                    ),
                },
                422,
            )

        # Lecture du JSON via le SSoT ``decode_afz_bytes`` : gzip transparent +
        # tolérance aux octets de queue + borne de décompression = QUOTA admin
        # (plus aucun cap hardcodé). Un .afz.json ne doit pas OOM le serveur ; le
        # quota borne le pic RAM (ce chemin re-matérialise le JSON pour
        # anonymisation → l'admin dimensionne le quota en conséquence).
        from app.services.classeur.reader import decode_afz_bytes
        from app.services.storage_manager import get_storage_quota_bytes

        try:
            file_bytes = await asyncio.to_thread(target.read_bytes)
            quota = await get_storage_quota_bytes()
            data = await asyncio.to_thread(
                decode_afz_bytes,
                file_bytes,
                source=target.name,
                max_decompressed_bytes=quota,
            )
        except (OSError, json.JSONDecodeError, UnicodeDecodeError, ValueError) as exc:
            logger.warning(
                "datastore download anonymisé : lecture .afz.json échouée : %s",
                exc,
                extra={"user_id": user.id},
            )
            return self.write_json(
                {"success": False, "error": "Fichier corrompu ou illisible"}, 500
            )

        if not isinstance(data, dict):
            return self.write_json({"success": False, "error": "Format .afz.json invalide"}, 400)

        term_count = 0
        tabs = data.get("tabs")
        if isinstance(tabs, list):
            from app.services.anonymization.export_filter import (
                anonymize_tabs_for_export_meta,
            )

            try:
                _anon = await anonymize_tabs_for_export_meta(user.id, tabs)
            except RuntimeError:
                logger.warning(
                    "datastore download anonymisé : fail-closed",
                    extra={"user_id": user.id},
                    exc_info=True,
                )
                return self.write_json(
                    {
                        "success": False,
                        "error": (
                            "Anonymisation impossible : un terme configuré sur "
                            "/data/privacy n'a pas pu être appliqué. Corrigez le "
                            "conflit puis réessayez."
                        ),
                    },
                    422,
                )
            data["tabs"] = _anon["tabs"]
            term_count = _anon["term_count"]

        # Sécurité (review adversariale 2026-06-01) : retirer la mémoire copilot
        # (chat Iris) de l'export anonymisé. Elle peut citer des valeurs réelles
        # dans la conversation, qu'on ne peut pas anonymiser de façon fiable
        # (texte libre). Un .afz.json anonymisé destiné au PARTAGE n'en a pas
        # besoin — on la strip plutôt que de risquer une fuite partielle.
        data.pop("copilot_memory", None)

        out_bytes = json.dumps(data, ensure_ascii=False).encode("utf-8")

        base = target.name
        anon_name = (
            base[: -len(".afz.json")] + "_anonymise.afz.json"
            if base.lower().endswith(".afz.json")
            else base + "_anonymise"
        )

        self.set_header("Content-Type", "application/json; charset=utf-8")
        self.set_header("Content-Disposition", _content_disposition(anon_name))
        self.set_header("Content-Length", len(out_bytes))
        self.set_header("Cache-Control", "no-store")
        # F12 : nombre de termes appliqués → le client avertit si 0 (fichier
        # identique au clair, anti fausse-impression de sécurité).
        self.set_header("X-Anon-Term-Count", str(term_count))
        self._schedule_audit_file_download(
            user_id=user.id,
            rel=rel,
            filename=anon_name,
            gzip=False,
            ip_address=self.request.remote_ip,
        )
        self.write(out_bytes)

    @authenticated
    async def get(self) -> None:
        user = self.current_user
        _check_rate_limit(_download_limiter, user.id, *RATE_LIMIT_DOWNLOAD)
        user_dir = _user_dir(user.id)
        rel = self.get_argument("path", "")

        if not rel:
            return self.write_json({"success": False, "error": "Chemin requis"}, 400)

        target = _safe_path(user_dir, rel)
        if target is None or not target.exists() or not target.is_file():
            return self.write_json({"success": False, "error": "Fichier introuvable"}, 404)

        # Mode anonymisé (« valeurs anonymisées ») : seul un classeur .afz.json
        # (ré-ouvrable) peut être ré-anonymisé à la volée selon /data/privacy.
        # Branche self-contained — ne touche pas les chemins de stream existants.
        if self.get_argument("anonymize", "") in ("1", "true", "True", "on"):
            return await self._serve_anonymized_afz(user, target, rel)

        # Gzip transparent : si c'est un .afz.json stocké compressé, on
        # décompresse à la volée pour servir un JSON lisible au client
        # (peu importe son tooling — l'extension reste .afz.json donc le
        # frontend Komptia s'attend à du JSON, et les téléchargements
        # manuels ouvrent un texte). Pour les autres extensions, on stream
        # le fichier tel quel.
        is_gzipped_afz = False
        if target.name.lower().endswith(".afz.json"):
            try:
                with open(target, "rb") as _f:
                    is_gzipped_afz = _f.read(2) == b"\x1f\x8b"
            except OSError:
                pass

        # DL-3 — ETag via FileMetadata.file_hash si disponible (calculé à
        # l'upload + persisté). Évite le double-read disque (read_bytes
        # pour hash + stream_file_to_handler pour body) sur les fichiers
        # < 50 MB. Sur les fichiers plus gros le hash BDD est la SEULE
        # façon d'obtenir un ETag (le re-read serait OOM).
        stored_hash: Optional[str] = None
        try:
            async with get_session() as db:
                result = await db.execute(
                    select(FileMetadata.file_hash)
                    .where(FileMetadata.user_id == user.id)
                    .where(FileMetadata.file_path == _normalize_rel_path(target, user_dir))
                )
                row = result.first()
                if row and row[0]:
                    stored_hash = row[0]
        except SQLAlchemyError as exc:
            logger.debug(
                "ETag lookup failed (best-effort)",
                extra={"user_id": user.id, "error": str(exc)},
            )

        if is_gzipped_afz:
            # ── Ouverture/téléchargement = LECTURE DISQUE DIRECTE ────────────
            # Symétrique au fix de sauvegarde : on NE décompresse PAS le
            # .afz.json en RAM côté serveur. On streame les octets gzip STOCKÉS
            # TELS QUELS avec ``Content-Encoding: gzip`` ; le navigateur les
            # décompresse NATIVEMENT (le ``fetch`` de « ouvrir le classeur » via
            # ``r.json()`` ET le téléchargement disque décodent l'encoding de
            # façon transparente — résultat identique à l'ancien comportement
            # « JSON décompressé », mais sans aucun pic mémoire serveur).
            #
            # Conséquences :
            #  • Plus de buffer de 600 Mio en RAM avant le flush → plus d'OOM ni
            #    de 413 quota à l'ouverture : un classeur s'ouvre quelle que soit
            #    sa taille décompressée (tant qu'il tient côté navigateur).
            #  • Octets transmis = taille COMPRESSÉE (~Mo) → réseau léger.
            #  • Le navigateur reste seul juge de sa propre RAM au parse JSON.
            self.set_header("Content-Type", "application/json; charset=utf-8")
            self.set_header("Content-Encoding", "gzip")
            self.set_header("Content-Disposition", _content_disposition(target.name))
            # Content-Length = octets ENVOYÉS (compressés), pas la taille décompressée.
            self.set_header("Content-Length", target.stat().st_size)
            if stored_hash:
                self.set_header("ETag", f'"{stored_hash}"')
            self._schedule_audit_file_download(
                user_id=user.id,
                rel=rel,
                filename=target.name,
                gzip=True,
                ip_address=self.request.remote_ip,
            )
            # Streaming des octets gzip bruts (fichier déjà compressé sur disque).
            await stream_file_to_handler(self, target, FILE_CHUNK_BYTES)
            return

        mime = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
        self.set_header("Content-Type", mime)
        self.set_header("Content-Disposition", _content_disposition(target.name))
        self.set_header("Content-Length", target.stat().st_size)

        # DL-3 — Si hash en BDD on l'utilise direct, sinon on skip (le
        # double-read est évité). Le hash sera calculé à l'upload suivant.
        if stored_hash:
            self.set_header("ETag", f'"{stored_hash}"')

        self._schedule_audit_file_download(
            user_id=user.id,
            rel=rel,
            filename=target.name,
            gzip=False,
            ip_address=self.request.remote_ip,
        )

        await stream_file_to_handler(self, target, FILE_CHUNK_BYTES)


# ─── API : Aperçu / Preview ─────────────────────────────────────


class DatastorePreviewAPIHandler(BaseHandler):
    """GET /api/datastore/preview?path=xxx — aperçu du contenu."""

    def _preview_text_file(self, file_path: Path, ext: str) -> dict:
        """Aperçu d'un fichier texte (CSV/JSON/text/SQL/XML/log/md).

        Cap du fichier lu en RAM à ``MAX_PREVIEW_FILE_BYTES`` (défense
        contre un upload 50 Mo suivi d'un hammer /preview → OOM).
        """
        try:
            try:
                file_size = file_path.stat().st_size
            except (FileNotFoundError, OSError):
                return {"type": "error", "content": "Fichier introuvable."}
            if file_size > MAX_PREVIEW_FILE_BYTES:
                return {
                    "type": "info",
                    "content": (
                        "Fichier trop volumineux pour la prévisualisation "
                        f"(max {MAX_PREVIEW_FILE_BYTES // (1024 * 1024)} Mo)."
                    ),
                }
            # Garde gzip : un fichier compressé lu en texte afficherait du binaire
            # (« \x1f\x8b… »). Les classeurs .afz.json sont normalement interceptés
            # en amont (``_preview_workbook``) ; ce filet couvre tout autre fichier
            # gzippé qui arriverait ici plutôt que d'afficher de la bouillie.
            with open(file_path, "rb") as _fh:
                if _is_gzip_magic(_fh.read(2)):
                    return {
                        "type": "info",
                        "content": "Fichier compressé (binaire) — aperçu non disponible.",
                    }
            # ``errors="replace"`` : on préfère afficher des ``?`` qu'un 500.
            text = file_path.read_text(encoding="utf-8", errors="replace")
            # A8-F2 — compter les lignes du fichier COMPLET avant de tronquer le
            # texte d'AFFICHAGE. Avant, ``total_lines`` (affiché « N lignes au
            # total » par le frontend) était calculé sur le texte déjà coupé à
            # MAX_PREVIEW_TEXT_BYTES (100 Ko) → un CSV jusqu'à 5 Mo / 40 000 lignes
            # affichait « ~2000 lignes au total » (donnée fausse silencieuse).
            # Le fichier est déjà borné à MAX_PREVIEW_FILE_BYTES (5 Mo) en amont,
            # donc ce count reste plafonné en RAM. ``+1`` si pas de newline final
            # (la dernière ligne sans ``\n`` compte quand même comme une ligne
            # physique — cohérent avec ``ws.max_row`` du preview Excel).
            full_line_count = text.count("\n") + (0 if (not text or text.endswith("\n")) else 1)
            if len(text) > MAX_PREVIEW_TEXT_BYTES:
                text = text[:MAX_PREVIEW_TEXT_BYTES] + "\n\n… (tronqué)"
            preview: dict = {"type": "text", "content": text}

            if ext == ".csv":
                # Séparateur détecté via le SSoT ``csv_loader._detect_separator``
                # (csv.Sniffer sur ``;,\t|`` + fallback ``;`` FR). Avant,
                # ``csv.reader`` prenait ``,`` par défaut → un CSV FR à ``;``
                # (export Excel/Sage standard) s'affichait en 1 SEULE colonne et
                # splittait sur la virgule décimale (« 1 234,56 ») = tableau
                # d'aperçu structurellement FAUX, silencieusement. Le loader
                # external-sheets sniffait déjà ; le preview ré-implémentait
                # naïvement — on factorise sur le même helper.
                from app.services.external_sheets.csv_loader import _detect_separator

                delimiter = _detect_separator(text)
                reader = csv.reader(io.StringIO(text), delimiter=delimiter)
                rows: list[list[str]] = []
                for i, row in enumerate(reader):
                    rows.append(row)
                    if i >= 50:
                        break
                if rows:
                    preview = {
                        "type": "csv",
                        "headers": rows[0] if rows else [],
                        "rows": rows[1:51],
                        "total_lines": full_line_count,  # A8-F2 : fichier complet
                    }
            elif ext == ".json":
                try:
                    obj = json.loads(text)
                    rendered = json.dumps(obj, indent=2, ensure_ascii=False)
                    # A8-F2 — signaler la troncature d'affichage (le pretty-print
                    # peut dépasser 50 Ko même pour un fichier < 100 Ko). Avant,
                    # la coupe ``[:50_000]`` était silencieuse → JSON d'apparence
                    # complète mais coupé. Le frontend affiche ``content`` dans un
                    # <pre> (pas de JSON.parse) → marqueur texte sûr.
                    if len(rendered) > 50_000:
                        rendered = rendered[:50_000] + "\n\n… (tronqué)"
                    preview = {
                        "type": "json",
                        "content": rendered,
                    }
                except json.JSONDecodeError:
                    pass

            return preview
        except (json.JSONDecodeError, csv.Error, ValueError) as e:
            logger.error("Erreur parsing texte: %s", e, exc_info=True)
            return {"type": "error", "content": "Impossible de prévisualiser ce fichier."}

    def _preview_excel(self, file_path: Path) -> dict:
        """Aperçu d'un fichier Excel (50 premières lignes, 200 cols max).

        Fix CRIT : ``"" if c is None else str(c)`` préserve les zéros et
        les booléens. Le précédent ``str(c) if c else ""`` convertissait
        ``0``, ``False``, ``""`` tous en chaîne vide — perte silencieuse
        de données cruciales pour un tableur financier.

        Borne le nombre de lignes itérées pour limiter l'exposition aux
        zip bombs xlsx (un xlsx de 100 Ko peut décompresser à 500 Mo).
        """
        try:
            try:
                import openpyxl  # type: ignore[import-untyped]
            except ImportError:
                return {
                    "type": "info",
                    "content": "Module openpyxl requis pour prévisualiser Excel",
                }

            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            ws = wb.active
            headers: list[str] = []
            rows: list[list[str]] = []

            def _cell(c: Any) -> str:
                # Préserve ``0``, ``False``, ``""`` et dates. ``None`` uniquement → "".
                return "" if c is None else str(c)

            # ``max_col`` borne la largeur itérée — anti zip-bomb sur un
            # xlsx avec 16k colonnes fantômes.
            for i, row in enumerate(ws.iter_rows(values_only=True, max_col=MAX_PREVIEW_XLSX_COLS)):
                if i == 0:
                    headers = [_cell(c) for c in row]
                elif i <= MAX_PREVIEW_XLSX_ROWS:
                    rows.append([_cell(c) for c in row])
                else:
                    break
            total_rows = ws.max_row or 0
            wb.close()
            return {
                "type": "csv",
                "headers": headers,
                "rows": rows,
                "total_lines": total_rows,
            }
        except (OSError, ValueError) as e:
            logger.error("Erreur parsing Excel: %s", e, exc_info=True)
            return {"type": "error", "content": "Impossible de prévisualiser ce fichier Excel."}

    def _preview_workbook(self, file_path: Path) -> dict:
        """Aperçu d'un classeur ``.afz.json`` (gzip transparent via SSoT).

        Les ``.afz.json`` sont stockés gzippés sur disque. Avant, un simple clic
        les routait vers ``_preview_text_file`` (extension ``.json``) →
        ``read_text`` sur des octets gzip → bouillie binaire (« \\x1f\\x8b… »)
        affichée dans le panneau d'aperçu. Ici on décode via ``decode_afz_bytes``
        (SSoT, gzip-aware + tolérant) et on renvoie un résumé d'onglets, en
        invitant au double-clic (``openWorkbook``) pour l'ouverture complète dans
        la grille. Le type ``info`` est déjà rendu par le frontend (aucune
        modification de renderer nécessaire).
        """
        from app.services.classeur.reader import decode_afz_bytes

        try:
            raw = file_path.read_bytes()
        except OSError:
            return {"type": "error", "content": "Fichier introuvable."}
        try:
            obj = decode_afz_bytes(raw, source=str(file_path))
        except (json.JSONDecodeError, ValueError) as exc:
            logger.warning("Aperçu classeur .afz.json échoué : %s", exc)
            return {"type": "error", "content": "Classeur illisible ou corrompu."}

        tabs = obj.get("tabs") if isinstance(obj, dict) else None
        if not isinstance(tabs, list) or not tabs:
            return {"type": "info", "content": "Classeur vide ou format non reconnu."}

        labels: list[str] = []
        for tab in tabs[:8]:
            if not isinstance(tab, dict):
                continue
            label = str(tab.get("label") or "Feuille")
            row_count = tab.get("row_count")
            if not isinstance(row_count, int):
                row_count = tab.get("totalRowCount")
            labels.append(f"{label} ({row_count} lignes)" if isinstance(row_count, int) else label)

        n_tabs = len(tabs)
        suffix = "…" if n_tabs > 8 else ""
        plural = "s" if n_tabs > 1 else ""
        return {
            "type": "info",
            "content": (
                f"Classeur Komptia — {n_tabs} onglet{plural} : "
                + ", ".join(labels)
                + suffix
                + ". Double-cliquez pour l'ouvrir dans la grille."
            ),
        }

    def _preview_image(self, file_path: Path) -> dict:
        """Aperçu d'une image (base64). Limite : ``MAX_PREVIEW_IMAGE_BYTES``."""
        try:
            file_size = file_path.stat().st_size
        except (FileNotFoundError, OSError):
            return {"type": "error", "content": "Fichier introuvable."}
        if file_size > MAX_PREVIEW_IMAGE_BYTES:
            return {
                "type": "info",
                "content": (
                    "Image trop volumineuse pour la prévisualisation "
                    f"(max {MAX_PREVIEW_IMAGE_BYTES // (1024 * 1024)} Mo)."
                ),
            }
        data_b64 = base64.b64encode(file_path.read_bytes()).decode("ascii")
        mime = mimetypes.guess_type(str(file_path))[0] or "image/png"
        return {"type": "image", "content": f"data:{mime};base64,{data_b64}"}

    @authenticated
    async def get(self) -> None:
        user = self.current_user
        _check_rate_limit(_preview_limiter, user.id, *RATE_LIMIT_PREVIEW)

        user_dir = _user_dir(user.id)
        rel = self.get_argument("path", "")

        if not rel:
            return self.write_json({"success": False, "error": "Chemin requis"}, 400)

        target = _safe_path(user_dir, rel)
        if target is None or not target.exists() or not target.is_file():
            return self.write_json({"success": False, "error": "Fichier introuvable"}, 404)

        info = _file_info(target, user_dir)
        if info is None:
            return self.write_json({"success": False, "error": "Fichier introuvable"}, 404)
        ext = target.suffix.lower()
        preview: dict = {"type": "unknown", "content": None}

        # Classeur .afz.json : stocké gzippé sur disque, son extension est ``.json``
        # donc il tomberait dans la branche texte ci-dessous → ``read_text`` sur du
        # gzip = bouillie binaire au simple clic. On l'intercepte ici pour décoder
        # via le SSoT ``decode_afz_bytes`` (gzip-aware) et renvoyer un résumé.
        if target.name.lower().endswith(".afz.json"):
            preview = await asyncio.to_thread(self._preview_workbook, target)
            return self.write_json({"success": True, "file": info, "preview": preview})

        # Les parsers (csv, json, openpyxl) sont sync → off-load pour ne
        # pas bloquer l'event-loop sur un fichier de 5 Mo.
        if ext in {".csv", ".txt", ".md", ".log", ".json", ".xml", ".sql"}:
            preview = await asyncio.to_thread(self._preview_text_file, target, ext)
            if ext == ".sql" and preview.get("type") == "text":
                preview = {"type": "sql", "content": preview.get("content", "")}
        elif ext in {".xlsx", ".xls"}:
            preview = await asyncio.to_thread(self._preview_excel, target)
        elif ext in {".png", ".jpg", ".jpeg"}:
            preview = await asyncio.to_thread(self._preview_image, target)

        self.write_json({"success": True, "file": info, "preview": preview})


# ─── API : Exécution SQL depuis le datastore ────────────────────


class DatastoreSqlExecuteAPIHandler(BaseHandler):
    """POST /api/datastore/sql/execute — exécute un fichier .sql stocké."""

    _MAX_SQL_FILE_BYTES: Final[int] = MAX_SQL_PAYLOAD_BYTES
    #: Filet de sécurité ultime côté caller — la VRAIE source de vérité
    #: du plafond effectif est ``DatabaseConnection.max_rows`` (configuré
    #: via ``/admin/database``). Le connector applique son propre cap
    #: avant de retourner les rows ; ce hard-cap n'intervient que si le
    #: connector ne respecte pas le contrat (mock test, connector tiers
    #: lazy). Ne PAS bumper cette valeur pour servir un cap user — passer
    #: par /admin/database, qui est l'unique levier admin.
    _HARD_CAP_RESULT_ROWS: Final[int] = 10_000_000

    @require_role("admin", "user")
    async def post(self) -> None:
        user = self.current_user
        _check_rate_limit(_sql_exec_limiter, user.id, *RATE_LIMIT_SQL_EXECUTE)

        user_dir = _user_dir(user.id)
        body = self.get_json_body() or {}
        rel = (body.get("path") or "").strip()
        inline_raw = body.get("sql")

        # Deux modes :
        #   - ``inline`` : ``sql`` fourni dans le body → on l'exécute tel quel
        #     (édition en cours dans l'UI, pas encore enregistrée).
        #   - ``file``   : seul ``path`` fourni → on lit le .sql du datastore.
        # Si les deux sont fournis, ``sql`` gagne — ``path`` ne sert que d'audit
        # contextuel (« cette édition transient venait de ce fichier »).
        if isinstance(inline_raw, str) and inline_raw.strip():
            source = "inline"
            sql = inline_raw.strip()
            # **CWE-158 (NUL byte injection)** : certains drivers ODBC tronquent
            # silencieusement la chaîne au premier NUL → la requête réellement
            # exécutée diffère de celle affichée à l'utilisateur (« données fausses
            # silencieusement »). On rejette en amont, comme ``_safe_path``.
            if "\x00" in sql:
                return self.write_json(
                    {"success": False, "error": "Caractère NUL interdit dans la requête."},
                    400,
                )
            sql_bytes = sql.encode("utf-8")
            if len(sql_bytes) > self._MAX_SQL_FILE_BYTES:
                return self.write_json(
                    {
                        "success": False,
                        "error": (
                            f"Requête trop volumineuse "
                            f"(max {self._MAX_SQL_FILE_BYTES // 1024} Ko)."
                        ),
                    },
                    413,
                )
            # ``path`` reste optionnel en mode inline ; s'il est fourni, on ne le
            # promeut en ``context_path`` audité QUE si ``_safe_path`` valide ET
            # que le fichier existe vraiment dans le user_dir. Sinon on log la
            # tentative comme « claim » (chaîne arbitraire fournie par le client,
            # non validée) pour ne pas polluer l'audit canonique avec des chaînes
            # forgées par un utilisateur authentifié.
            context_path: Optional[str] = None
            context_path_claim: Optional[str] = rel or None
            if context_path_claim is not None:
                resolved = _safe_path(user_dir, context_path_claim)
                if resolved is not None and resolved.exists() and resolved.is_file():
                    try:
                        context_path = str(resolved.relative_to(user_dir))
                    except ValueError:
                        context_path = None
            if not PYODBC_AVAILABLE:
                return self.write_json(
                    {
                        "success": False,
                        "error": "Exécution SQL indisponible (pyodbc non installé).",
                    },
                    503,
                )
            logger.info(
                "SQL execute (inline)",
                extra={
                    "user_id": user.id,
                    "source": source,
                    "context_path": context_path,
                    "context_path_claim": context_path_claim,
                    "sql_bytes": len(sql_bytes),
                },
            )
        elif inline_raw is not None and not isinstance(inline_raw, str):
            # Garde-fou contre les payloads malformés (sql: 42, sql: {…}).
            return self.write_json(
                {"success": False, "error": "Format ``sql`` invalide (chaîne attendue)."},
                400,
            )
        else:
            source = "file"
            if not rel:
                return self.write_json({"success": False, "error": "Chemin requis."}, 400)

            target = _safe_path(user_dir, rel)
            if target is None or not target.exists() or not target.is_file():
                return self.write_json({"success": False, "error": "Fichier introuvable."}, 404)

            if target.suffix.lower() != ".sql":
                return self.write_json(
                    {"success": False, "error": "Seuls les fichiers .sql peuvent être exécutés."},
                    400,
                )

            size = target.stat().st_size
            if size > self._MAX_SQL_FILE_BYTES:
                return self.write_json(
                    {
                        "success": False,
                        "error": (
                            f"Fichier trop volumineux "
                            f"(max {self._MAX_SQL_FILE_BYTES // 1024} Ko)."
                        ),
                    },
                    413,
                )

            if not PYODBC_AVAILABLE:
                return self.write_json(
                    {
                        "success": False,
                        "error": "Exécution SQL indisponible (pyodbc non installé).",
                    },
                    503,
                )

            try:
                # ``utf-8-sig`` strip le BOM EF BB BF que certains éditeurs
                # Windows (Notepad, SSMS) insèrent au début. Sans ça, le BOM
                # reste dans le SQL et le validator tronque au mauvais endroit.
                sql = await asyncio.to_thread(
                    target.read_text, encoding="utf-8-sig", errors="replace"
                )
            except OSError as exc:
                logger.warning(
                    "Erreur lecture fichier SQL",
                    extra={"user_id": user.id, "target_name": target.name, "error": str(exc)},
                )
                return self.write_json(
                    {"success": False, "error": "Impossible de lire le fichier."}, 500
                )

            sql = sql.strip()
            if not sql:
                return self.write_json({"success": False, "error": "Requête SQL vide."}, 400)
            # Voir mode inline : NUL → troncature silencieuse driver-dépendante.
            if "\x00" in sql:
                return self.write_json(
                    {
                        "success": False,
                        "error": (
                            "Le fichier SQL contient un caractère NUL. "
                            "Ouvrez-le dans un éditeur et retirez les octets nuls."
                        ),
                    },
                    400,
                )

        executor = QueryExecutor()
        # ``MAX+1`` pour distinguer "exactement MAX" vs "plus que MAX".
        result = await executor.execute_for_ai(
            sql,
            max_rows=self._HARD_CAP_RESULT_ROWS + 1,
            user=self.current_user,
            rls_source="datastore_execute_for_ai",
        )

        if not result.get("success"):
            raw_error = result.get("error") or "Erreur lors de l'exécution."
            # P2.3 (audit 2026-05-26) — Migration vers ``sanitize_sql_for_client``
            # (SSoT P2.1). Avant : tout non-admin recevait LE MÊME message
            # « La requête contient une référence invalide. Vérifiez les noms
            # des tables et colonnes. » quel que soit le type d'erreur SQL
            # Server (FK violation, timeout, deadlock, type mismatch, syntax,
            # permission, etc.) → impossible de diagnostiquer.
            #
            # Maintenant : helper unique qui catégorise + sanitize PII selon
            # l'audience :
            # - admin : ``audience="admin"`` → message verbeux + SQLSTATE +
            #   detail_for_admin (utile pour debug + tickets).
            # - non-admin : ``audience="user"`` → hint catégoriel FR adapté
            #   (referential / type / timeout / permission / etc.) avec
            #   sanitization automatique « Invalid object name 'F_X' » →
            #   message générique mode invisible si user a des règles deny.
            from app.services.data_access.error_messages import (
                sanitize_sql_for_client,
            )
            from app.handlers.base import is_admin as _is_admin

            # P2.3 SSoT — utiliser ``is_admin`` (base.py:901) qui gère
            # robustement enum/string. L'ancien check ``user_role == "admin"``
            # comparait un UserRole enum à une string → toujours False (bug
            # pré-existant : admin recevait silencieusement le message
            # non-admin générique). Fix collateral via ce helper.
            _is_admin_user = _is_admin(self.current_user)
            audience = "admin" if _is_admin_user else "user"
            payload = await sanitize_sql_for_client(raw_error, self.current_user, audience=audience)
            if not _is_admin_user:
                # Log côté serveur le raw pour permettre debug admin via /admin
                logger.warning(
                    "SQL exec error (non-admin user)",
                    extra={
                        "user_id": user.id,
                        "raw_error": raw_error,
                        "sqlstate": payload.get("sqlstate"),
                        "category": payload.get("category"),
                    },
                )
            return self.write_json(
                {
                    "success": False,
                    "error": payload["message"],
                    "category": payload["category"],
                    "sqlstate": payload["sqlstate"],
                    "sql": sql,
                },
                400,
            )

        rows = result.get("data") or []
        # A8-F1 — la troncature AUTORITATIVE vient du connector (cap effectif
        # ``min(caller, DatabaseConnection.max_rows)``), propagée par
        # ``execute_for_ai``. Avant, on ne testait QUE le hard-cap local (10M)
        # que le connector n'atteint jamais (il cape au max_rows admin AVANT) →
        # ``truncated`` toujours False → l'utilisateur croyait voir TOUTES les
        # lignes alors que le résultat était coupé au cap admin (ex: 1000).
        # On OR avec le hard-cap local (defense-in-depth si connector non
        # conforme).
        connector_truncated = bool(result.get("truncated"))
        hard_capped = len(rows) > self._HARD_CAP_RESULT_ROWS
        if hard_capped:
            rows = rows[: self._HARD_CAP_RESULT_ROWS]
        truncated = connector_truncated or hard_capped

        return self.write_json(
            {
                "success": True,
                "sql": sql,
                "columns": result.get("columns") or [],
                "rows": rows,
                "row_count": len(rows),
                "execution_time_ms": result.get("execution_time_ms", 0),
                # A8-F1 — ``max_rows`` = hard-cap local (10M), PAS le cap effectif
                # (qui vit dans ``DatabaseConnection.max_rows``, non exposé par le
                # connector). NE PAS l'afficher comme borne de troncature côté UI :
                # quand ``truncated`` vient du connector (cap admin ex 1000),
                # ce 10M serait trompeur. Le badge « limité » s'appuie sur
                # ``truncated`` + ``row_count``, jamais sur ``max_rows``.
                "max_rows": self._HARD_CAP_RESULT_ROWS,
                "truncated": truncated,
            }
        )


# ─── API : Sauvegarde d'une requête SQL depuis la grille ────────


class DatastoreSqlSaveAPIHandler(BaseHandler):
    """POST /api/datastore/sql/save — enregistre une requête SQL comme fichier .sql."""

    @staticmethod
    def _sanitize_filename(raw: Any) -> Optional[str]:
        """Forme ``_sanitize_user_filename`` spécialisée pour ``.sql``.

        Legacy API — conservée pour les tests qui appellent
        ``DatastoreSqlSaveAPIHandler._sanitize_filename`` directement.
        """
        return _sanitize_user_filename(raw, require_extension=True, default_ext=".sql")

    @require_role("admin", "user")
    async def post(self) -> None:
        user = self.current_user
        _check_rate_limit(_save_search_limiter, user.id, *RATE_LIMIT_SAVE_SEARCH)

        user_dir = _user_dir(user.id)
        body = self.get_json_body() or {}

        filename = _sanitize_user_filename(
            body.get("filename") or "", require_extension=True, default_ext=".sql"
        )
        if not filename:
            return self.write_json({"success": False, "error": "Nom de fichier invalide."}, 400)

        sql = body.get("sql")
        if not isinstance(sql, str) or not sql.strip():
            return self.write_json({"success": False, "error": "Requête SQL vide."}, 400)

        sql_bytes = sql.encode("utf-8")
        if len(sql_bytes) > MAX_SQL_PAYLOAD_BYTES:
            return self.write_json(
                {
                    "success": False,
                    "error": (
                        f"Requête trop volumineuse " f"(max {MAX_SQL_PAYLOAD_BYTES // 1024} Ko)."
                    ),
                },
                413,
            )

        overwrite = bool(body.get("overwrite"))

        target = _safe_path(user_dir, filename)
        if target is None:
            return self.write_json({"success": False, "error": "Chemin invalide."}, 400)

        try:
            await asyncio.to_thread(user_dir.mkdir, parents=True, exist_ok=True)
        except OSError as exc:
            logger.warning(
                "Erreur mkdir user_dir",
                extra={"user_id": user.id, "error": str(exc)},
            )
            return self.write_json(
                {"success": False, "error": "Impossible de créer le dossier utilisateur."},
                500,
            )

        try:
            rel = str(target.relative_to(user_dir))
        except ValueError:
            rel = filename

        file_size = len(sql_bytes)
        file_hash = calculate_hash_from_bytes(sql_bytes)

        # T10 — alignement sur le pipeline upload SSoT. L'ancien handler écrivait
        # en direct (``open("w"/"x")``) SANS : (a) compter dans le quota, (b) créer
        # de FileMetadata (→ .sql invisible des stats/context-files/cleanup,
        # incohérence FS↔BDD dès la création — chemin nominal), (c) écriture
        # atomique (crash = .sql tronqué + original perdu), (d) audit. On réutilise
        # les mêmes helpers que ``DatastoreUploadAPIHandler``.
        target_exists = await asyncio.to_thread(target.exists)
        if target_exists and not overwrite:
            return self.write_json(
                {
                    "success": False,
                    "error": "Un fichier portant ce nom existe déjà.",
                    "code": "exists",
                    "filename": filename,
                },
                409,
            )

        # Quota check AVANT écriture (session courte isolée, comme l'upload).
        async with get_session() as quota_db:
            storage_mgr_quota = StorageManager(quota_db, DATASTORE_DIR)
            can_save, quota_error = await storage_mgr_quota.check_quota(user.id, file_size)
        if not can_save:
            return self.write_json(
                {"success": False, "error": quota_error, "error_code": "QUOTA_EXCEEDED"},
                413,
            )

        # Overwrite : décrémenter + délier l'ancien AVANT de réécrire.
        # ``register_upload`` n'UPSERT PAS (ré-INSERT + ré-incrément atomique) →
        # sans deletion préalable on dupliquerait le FileMetadata et doublerait le
        # quota. ``register_deletion`` no-op gracieusement (False) si l'ancien .sql
        # est legacy (créé avant T10, sans metadata).
        if target_exists and overwrite:
            await _storage_register_deletion_with_retry(user.id, rel)
            try:
                await asyncio.to_thread(target.unlink)
            except FileNotFoundError:
                pass

        def _write_atomic_sql() -> Optional[str]:
            """Écrit le .sql atomiquement. ``"exists"`` si une course a recréé le fichier."""
            if overwrite:
                # Cible déjà déliée → ``.tmp`` UNIQUE + ``os.replace`` pour que le
                # fichier visible apparaisse atomiquement (un reader ne voit jamais
                # un .sql partiel). Token unique (T9) = pas de collision .tmp entre
                # 2 saves concurrents. Cleanup du .tmp orphelin si crash write↔replace.
                tmp_target = _unique_tmp_path(target)
                try:
                    tmp_target.write_bytes(sql_bytes)
                    os.replace(str(tmp_target), str(target))
                except BaseException:
                    try:
                        if tmp_target.exists():
                            tmp_target.unlink()
                    except OSError:
                        pass
                    raise
                return None
            # Non-overwrite : création exclusive atomique (``O_EXCL``) = race-safe
            # (pas d'overwrite silencieux d'un fichier créé entre le check et l'écriture).
            try:
                fd = os.open(str(target), os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
            except FileExistsError:
                return "exists"
            with os.fdopen(fd, "wb") as f:
                f.write(sql_bytes)
            return None

        try:
            outcome = await asyncio.to_thread(_write_atomic_sql)
        except OSError as exc:
            logger.warning(
                "Erreur écriture fichier SQL",
                extra={"user_id": user.id, "target_name": target.name, "error": str(exc)},
            )
            return self.write_json(
                {"success": False, "error": "Impossible d'écrire le fichier."}, 500
            )

        if outcome == "exists":
            return self.write_json(
                {
                    "success": False,
                    "error": "Un fichier portant ce nom existe déjà.",
                    "code": "exists",
                    "filename": filename,
                },
                409,
            )

        # FileMetadata + quota (session isolée + retry, fichier déjà sur disque),
        # puis audit fire-and-forget — exactement le pipeline de l'upload.
        metadata = await _storage_register_upload_with_retry(
            user.id,
            target,
            rel,
            file_size=file_size,
            file_hash=file_hash,
        )
        _schedule_audit_fire_and_forget(
            action=AuditAction.FILE_UPLOAD,
            user_id=user.id,
            entity_id=metadata.id,
            details={"filename": rel, "size": file_size},
            ip_address=self.request.remote_ip,
            op_label="sql_save",
        )

        # ``file_hash`` renvoyé pour parité ETag avec l'upload (If-Match ultérieur).
        return self.write_json(
            {"success": True, "filename": filename, "path": rel, "file_hash": file_hash}, 201
        )


# ─── API : Lister les dossiers (pour modale de déplacement) ─────


class DatastoreFoldersAPIHandler(BaseHandler):
    """GET /api/datastore/folders — liste tous les dossiers avec profondeur."""

    @authenticated
    async def get(self) -> None:
        user = self.current_user
        _check_rate_limit(_folders_limiter, user.id, *RATE_LIMIT_FOLDERS)
        user_dir = _user_dir(user.id)
        # M-11 : ``?exclude_source=<path>`` filtre la source elle-même, ses
        # descendants ET son parent direct (no-op destination). Évite que
        # l'UI move propose une cible invalide qui échouera côté serveur.
        exclude_rel = (self.get_argument("exclude_source", "") or "").strip()
        # ``rglob`` + ``is_dir`` peut être coûteux → off-load.
        folders = await asyncio.to_thread(_walk_folders, user_dir)

        if exclude_rel:
            exclude_path = _safe_path(user_dir, exclude_rel)
            if exclude_path is not None:
                try:
                    exclude_norm = _normalize_rel_path(exclude_path, user_dir)
                except ValueError:
                    exclude_norm = exclude_rel
                exclude_parent: Optional[str]
                try:
                    parent = exclude_path.parent
                    if parent == user_dir:
                        exclude_parent = ""
                    else:
                        exclude_parent = _normalize_rel_path(parent, user_dir)
                except (ValueError, OSError):
                    exclude_parent = None
                prefix = f"{exclude_norm}/"
                folders = [
                    f
                    for f in folders
                    if f["path"] != exclude_norm
                    and not f["path"].startswith(prefix)
                    and f["path"] != exclude_parent
                ]
        self.write_json({"success": True, "folders": folders})


# ─── API : Déplacer un fichier/dossier ──────────────────────────


class DatastoreMoveAPIHandler(BaseHandler):
    """POST /api/datastore/move — déplace un fichier/dossier."""

    def _validate_move_paths(
        self,
        source_rel: str,
        dest_rel: str,
        allow_root: bool,
        user_dir: Path,
    ) -> tuple[Optional[Path], Optional[Path]]:
        """Valide les chemins source et destination.

        M-5 : ``dest_rel`` vide est REJETÉ sauf si ``allow_root=True``
        (flag explicite côté client). Auparavant un destination vide
        déplaçait silencieusement le fichier à la racine — comportement
        destructif par défaut si l'utilisateur clique "Déplacer" sans
        sélectionner de cible.
        """
        if not source_rel:
            self.write_json({"success": False, "error": "Source manquante"}, 400)
            return None, None

        source_path = _safe_path(user_dir, source_rel)
        if not source_path or not source_path.exists():
            self.write_json({"success": False, "error": "Fichier source introuvable"}, 404)
            return None, None

        if dest_rel:
            dest_folder = _safe_path(user_dir, dest_rel)
            if not dest_folder or not dest_folder.is_dir():
                self.write_json({"success": False, "error": "Dossier de destination invalide"}, 400)
                return None, None
        else:
            if not allow_root:
                self.write_json(
                    {
                        "success": False,
                        "error": "Destination requise. Pour déplacer à la racine, utiliser explicitement la cible 'Racine'.",
                    },
                    400,
                )
                return None, None
            dest_folder = user_dir

        # M-3 : no-op si source est déjà dans dest_folder (parent identique).
        if source_path.parent == dest_folder:
            self.write_json({"success": True, "message": "Aucun changement"})
            return None, None

        # Déplacer un dossier dans lui-même / un descendant → boucle.
        if source_path.is_dir():
            try:
                dest_folder.relative_to(source_path)
                self.write_json(
                    {
                        "success": False,
                        "error": "Impossible de déplacer un dossier dans lui-même",
                    },
                    400,
                )
                return None, None
            except ValueError:
                pass

        return source_path, dest_folder

    async def _update_file_metadata_after_move(
        self,
        session: Any,
        source_path: Path,
        dest_path: Path,
        user: Any,
        user_dir: Path,
    ) -> int:
        """Met à jour les ``FileMetadata`` APRÈS le move FS (FS-first).

        Retourne le nombre de FileMetadata effectivement mis à jour /
        créés. Si l'entrée n'existe pas pour un fichier (sync dérivé),
        on appelle ``register_upload(dest_path)`` — sûr maintenant que
        ``dest_path`` existe physiquement (M-7 — auparavant cet appel
        était fait AVANT le ``shutil.move``, causant un FileNotFoundError
        ou un quota silencieusement faux).
        """
        old_rel_path = _normalize_rel_path(source_path.parent / source_path.name, user_dir)
        # Recompute avec ``Path`` post-move : ``source_path`` peut être
        # invalidé (Pathlib ne tracke pas le rename), donc on calcule
        # ``old_rel_path`` à partir de la composition originale.
        new_rel_path = _normalize_rel_path(dest_path, user_dir)
        now_ts = int(clock.timestamp())
        affected = 0

        if dest_path.is_file():
            result = await session.execute(
                select(FileMetadata)
                .where(FileMetadata.user_id == user.id)
                .where(FileMetadata.file_path == old_rel_path)
            )
            metadata = result.scalar_one_or_none()
            if metadata:
                metadata.file_path = new_rel_path
                metadata.filename = dest_path.name
                metadata.extension = dest_path.suffix.lower()
                metadata.updated_at = now_ts
                affected = 1
            else:
                storage_manager = StorageManager(session, user_dir)
                await storage_manager.register_upload(user.id, dest_path, new_rel_path)
                affected = 1
        elif dest_path.is_dir():
            old_prefix = f"{old_rel_path}/"
            new_prefix = f"{new_rel_path}/"
            escaped_base = (
                old_rel_path.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            )
            result = await session.execute(
                select(FileMetadata).where(
                    FileMetadata.user_id == user.id,
                    FileMetadata.file_path.like(f"{escaped_base}/%", escape="\\"),
                )
            )
            for file_meta in result.scalars().all():
                if file_meta.file_path.startswith(old_prefix):
                    file_meta.file_path = new_prefix + file_meta.file_path[len(old_prefix) :]
                    file_meta.updated_at = now_ts
                    affected += 1
            if affected:
                await session.flush()
        return affected

    @require_role("admin", "user")
    async def post(self) -> None:
        """Déplace un fichier ou dossier vers une autre destination.

        Role: ``admin``/``user`` uniquement (fix security : l'ancien
        décorateur ``@authenticated`` laissait ``reader`` déplacer les
        fichiers d'autrui — implicite via session cookie compromise).
        """
        user = self.current_user
        _check_rate_limit(_move_limiter, user.id, *RATE_LIMIT_MOVE)
        user_dir = _user_dir(user.id)

        try:
            data = self.get_json_body()
            source_rel = (data.get("source", "") or "").strip()
            dest_rel = (data.get("destination", "") or "").strip()
            allow_root = bool(data.get("move_to_root", False))

            source_path, dest_folder = self._validate_move_paths(
                source_rel, dest_rel, allow_root, user_dir
            )
            if source_path is None or dest_folder is None:
                return

            dest_path = dest_folder / source_path.name
            was_dir = source_path.is_dir()
            source_rel_norm = _normalize_rel_path(source_path, user_dir)

            # FS first via helper atomique fail-if-exists (M-1 race silent
            # overwrite). DB après — si DB échoue, on log un warning mais
            # l'état FS est cohérent.
            try:
                await asyncio.to_thread(_atomic_move_failing_if_exists, source_path, dest_path)
            except FileExistsError:
                self.write_json(
                    {
                        "success": False,
                        "error": f"'{source_path.name}' existe déjà dans la destination",
                    },
                    409,
                )
                return

            dest_rel_norm = _normalize_rel_path(dest_path, user_dir)
            affected_files_count = 0
            try:
                async with get_session() as session:
                    affected_files_count = await self._update_file_metadata_after_move(
                        session, source_path, dest_path, user, user_dir
                    )
            except SQLAlchemyError as exc:
                logger.warning(
                    "Quota drift après move FS — DB unreachable",
                    extra={
                        "user_id": user.id,
                        "source": source_rel_norm,
                        "destination": dest_rel_norm,
                        "error": str(exc),
                    },
                )

            _schedule_audit_fire_and_forget(
                action=AuditAction.FILE_MOVE,
                user_id=user.id,
                details={
                    "source": source_rel_norm,
                    "destination": dest_rel_norm,
                    "was_dir": was_dir,
                    "affected_files_count": affected_files_count,
                },
                ip_address=self.request.remote_ip,
                op_label=f"move {source_rel_norm} -> {dest_rel_norm}",
            )

            logger.info(
                "Fichier déplacé",
                extra={
                    "user_id": user.id,
                    "source": source_rel_norm,
                    "destination": dest_rel_norm,
                },
            )
            self.write_json({"success": True})

        except OSError as e:
            logger.error("Erreur déplacement fichier: %s", e, exc_info=True)
            self.write_json(
                {
                    "success": False,
                    "error": "Une erreur est survenue lors du déplacement du fichier.",
                },
                500,
            )


# ─── API : Sauvegarder une recherche ────────────────────────────────


class SaveSearchAPIHandler(BaseHandler):
    """POST /api/datastore/save-search — sauvegarde une recherche en CSV."""

    # Caller niveau infini. ``sage_connector.execute()`` applique
    # ``min(_MAX_ROWS_EXPORT, db_conn.max_rows)`` — admin /admin/database
    # est l'UNIQUE source de vérité du plafond effectif.
    _MAX_ROWS_EXPORT: Final[int] = 1_000_000_000

    @authenticated
    async def post(self) -> None:
        user = self.current_user
        _check_rate_limit(_save_search_limiter, user.id, *RATE_LIMIT_SAVE_SEARCH)

        try:
            data = self.get_json_body()
            raw_search_id = data.get("search_id")
            folder = data.get("folder", "")

            if not raw_search_id:
                return self.write_json({"success": False, "error": "search_id requis"}, 400)

            try:
                search_id = int(raw_search_id)
            except (ValueError, TypeError):
                return self.write_json({"success": False, "error": "search_id invalide"}, 400)

            if not PYODBC_AVAILABLE:
                logger.warning(
                    "pyodbc non disponible — export impossible en mode dev",
                    extra={"user_id": user.id, "search_id": search_id},
                )
                return self.write_json(
                    {"success": False, "error": "Export impossible en mode dev sans pyodbc"},
                    503,
                )

            # Un seul ``get_session`` pour toute la séquence (atomicité :
            # si le write FS échoue après ``register_upload``, rollback).
            async with get_session() as session:
                result = await session.execute(
                    select(SearchHistory).where(
                        SearchHistory.id == search_id,
                        SearchHistory.user_id == user.id,
                    )
                )
                search_record = result.scalar_one_or_none()

                if not search_record:
                    return self.write_json(
                        {"success": False, "error": "Recherche introuvable"}, 404
                    )
                if not search_record.success:
                    return self.write_json(
                        {"success": False, "error": "Cette recherche a échoué, pas de résultats"},
                        400,
                    )

                # Re-exécution (note : les résultats originaux ne sont pas
                # stockés — export = snapshot "maintenant" qui peut différer
                # de la recherche originale si la BDD a bougé).
                executor = QueryExecutor()
                query_result = await executor.execute(
                    search_record.sql_validated,
                    max_rows=self._MAX_ROWS_EXPORT,
                    user=self.current_user,
                    rls_source="datastore_export",
                )
                results = query_result.to_dicts()

                if not results:
                    return self.write_json({"success": False, "error": "Aucun résultat"}, 404)

                # Délègue au service unifié ``csv_export.to_csv_bytes`` :
                # bytes UTF-8 BOM (Excel-friendly) + sanitisation
                # OWASP-CSV-Injection sur headers ET valeurs par défaut
                # — un alias SQL ``=cmd|`` ne pouvait plus glisser ici.
                fieldnames = list(results[0].keys())
                csv_bytes = to_csv_bytes(results, columns=fieldnames)

                user_dir = _user_dir(user.id)
                safe_question = (
                    "".join(
                        c if c.isalnum() or c in (" ", "-", "_") else "_"
                        for c in (search_record.question or "")[:50]
                    ).strip()
                    or "recherche"
                )
                filename_candidate = f"{safe_question}_{search_id}.csv"
                filename = _sanitize_user_filename(
                    filename_candidate, require_extension=True, default_ext=".csv"
                )
                if filename is None:
                    # Fallback defensive : nom 100% ASCII/safe.
                    filename = f"recherche_{search_id}.csv"

                if folder:
                    dest_dir = _safe_path(user_dir, folder)
                    if dest_dir is None:
                        return self.write_json({"success": False, "error": "Chemin invalide"}, 400)
                    dest_dir.mkdir(parents=True, exist_ok=True)
                else:
                    dest_dir = user_dir

                file_path = dest_dir / filename
                file_size = len(csv_bytes)
                file_hash = calculate_hash_from_bytes(csv_bytes)

                storage_mgr = StorageManager(session, DATASTORE_DIR)
                can_upload, error_msg = await storage_mgr.check_quota(user.id, file_size)
                if not can_upload:
                    # ``error_code`` machine-readable : le client distingue le
                    # quota app (JSON, status 413) de l'oversize passerelle nginx
                    # (HTML, status 413 aussi) — cf. _saveToPathAsync côté JS.
                    return self.write_json(
                        {
                            "success": False,
                            "error": error_msg,
                            "error_code": "QUOTA_EXCEEDED",
                        },
                        413,
                    )

                relative_path = str(file_path.relative_to(user_dir))
                metadata = await storage_mgr.register_upload(
                    user.id,
                    file_path,
                    relative_path,
                    description=f"Recherche: {search_record.question}",
                    file_size=file_size,
                    file_hash=file_hash,
                )
                metadata.search_history_id = search_id

                # Atomic exclusive create (pas d'overwrite silencieux).
                def _atomic_write() -> Optional[str]:
                    try:
                        fd = os.open(
                            str(file_path),
                            os.O_WRONLY | os.O_CREAT | os.O_EXCL,
                            0o600,
                        )
                    except FileExistsError:
                        return "exists"
                    try:
                        # ``wb`` car ``csv_bytes`` inclut déjà le BOM UTF-8
                        # et est encodé en utf-8 par ``to_csv_bytes``.
                        with os.fdopen(fd, "wb") as f:
                            f.write(csv_bytes)
                    except OSError:
                        # fd déjà consommé par fdopen ; pas de second close.
                        raise
                    return None

                write_outcome = await asyncio.to_thread(_atomic_write)
                if write_outcome == "exists":
                    return self.write_json(
                        {
                            "success": False,
                            "error": "Un fichier portant ce nom existe déjà.",
                            "filename": filename,
                        },
                        409,
                    )

                session.add(
                    AuditLog.log_action(
                        action=AuditAction.FILE_SEARCH_EXPORT,
                        user_id=user.id,
                        entity_type="file",
                        entity_id=metadata.id,
                        details={
                            "filename": relative_path,
                            "search_id": search_id,
                            "size": file_size,
                        },
                        ip_address=self.request.remote_ip,
                    )
                )

            logger.info(
                "Recherche sauvegardée",
                extra={
                    "user_id": user.id,
                    "saved_filename": filename,
                    "size_bytes": file_size,
                    "search_id": search_id,
                },
            )
            self.write_json(
                {
                    "success": True,
                    "filename": filename,
                    "path": relative_path,
                    "size": file_size,
                    "size_human": _human_size(file_size),
                    # #18e (triage caps 2026-06-10) — le connector applique
                    # min(caller, cap admin) : si le cap admin a tronqué la
                    # ré-exécution, le CSV sauvegardé est PARTIEL. Avant ce
                    # flag, aucun signal — l'utilisateur archivait/diffusait
                    # un fichier incomplet en silence.
                    "truncated": bool(getattr(query_result, "truncated", False)),
                    **(
                        {
                            "warning": (
                                "Résultat tronqué au cap de lignes "
                                "(/admin/database) — le CSV sauvegardé est "
                                "PARTIEL."
                            )
                        }
                        if getattr(query_result, "truncated", False)
                        else {}
                    ),
                }
            )

        except (csv.Error, OSError) as e:
            logger.error("Erreur sauvegarde recherche: %s", e, exc_info=True)
            self.write_json(
                {
                    "success": False,
                    "error": "Une erreur est survenue lors de la sauvegarde de la recherche.",
                },
                500,
            )


# ─── API : Liste fichiers avec contexte ────────────────────────────────


class ContextFilesAPIHandler(BaseHandler):
    """GET /api/datastore/context-files — fichiers avec SQL ou colonnes."""

    @authenticated
    async def get(self) -> None:
        user = self.current_user
        try:
            async with get_session() as session:
                # Fetch ``FileMetadata`` + ``SearchHistory`` en une seule passe.
                # Fix N+1 : l'ancien code faisait un ``select(SearchHistory)``
                # par fichier → 1 + N queries. On agrège par ``search_id``
                # en amont pour tout résoudre en 2 queries max.
                files_stmt = (
                    select(FileMetadata)
                    .where(
                        FileMetadata.user_id == user.id,
                        (FileMetadata.search_history_id.isnot(None))
                        | (FileMetadata.columns_json.isnot(None)),
                    )
                    .order_by(FileMetadata.created_at.desc())
                )
                files_result = await session.execute(files_stmt)
                files = files_result.scalars().all()

                search_ids = {f.search_history_id for f in files if f.search_history_id}
                sql_by_id: dict[int, str] = {}
                if search_ids:
                    # Defense-in-depth : la 1re query filtre déjà ``FileMetadata.user_id == user.id``,
                    # donc en théorie ces ``search_history_id`` pointent tous sur des SearchHistory
                    # du même user. Mais si un import/restore/migration ratés assignaient un FK
                    # cross-user, on leakerait le texte SQL d'un autre compte (noms de tables +
                    # filtres avec valeurs réelles). On re-filtre par ``user_id`` ici comme
                    # garde-fou — strict subset = même résultat dans le happy path, fail-closed
                    # sur les rows mal liés.
                    sh_result = await session.execute(
                        select(SearchHistory).where(
                            SearchHistory.id.in_(search_ids),
                            SearchHistory.user_id == user.id,
                        )
                    )
                    for record in sh_result.scalars().all():
                        sql_by_id[record.id] = record.sql_validated or record.sql_generated or ""

                files_data: list[dict] = []
                for file in files:
                    # ``updated_at`` peut être 0 après migration → fallback created_at.
                    ts = file.updated_at or file.created_at or 0
                    try:
                        modified_human = clock.local_from_timestamp(ts).strftime("%d/%m/%Y %H:%M")
                    except (OverflowError, OSError, ValueError):
                        modified_human = ""
                    file_info: dict = {
                        "id": file.id,
                        "filename": file.filename,
                        "description": file.description,
                        "size_bytes": file.size_bytes,
                        "size_human": _human_size(file.size_bytes),
                        "context_type": file.context_type,
                        "modified_human": modified_human,
                    }
                    if file.context_type == "search" and file.search_history_id:
                        sql = sql_by_id.get(file.search_history_id)
                        if sql:
                            file_info["sql"] = sql
                    elif file.context_type == "import":
                        file_info["columns"] = file.columns
                    files_data.append(file_info)

                self.write_json({"success": True, "files": files_data})

        except SQLAlchemyError as e:
            logger.error("Erreur récupération fichiers contexte: %s", e, exc_info=True)
            self.write_json(
                {
                    "success": False,
                    "error": "Une erreur est survenue lors de la récupération des fichiers.",
                },
                500,
            )
