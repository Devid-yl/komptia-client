"""Builder XLSX serveur — Piste 3 (export complet d'un classeur Iris).

Prend en entrée le payload JSON renvoyé par ``GridTabManager.serialize()``
côté frontend (la même forme que celle persistée dans un ``.afz.json``),
réexécute les requêtes SQL de chaque onglet via ``QueryExecutor`` avec un
cap élevé (par défaut 100 000 lignes) — donc avec **les données complètes**
plutôt que le snapshot tronqué à 500 lignes du frontend — puis assemble un
fichier ``.xlsx`` multi-feuilles avec :

- Hyperliens cliquables cellule → feuille de détail (cellules dashboard /
  emit_tab / copilot dont le détail est reconstructible par filtrage du tab
  source via ``match`` / ``match_exclude``)
- Couleur d'onglet bleu clair pour distinguer visuellement les feuilles de
  détail des feuilles principales
- Headers stylés (bold blanc sur fond bleu indigo) + ligne figée
- Format nombre français (#,##0 / #,##0.00)
- Fusions de cellules (mergeCells)

Sécurité — la promesse RLS est tenue parce que chaque réexécution passe par
``QueryExecutor.execute(user=current_user, rls_source="iris_export_xlsx_full")``
qui invoque ``enforcer.enforce_for_executor`` AVANT envoi à SQL Server.
Si l'utilisateur n'a pas accès à une table/colonne/ligne, l'onglet
correspondant est exporté avec son snapshot frontend (déjà filtré par RLS
au moment de l'exécution initiale) — pas de fuite de données.

Pseudonymisation — non appliquée. ``QueryExecutor`` ne pseudonymise pas
(c'est le pipeline copilot qui anonymise les valeurs avant envoi LLM,
hors scope de l'export utilisateur final).
"""

from __future__ import annotations

import asyncio
import io
import logging
import re
from typing import Any, Dict, List, Optional, Set, Tuple

from app.services.database.query_executor import QueryExecutor, get_query_executor
from app.services.data_access.enforcer import DataAccessDeniedError

logger = logging.getLogger(__name__)


# ── Constantes ─────────────────────────────────────────────────────────────

# (Historique) ``_DEFAULT_MAX_ROWS_PER_TAB = 100_000`` a été RETIRÉ le
# 2026-06-10 : ce défaut hardcodé écrasait le cap admin
# ``DatabaseConnection.max_rows`` dans les deux sens (doctrine no-double-cap,
# finding #18b du triage caps). Les défauts de ``build_iris_xlsx`` /
# ``materialize_workbook_sql_tabs`` sont désormais ``None`` = cap admin
# résolu par ``QueryExecutor.execute``. Un caller peut toujours passer un
# cap explicite (preview, param client de /api/iris/export-xlsx-full).

_MAX_DETAIL_ROWS: int = 5_000
"""Cap par feuille de détail. Plus bas que le cap principal parce que
chaque cellule de dashboard peut générer une feuille — limiter pour ne
pas faire exploser la taille du fichier final."""

_MAX_TABS_PER_EXPORT: int = 200
"""Garde-fou : un classeur avec >200 onglets indique probablement un usage
hors-cadre. Le builder refuse au-delà pour protéger la RAM serveur."""

_MAX_DETAIL_SHEETS_PER_EXPORT: int = 1_000
"""Garde-fou similaire pour les feuilles de détail (cellDetails)."""

# T12a (2026-06-10) — plus de timeout hardcodé. ``sql_timeout_s`` défaut ``None``
# → ``QueryExecutor.execute`` résout le timeout admin (``connector.timeout``,
# configuré via /admin/database), SSoT unique (même doctrine que ``max_rows_per_tab``
# et les 4 call-sites nettoyés le 2026-06-08 : copilot_iris_bridge ×2,
# sql_rewrite_service, training_store). L'ancien ``180`` hardcodé IGNORAIT
# silencieusement un timeout admin plus strict (ex. 120s). Un caller qui a
# légitimement besoin de plus pour un export lourd peut passer un override
# explicite ; un onglet qui timeoute est skippé avec un warning VISIBLE
# (surfacé jusqu'au récap email/PDF depuis T8).

_RLS_SOURCE = "iris_export_xlsx_full"

# Couleurs (cohérentes avec le frontend iris-grid.js export Excel client-side).
_HEADER_FONT_RGB = "FFFFFFFF"
_HEADER_FILL_RGB = "FF4472C4"  # Bleu Komptia (cf. styles.xml côté JS)
_HYPERLINK_FONT_RGB = "FF0563C1"  # Bleu hyperlink standard Office
_DETAIL_TAB_COLOR = "FF8DB4E2"  # Bleu clair pour distinguer les feuilles de détail
_HEADER_NUMBER_FORMAT_INT = "#,##0"
_HEADER_NUMBER_FORMAT_DEC = "#,##0.00"

# Largeurs FIXES par type — matche le comportement de l'export JS
# client-side (iris-grid.js _buildSheetXml). Pas d'auto-fit-to-content :
# une cellule avec une valeur très longue n'élargit pas toute la colonne.
_WIDTH_NUMBER = 14
_WIDTH_DATE = 16
_WIDTH_DEFAULT = 20

_INVALID_SHEET_NAME_CHARS = re.compile(r"[/\\?*\[\]:]")

# Caractères interdits dans le XML 1.0 (et donc dans XLSX). Les couches
# qui rejettent :
#  - openpyxl (``IllegalCharacterError`` au cell.value=) sur les contrôles
#    C0 sauf \t \n \r.
#  - lxml (au moment de la sérialisation finale du workbook) sur le même
#    ensemble + surrogates invalides + non-caractères ￾ / ￿.
# Cas typique : champ BLOB Sage casté en string par pyodbc qui contient
# des octets arbitraires, ou décodage UTF-8 partiellement raté qui laisse
# des surrogates orphelins.
#
# Plage couverte :
#   \x00-\x08, \x0B, \x0C, \x0E-\x1F  → contrôles C0 (sauf \t \n \r)
#   \uD800-\uDFFF                     → surrogates Unicode (toujours invalides
#                                       hors paire — illégaux en XML 1.0)
#   ￾, ￿                    → non-caractères réservés
_XLSX_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\uD800-\uDFFF￾￿]")


async def _scrub_tab_labels_for_user(
    tabs: List[Dict[str, Any]], user: Any
) -> List[Dict[str, Any]]:
    """**#144** — Scrub les ``label`` des onglets pour retirer les noms
    de tables denied par le user courant.

    Comportement :

    - ``user is None`` → renvoie ``tabs`` tel quel (caller backend sans
      contexte user — cas scheduled job sans user owner identifiable).
    - admin / sans restrictions → ``scrub_text_for_user`` court-circuit
      O(1), renvoie le label inchangé.
    - sinon → remplace les noms denied par ``[…]`` via la même fonction
      que partout ailleurs (#124, #140, #141).

    **Isolation des dicts** : on renvoie une nouvelle list avec des
    SHALLOW copies des tabs (chaque dict tab est copié pour ne pas
    muter la source — workbook persisté en BDD potentiellement). Les
    valeurs imbriquées (``columns``, ``rows``) restent partagées par
    référence car on ne les touche pas (et leur copy coûterait cher
    sur 100k lignes).

    **Fail-safe** : si le scrub d'un label crash, on garde l'original
    (mieux qu'un export vide).
    """
    if user is None or not tabs:
        return tabs
    from app.services.data_access.error_messages import scrub_text_for_user

    out: List[Dict[str, Any]] = []
    for tab in tabs:
        if not isinstance(tab, dict):
            out.append(tab)
            continue
        label = tab.get("label")
        new_tab = dict(tab)  # shallow copy
        if isinstance(label, str) and label:
            try:
                new_tab["label"] = await scrub_text_for_user(
                    label, user, context_label="xlsx_export_sheet_name"
                )
            except Exception:
                new_tab["label"] = label  # fail-safe
        out.append(new_tab)
    return out


# ── API publique ────────────────────────────────────────────────────────────


async def build_iris_xlsx(
    payload: Dict[str, Any],
    user: Any,
    *,
    max_rows_per_tab: Optional[int] = None,
    sql_timeout_s: Optional[int] = None,
    executor: Optional[QueryExecutor] = None,
    anonymize: bool = False,
) -> Dict[str, Any]:
    """Re-exécute les SQL des onglets et construit le .xlsx complet.

    Args:
        payload: dict au format ``serialize()`` (clé ``tabs`` requise)
        user: utilisateur authentifié pour la RLS sur la réexécution SQL
        max_rows_per_tab: cap par onglet. ``None`` (défaut depuis le
            2026-06-10) = cap admin ``DatabaseConnection.max_rows``
            (/admin/database), résolu par ``QueryExecutor.execute`` —
            doctrine « admin = unique source de vérité, pas de double cap » :
            l'ancien défaut 100 000 hardcodé ÉCRASAIT la config admin dans
            les deux sens (admin 10k → l'export prenait 100k ; admin 500k →
            cappé à 100k). Un appelant peut toujours passer un cap explicite
            (ex. preview à 100 lignes).
        sql_timeout_s: timeout par requête en secondes. ``None`` (défaut) →
            ``QueryExecutor`` résout le timeout admin (``connector.timeout``,
            SSoT /admin/database). Passer une valeur pour un override explicite.
        executor: injectable pour tests, sinon ``get_query_executor()``
        anonymize: si ``True``, applique les pseudonymes configurés par
            l'utilisateur sur ``/data/privacy`` aux VALEURS de cellules avant
            la sérialisation (export « valeurs anonymisées »). Fail-closed si
            un terme configuré ne peut être chargé. Default ``False`` = export
            en clair (comportement historique).

    Returns:
        dict avec :
            - ``content`` (bytes) : le fichier .xlsx
            - ``stats`` (dict) : tabs_count, total_rows, sql_re_executed,
              sql_skipped (avec raison), detail_sheets_count, truncated_tabs
            - ``warnings`` (list[str]) : avertissements non-bloquants

    Raises:
        ValueError : payload invalide (forme inattendue, trop d'onglets…)
        ImportError : openpyxl non installé (transformé en ValueError pour
            uniformiser le code de retour côté handler)
    """
    if not isinstance(payload, dict):
        raise ValueError("Payload doit être un dict")
    tabs = payload.get("tabs")
    if not isinstance(tabs, list):
        raise ValueError("Payload.tabs doit être une liste")
    if len(tabs) == 0:
        raise ValueError("Aucun onglet à exporter")
    if len(tabs) > _MAX_TABS_PER_EXPORT:
        raise ValueError(
            f"Trop d'onglets ({len(tabs)} > {_MAX_TABS_PER_EXPORT}) — refus pour "
            "protéger la RAM serveur"
        )

    if executor is None:
        executor = get_query_executor()

    # Étape 1 : ré-exécuter les SQL via le helper extrait, partagé avec
    # les autres consommateurs de classeur côté backend (copilot bridge,
    # PDF report, email conversion implicite, CSV automation).
    materialization = await materialize_workbook_sql_tabs(
        tabs,
        user,
        max_rows_per_tab=max_rows_per_tab,
        sql_timeout_s=sql_timeout_s,
        executor=executor,
        rls_source=_RLS_SOURCE,
        logger_prefix="iris_export_xlsx_full",
    )
    enriched_tabs = materialization["tabs"]
    warnings = materialization["warnings"]
    sql_re_executed = materialization["sql_re_executed"]
    sql_skipped = materialization["sql_skipped"]
    truncated_tabs = materialization["truncated_tabs"]

    # **#144** — Scrub des labels d'onglets pour mode invisible
    # rétroactif. Si user a nommé un onglet ``F_SALAIRES`` avant
    # d'être denied sur cette table, l'export sinon leak le nom via
    # le titre de feuille Excel (visible dans la barre d'onglets +
    # dans les noms des feuilles de détail D-{label}-r_c).
    enriched_tabs = await _scrub_tab_labels_for_user(enriched_tabs, user)

    # **Anonymisation optionnelle** — export « valeurs anonymisées ». Applique
    # les pseudonymes configurés par l'utilisateur sur ``/data/privacy`` (table
    # ``anonymization_terms``, SEULE source de vérité) aux VALEURS de cellules,
    # après matérialisation SQL (donc sur les données fraîches) et après le
    # scrub RLS des labels. Les en-têtes de colonnes restent en clair (= schéma,
    # cohérent avec le contrat ``Pseudonymizer.anonymize``). Fail-closed : si un
    # terme configuré ne peut être chargé, ``anonymize_tabs_for_export`` lève —
    # on préfère échouer l'export que livrer un fichier où une vraie valeur
    # fuiterait silencieusement.
    if anonymize:
        from app.services.anonymization.export_filter import anonymize_tabs_for_export_meta

        _anon = await anonymize_tabs_for_export_meta(getattr(user, "id", None), enriched_tabs)
        enriched_tabs = _anon["tabs"]
        if _anon["term_count"] == 0:
            # Anti fausse-impression de sécurité : l'utilisateur a coché
            # « Anonymisé » mais n'a aucun terme configuré → le fichier est
            # identique au clair. On le signale (warning non-bloquant, surfacé
            # en toast côté client comme les autres warnings d'export).
            warnings.append(
                "Export anonymisé demandé mais aucun terme n'est configuré sur "
                "/data/privacy — le fichier est identique à un export en clair."
            )

    # Étape 2 : construction du XLSX (sync) hors event loop pour ne pas
    # bloquer Tornado pendant openpyxl.save() qui peut prendre 1-3s sur
    # 100k lignes.
    content, build_stats = await asyncio.to_thread(_build_xlsx_sync, enriched_tabs)

    stats = {
        "tabs_count": len(enriched_tabs),
        "total_rows": build_stats["total_rows"],
        "sql_re_executed": sql_re_executed,
        "sql_skipped": sql_skipped,
        "detail_sheets_count": build_stats["detail_sheets_count"],
        "truncated_tabs": truncated_tabs,
        "anonymized": bool(anonymize),
    }
    return {"content": content, "stats": stats, "warnings": warnings}


async def materialize_workbook_sql_tabs(
    tabs: List[Dict[str, Any]],
    user: Any,
    *,
    max_rows_per_tab: Optional[int] = None,
    sql_timeout_s: Optional[int] = None,
    executor: Optional[QueryExecutor] = None,
    rls_source: str = _RLS_SOURCE,
    logger_prefix: str = "materialize_workbook_sql_tabs",
) -> Dict[str, Any]:
    """Re-exécute les SQL des SQL tabs et retourne les onglets enrichis.

    Extrait depuis ``build_iris_xlsx`` pour pouvoir être réutilisé par les
    autres consommateurs de classeur backend (copilot bridge, PDF report,
    email/CSV automation). Le contrat est identique : RLS appliquée via
    ``QueryExecutor.execute(user=, rls_source=)``, fail-closed sur
    ``DataAccessDeniedError`` (snapshot frontend conservé), warning + skip
    sur erreur SQL générique.

    Args:
        tabs: liste d'onglets ``serialize()`` (chaque tab : dict avec ``sql``,
            ``columns``, ``rows``, ``label`` …)
        user: utilisateur authentifié (forwardé à ``QueryExecutor`` pour la
            RLS). ``None`` accepté pour les contextes système (ex: jobs
            cron) — le QueryExecutor refusera fail-closed si l'enforcer
            l'exige.
        max_rows_per_tab: cap par onglet. ``None`` (défaut depuis le
            2026-06-10) = cap admin ``DatabaseConnection.max_rows``, résolu
            par ``QueryExecutor.execute`` (doctrine no-double-cap — l'ancien
            défaut 100k hardcodé écrasait la config admin dans les 2 sens)
        sql_timeout_s: timeout par requête. ``None`` (défaut) → timeout admin
            (``connector.timeout``, SSoT). Override explicite possible.
        executor: injectable pour tests
        rls_source: identifiant RLS pour audit (different par caller)
        logger_prefix: préfixe pour les logs (different par caller)

    Returns:
        dict avec :
            - ``tabs`` (List[Dict]) : onglets en array-of-arrays, rows
              hydratées depuis SQL si ``sql`` présent, sinon snapshot
              d'entrée préservé
            - ``warnings`` (list[str]) : messages user-facing (RLS denied,
              SQL error)
            - ``sql_re_executed`` (int) : nombre de SQL effectivement
              ré-exécutés
            - ``sql_skipped`` (list[dict]) : détails par skip
            - ``truncated_tabs`` (list[str]) : labels des onglets dont la
              ré-exec a hit le cap
    """
    if executor is None:
        executor = get_query_executor()

    enriched_tabs: List[Dict[str, Any]] = []
    warnings: List[str] = []
    sql_re_executed = 0
    sql_skipped: List[Dict[str, Any]] = []
    truncated_tabs: List[str] = []

    for idx, tab in enumerate(tabs):
        if not isinstance(tab, dict):
            warnings.append(f"Onglet #{idx} ignoré (format invalide)")
            continue

        sql_text = (tab.get("sql") or "").strip()
        label = tab.get("label") or f"Onglet {idx + 1}"
        normalized = _normalize_tab_to_array_format(tab)
        re_exec_failed = False  # #16 — ré-exec SQL KO → fallback snapshot

        if sql_text:
            try:
                result = await executor.execute(
                    sql_text,
                    max_rows=max_rows_per_tab,
                    timeout=sql_timeout_s,
                    user=user,
                    rls_source=rls_source,
                )
                normalized["columns"] = list(result.columns)
                normalized["rows"] = [list(r) for r in result.rows]
                normalized["totalRowCount"] = result.row_count
                normalized["truncated"] = bool(result.truncated)
                normalized["isArrayFormat"] = True
                sql_re_executed += 1
                if result.truncated:
                    truncated_tabs.append(label)
            except DataAccessDeniedError as exc:
                msg = (
                    f"Onglet « {label} » : RLS a refusé la réexécution "
                    f"({exc.user_message or 'accès denied'}). Snapshot d'entrée utilisé."
                )
                warnings.append(msg)
                re_exec_failed = True
                sql_skipped.append({"idx": idx, "label": label, "reason": "rls_denied"})
                logger.info(
                    "%s: RLS denied for tab %d (%s): %s",
                    logger_prefix,
                    idx,
                    label,
                    exc.user_message,
                )
            except Exception as exc:  # noqa: BLE001 — robustesse pipeline
                msg = (
                    f"Onglet « {label} » : réexécution SQL échouée "
                    f"({type(exc).__name__}). Snapshot d'entrée utilisé."
                )
                warnings.append(msg)
                re_exec_failed = True
                sql_skipped.append(
                    {
                        "idx": idx,
                        "label": label,
                        "reason": "sql_error",
                        "detail": str(exc)[:200],
                    }
                )
                logger.warning(
                    "%s: SQL exec failed for tab %d (%s)",
                    logger_prefix,
                    idx,
                    label,
                    exc_info=True,
                )

        # #134 — surfacer TOUTE troncature de l'onglet en warning VISIBLE (sinon
        # l'utilisateur télécharge un fichier incomplet présenté comme complet =
        # données fausses silencieuses). Deux origines :
        #  - ré-exec SQL qui a hit le cap (``normalized["truncated"]`` posé plus haut) ;
        #  - snapshot SOURCE déjà tronqué au chargement (``tab["truncated"]``, cf.
        #    #133/#86 — non recopié par ``_normalize_tab_to_array_format``).
        # Cohérent avec les warnings RLS / SQL-error émis ci-dessus, surfacés
        # côté client via le header ``X-Iris-Export-Warnings`` (iris-grid.js).
        #
        # ⚠️ Pour un onglet SQL, la RÉ-EXEC serveur est AUTORITAIRE (son résultat
        # est dans ``normalized["truncated"]``, posé plus haut) et SUPERSÈDE le
        # flag client ``tab["truncated"]`` — ce dernier ne reflète que le cap
        # d'AFFICHAGE (~500 lignes) du snapshot frontend, justement ce que
        # l'export full corrige. On ne consulte donc le flag source client QUE
        # pour les onglets SANS SQL (snapshot pur) ; sinon un onglet que le
        # serveur vient de re-récupérer COMPLET déclencherait un faux « tronqué ».
        # #16 fix 2026-06-11 — si la ré-exec SQL a ÉCHOUÉ (RLS/erreur), on est
        # retombé sur le snapshot d'entrée → la ré-exec n'est PAS autoritaire, on
        # consulte donc AUSSI le flag source (sinon un snapshot tronqué partirait
        # sans signal de troncature, l'user ne voyant que « ré-exec échouée »).
        # NB (revue adv. 2026-06-11) : le terme ``or re_exec_failed`` est en
        # pratique BELT-AND-SUSPENDERS dans le code actuel — ``_normalize_tab_to_
        # array_format`` fait ``out = dict(tab)`` (recopie ``truncated``) et la
        # ligne 370 ne l'écrase QUE dans la branche SQL réussie ; donc sur échec,
        # ``normalized.get("truncated")`` est DÉJÀ vrai. On garde ce terme comme
        # garde explicite au cas où ``_normalize`` cesserait un jour de recopier
        # ``truncated`` (ne pas le retirer en croyant qu'il est mort).
        source_truncated = (not sql_text or re_exec_failed) and bool(tab.get("truncated"))
        if normalized.get("truncated") or source_truncated:
            normalized["truncated"] = True
            if label not in truncated_tabs:
                truncated_tabs.append(label)
            warnings.append(
                f"Onglet « {label} » : données tronquées — l'export ne contient "
                f"pas toutes les lignes du résultat complet."
            )

        enriched_tabs.append(normalized)

    return {
        "tabs": enriched_tabs,
        "warnings": warnings,
        "sql_re_executed": sql_re_executed,
        "sql_skipped": sql_skipped,
        "truncated_tabs": truncated_tabs,
    }


# ── Internals : normalisation onglet → array-of-arrays ──────────────────────


def _normalize_tab_to_array_format(tab: Dict[str, Any]) -> Dict[str, Any]:
    """Retourne une copie de ``tab`` avec ``rows`` garanti en array-of-arrays
    et ``isArrayFormat`` à True. Préserve toutes les autres clés."""
    columns = tab.get("columns") or []
    if not isinstance(columns, list):
        columns = []
    raw_rows = tab.get("rows") or []
    if not isinstance(raw_rows, list):
        raw_rows = []

    is_array_fmt = bool(tab.get("isArrayFormat"))
    if not is_array_fmt and raw_rows:
        # Format dict : on convertit en array selon ``columns``.
        normalized_rows = [
            [row.get(c) if isinstance(row, dict) else None for c in columns] for row in raw_rows
        ]
    else:
        normalized_rows = [list(r) if isinstance(r, list) else list(r or []) for r in raw_rows]

    out = dict(tab)  # shallow copy — assez parce qu'on remplace les clés muables
    out["columns"] = list(columns)
    out["rows"] = normalized_rows
    out["isArrayFormat"] = True
    return out


# ── Internals : construction openpyxl (sync) ────────────────────────────────


def _build_xlsx_sync(tabs: List[Dict[str, Any]]) -> Tuple[bytes, Dict[str, Any]]:
    """Construit le workbook openpyxl. Bloquant : appelé via
    ``asyncio.to_thread``. Retourne (bytes, stats)."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.hyperlink import Hyperlink
    except ImportError as exc:
        raise ValueError("openpyxl indisponible côté serveur — export XLSX impossible") from exc

    wb = openpyxl.Workbook()
    # Supprime la feuille par défaut "Sheet" — on créera les nôtres.
    default = wb.active
    if default is not None:
        wb.remove(default)

    # **#144** — Cadenasser les métadonnées workbook pour mode invisible.
    # openpyxl positionne par défaut ``creator="openpyxl"`` et
    # ``last_modified_by="openpyxl"`` ; pas de leak intrinsèque mais
    # défensif : si un upstream injecte un user/admin name un jour, on
    # garantit qu'aucun nom métier ne fuit ici. Champs ``title``,
    # ``subject``, ``description``, ``keywords``, ``category`` mis à
    # vide pour éviter toute fuite via ``wb.properties.*`` exposé par
    # Excel dans Fichier > Propriétés.
    from app.services.branding import get_company_name

    _company = get_company_name()
    try:
        wb.properties.creator = _company
        wb.properties.last_modified_by = _company
        wb.properties.title = ""
        wb.properties.subject = ""
        wb.properties.description = ""
        wb.properties.keywords = ""
        wb.properties.category = ""
    except Exception:
        # Compatibilité versions openpyxl très anciennes — ne pas casser
        # l'export pour un champ properties non assignable.
        pass

    used_sheet_names: Set[str] = set()
    main_ws_by_idx: Dict[int, Any] = {}

    header_font = Font(bold=True, color=_HEADER_FONT_RGB)
    header_fill = PatternFill(
        start_color=_HEADER_FILL_RGB, end_color=_HEADER_FILL_RGB, fill_type="solid"
    )
    hyperlink_font = Font(color=_HYPERLINK_FONT_RGB, underline="single")
    header_align = Alignment(horizontal="left", vertical="center")

    total_rows = 0
    detail_sheets_count = 0

    # Pass 0 : créer les feuilles principales et réserver leurs noms (pour
    # éviter qu'un nom de feuille de détail entre en collision avec un main).
    for idx, tab in enumerate(tabs):
        label = tab.get("label") or f"Onglet {idx + 1}"
        name = _unique_sheet_name(_sanitize_sheet_name(label), used_sheet_names)
        ws = wb.create_sheet(title=name)
        main_ws_by_idx[idx] = ws

    # Pass 1 : écrire les feuilles principales (sans hyperliens encore — on
    # les ajoute en pass 3, après avoir créé les feuilles de détail).
    pending_hyperlinks: List[Dict[str, Any]] = []  # {ws, cell_ref, target_sheet, tooltip}

    for idx, tab in enumerate(tabs):
        ws = main_ws_by_idx[idx]
        rows_written = _populate_sheet(
            ws,
            tab.get("columns") or [],
            tab.get("rows") or [],
            tab.get("merges") or [],
            header_font,
            header_fill,
            header_align,
        )
        total_rows += rows_written

    # Pass 2 : pour chaque cellDetail, soit on a des rows cachés (le user a
    # déjà drill-down côté frontend) — on les utilise tels quels —, soit on
    # reconstruit par filtrage du tab source (équivalent du
    # ``_reconstructDetailRowsFromMatch`` côté JS, mais avec la donnée
    # fraîche du re-exec).
    for idx, tab in enumerate(tabs):
        cell_details = tab.get("cellDetails") or {}
        if not isinstance(cell_details, dict) or not cell_details:
            continue

        for cell_key, detail in list(cell_details.items()):
            if detail_sheets_count >= _MAX_DETAIL_SHEETS_PER_EXPORT:
                logger.warning(
                    "iris_export_xlsx_full: cap _MAX_DETAIL_SHEETS_PER_EXPORT atteint, "
                    "feuilles de détail restantes ignorées"
                )
                break
            if not isinstance(detail, dict):
                continue
            try:
                row_idx_str, col_idx_str = cell_key.split(",", 1)
                row_idx = int(row_idx_str.strip())
                col_idx = int(col_idx_str.strip())
            except (ValueError, AttributeError):
                continue

            effective = _resolve_detail_rows(detail, idx, tabs)
            if effective is None:
                continue
            d_columns, d_rows, total_match = effective
            if not d_columns:
                continue

            # Création de la feuille de détail.
            tab_label = tab.get("label") or f"Onglet {idx + 1}"
            raw_name = f"D-{_sanitize_sheet_name(tab_label)}-{row_idx}_{col_idx}"
            base_name = _sanitize_sheet_name(raw_name)  # re-cap à 31
            detail_sheet_name = _unique_sheet_name(base_name, used_sheet_names)

            d_ws = wb.create_sheet(title=detail_sheet_name)
            d_ws.sheet_properties.tabColor = _DETAIL_TAB_COLOR

            # Row 1 = lien "← Retour" vers la cellule parente. Sans ce lien,
            # depuis une feuille de détail Excel l'utilisateur doit chercher
            # la feuille d'origine dans la barre d'onglets (impraticable
            # quand il y a 50+ feuilles de détail dans un classeur).
            ws_main_for_return = main_ws_by_idx[idx]
            parent_sheet_name = ws_main_for_return.title
            try:
                parent_cell_coord = ws_main_for_return.cell(
                    row=row_idx + 2, column=col_idx + 1
                ).coordinate
            except (ValueError, IndexError):
                parent_cell_coord = "A1"
            return_quoted = "'" + parent_sheet_name.replace("'", "''") + "'"
            return_location = f"{return_quoted}!{parent_cell_coord}"
            return_text = f'← Retour à "{parent_sheet_name}" ' f"cellule {parent_cell_coord}"
            return_cell = d_ws.cell(row=1, column=1, value=return_text)
            return_cell.hyperlink = Hyperlink(
                ref=return_cell.coordinate,
                location=return_location,
                display=return_text,
                tooltip=(
                    f"Retour vers la cellule {parent_cell_coord} "
                    f"de la feuille {parent_sheet_name}"
                ),
            )
            return_cell.font = hyperlink_font

            _populate_sheet(
                d_ws,
                d_columns,
                d_rows,
                [],  # pas de merges sur les feuilles de détail
                header_font,
                header_fill,
                header_align,
                header_row=2,  # row 1 = lien Retour, header en row 2
            )
            detail_sheets_count += 1

            # Hyperlien sur la cellule du tab principal — coordonnée Excel
            # = (row_idx + 2, col_idx + 1) (ligne 1 = header, openpyxl 1-based).
            ws_main = main_ws_by_idx[idx]
            try:
                cell = ws_main.cell(row=row_idx + 2, column=col_idx + 1)
            except (ValueError, IndexError):
                continue
            # IMPORTANT : pour un hyperlien INTERNE (vers une autre feuille du
            # même classeur), il faut utiliser ``Hyperlink(location=...)`` et
            # PAS ``cell.hyperlink = "#'Sheet'!A1"`` (assignation string).
            # L'assignation string crée un Hyperlink avec ``target=...`` qui
            # devient une relation External dans le XLSX → Excel essaie de
            # suivre comme une URL → "problèmes avec le contenu" + récupération
            # qui PERD les hyperliens. Avec ``location``, c'est traité comme
            # une référence interne native qu'Excel suit normalement.
            #
            # Format de location : ``'NomFeuille'!A1`` (apostrophes obligatoires
            # si le nom a un espace ou un caractère spécial ; doublées si le
            # nom lui-même contient un apostrophe).
            quoted_name = "'" + detail_sheet_name.replace("'", "''") + "'"
            # A2 = header de la feuille de détail (A1 = lien Retour).
            internal_location = f"{quoted_name}!A2"
            # NOTE : on N'AJOUTE PAS de cell.comment ici. openpyxl écrit
            # les comments via un fichier VML drawing (xl/drawings/*.vml)
            # avec des préfixes namespace non-standard (ns0:, ns1:, ns2:
            # au lieu de v:, o:, x:). Excel desktop refuse parfois le
            # fichier entier ("possibly corrupt or unsupported") à cause
            # de cette VML.
            #
            # Bug openpyxl connu : les comments produisent un VML que les
            # versions strictes d'Excel rejettent. L'hyperlien lui-même
            # affiche un tooltip natif Excel quand on survole — le comment
            # était juste de la redondance.
            tooltip_count = total_match if total_match is not None else len(d_rows)
            displayed_count = len(d_rows)
            display_text = f"{tooltip_count} ligne{'s' if tooltip_count > 1 else ''}"
            if displayed_count < tooltip_count:
                display_text += f" ({displayed_count} aff.)"
            cell.hyperlink = Hyperlink(
                ref=cell.coordinate,
                location=internal_location,
                # ``display`` apparaît dans la barre d'état Excel + tooltip
                # natif au survol — remplace le besoin d'un Comment séparé.
                display=display_text,
                tooltip=f"Voir {display_text} de détail",
            )
            cell.font = hyperlink_font

    # Sauvegarde avec filet de sécurité : si lxml rejette malgré la
    # sanitization upstream (cas exotique non couvert), on fait une passe
    # de rinçage ASCII-only sur TOUTES les cellules string et on retente
    # une fois. Mieux vaut un export "moche" qu'un export raté.
    buf = io.BytesIO()
    try:
        wb.save(buf)
    except Exception as exc:  # noqa: BLE001 — defensive
        logger.warning(
            "iris_xlsx: wb.save() rejet lxml/openpyxl (%s) — "
            "rinçage ASCII-only et nouvelle tentative",
            type(exc).__name__,
        )
        _aggressive_ascii_sanitize_all_cells(wb)
        buf = io.BytesIO()
        wb.save(buf)  # si ça échoue encore, on laisse remonter
    return buf.getvalue(), {
        "total_rows": total_rows,
        "detail_sheets_count": detail_sheets_count,
    }


def _aggressive_ascii_sanitize_all_cells(wb: Any) -> None:
    """Réécrit toutes les cellules ``str`` du workbook en ASCII printable
    pur (+\t \n \r). Appelé en dernier recours quand lxml rejette le
    workbook au save() malgré la sanitization standard."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    cleaned = "".join(
                        c for c in v if (0x20 <= ord(c) <= 0x7E) or c in ("\t", "\n", "\r")
                    )
                    if cleaned != v:
                        cell.value = cleaned


# ── Internals : population d'une feuille ────────────────────────────────────


def _populate_sheet(
    ws: Any,
    columns: List[str],
    rows: List[List[Any]],
    merges: List[Dict[str, int]],
    header_font: Any,
    header_fill: Any,
    header_align: Any,
    header_row: int = 1,
) -> int:
    """Écrit headers + rows + merges + freeze sur ``ws``. Retourne le
    nombre de rows de données écrites (hors header).

    ``header_row`` permet de décaler le header (par exemple si une row 1
    est déjà occupée par un lien "Retour" sur une feuille de détail).
    Data rows commencent à ``header_row + 1``. Freeze panes placé sous
    la dernière row "structure" (header + offset).
    """
    if not columns:
        return 0

    # #30 (revue adv.) — SSoT formule-safe partagée CSV/XLSX (output_safety).
    from app.utils.output_safety import excel_safe_cell

    # Headers à la row ``header_row``.
    for col_idx, col_name in enumerate(columns, start=1):
        # Strip illégaux XLSX sur le nom aussi (rare mais robuste).
        header_text = _XLSX_ILLEGAL_CHARS_RE.sub("", str(col_name)) if col_name is not None else ""
        # #30 fix 2026-06-11 (revue adv. → SSoT) — injection de formule (CWE-1236)
        # sur l'EN-TÊTE, neutralisée via ``excel_safe_cell`` (SSoT partagée CSV/
        # XLSX, output_safety.py) : préfixe `'` sur TOUS les préfixes dangereux
        # (=, +, -, @, tab, CR — cf. CSV_FORMULA_PREFIXES), types natifs préservés.
        # Avant : `data_type='f'→'s'` ne couvrait QUE `=` (openpyxl ne classe 'f'
        # que pour `=`) → +,-,@,tab,CR restaient injectables. Un alias SQL
        # `AS "=cmd|..."` / nom de colonne Sage / pseudonyme ne peut plus devenir
        # une formule exécutable chez un destinataire EXTERNE.
        c = ws.cell(row=header_row, column=col_idx, value=excel_safe_cell(header_text))
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # Data rows à partir de ``header_row + 1``.
    data_start_row = header_row + 1
    n_cols = len(columns)
    rows_written = 0
    for r_idx, row in enumerate(rows, start=data_start_row):
        # row peut être array OU dict — _normalize_tab_to_array_format a
        # déjà aligné pour les onglets ; mais on reste défensif pour les
        # détails reconstruits.
        if isinstance(row, dict):
            values = [row.get(columns[c]) for c in range(n_cols)]
        else:
            try:
                values = list(row)[:n_cols]
            except TypeError:
                values = []
            # Pad si row plus courte que columns.
            if len(values) < n_cols:
                values = values + [None] * (n_cols - len(values))

        for col_idx, val in enumerate(values, start=1):
            coerced = _coerce_cell_value(val)
            try:
                # #30 (revue adv. → SSoT) — excel_safe_cell neutralise l'injection
                # de formule (=,+,-,@,tab,CR) sur les chaînes ET préserve les types
                # natifs (nombres/dates restent sommables/triables). Remplace
                # l'ancien backstop `data_type='f'→'s'` qui ne couvrait que `=`.
                cell = ws.cell(row=r_idx, column=col_idx, value=excel_safe_cell(coerced))
            except Exception:  # noqa: BLE001 — defensive: ne JAMAIS crash l'export
                # Filet de sécurité : si openpyxl ou lxml rejette malgré la
                # sanitization (surrogate orphelin, encodage partiel raté,
                # cas exotique non couvert par notre regex), on remplace par
                # une représentation ASCII safe au lieu de planter le build.
                fallback = _safe_ascii_fallback(coerced)
                try:
                    cell = ws.cell(row=r_idx, column=col_idx, value=excel_safe_cell(fallback))
                except Exception:  # noqa: BLE001
                    cell = ws.cell(row=r_idx, column=col_idx, value="[non exportable]")

            # IMPORTANT : openpyxl interprète AUTOMATIQUEMENT toute string
            # commençant par ``=`` comme une formule (data_type='f'). Si
            # cette "formule" est en fait du texte invalide (ex: champ Sage
            # qui contient ``=signed by John``), Excel à l'ouverture lance
            # la "Validation et réparation de niveau fichier" et SUPPRIME
            # ces enregistrements ("Enregistrements supprimés: Formule").
            #
            # Pour les exports de données — où on veut écrire le texte tel
            # quel, jamais le faire évaluer comme formule — on force
            # data_type='s'. La cellule finit en ``<c t="inlineStr">...``
            # qu'Excel affiche comme du texte sans message d'avertissement.
            if cell.data_type == "f":
                cell.data_type = "s"
        rows_written += 1

    # Merges — 4 entiers 0-based (r1, c1, r2, c2) → coordonnées Excel
    # 1-based + offset structure (toutes les rows avant data_start_row).
    for m in merges or []:
        try:
            r1 = int(m.get("r1", 0)) + data_start_row
            r2 = int(m.get("r2", 0)) + data_start_row
            c1 = int(m.get("c1", 0)) + 1
            c2 = int(m.get("c2", 0)) + 1
        except (ValueError, TypeError):
            continue
        if r1 < data_start_row or c1 < 1 or r2 < r1 or c2 < c1:
            continue
        if r2 > data_start_row + len(rows) - 1 or c2 > n_cols:
            continue
        try:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        except (ValueError, Exception):  # noqa: BLE001 — defensive merge
            logger.debug("merge ignorée: %s", m)

    # Freeze toutes les rows "structure" (retour + header) — première data row
    # devient la 1ʳᵉ row scrollable.
    ws.freeze_panes = f"A{data_start_row}"

    # Largeurs FIXES par type — pas d'auto-fit (une cellule longue
    # n'élargit pas toute la colonne).
    _set_fixed_column_widths(ws, columns, rows[:50])

    return rows_written


def _set_fixed_column_widths(ws: Any, columns: List[str], sample_rows: List[Any]) -> None:
    """Largeur fixe par colonne, basée sur le type uniforme dans le sample.

    Règle : si TOUTES les valeurs non-null du sample sont du même type
    (numérique OU date), on applique la largeur dédiée (14 ou 16). Au
    moindre mélange ou s'il y a une string, fallback sur 20. Une string
    qui ressemble à un nombre reste considérée comme texte (Excel
    n'imposera pas de format numérique sur ces cellules — voulu)."""
    from datetime import date, datetime, time

    from openpyxl.utils import get_column_letter

    for col_idx in range(1, len(columns) + 1):
        candidate = None  # None → aucune valeur vue ; sinon "number" | "date" | "default"
        for row in sample_rows:
            if isinstance(row, list):
                val = row[col_idx - 1] if col_idx - 1 < len(row) else None
            elif isinstance(row, dict):
                val = row.get(columns[col_idx - 1])
            else:
                val = None
            if val is None or val == "":
                continue
            if isinstance(val, bool):
                this_kind = "default"
            elif isinstance(val, (int, float)):
                this_kind = "number"
            elif isinstance(val, (datetime, date, time)):
                this_kind = "date"
            else:
                this_kind = "default"
            if candidate is None:
                candidate = this_kind
            elif candidate != this_kind:
                candidate = "default"
                break

        if candidate == "number":
            width = _WIDTH_NUMBER
        elif candidate == "date":
            width = _WIDTH_DATE
        else:
            width = _WIDTH_DEFAULT
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ── Internals : reconstruction des rows de détail ───────────────────────────


def _resolve_detail_rows(
    detail: Dict[str, Any],
    cur_tab_idx: int,
    all_tabs: List[Dict[str, Any]],
) -> Optional[Tuple[List[str], List[List[Any]], int]]:
    """Pour un cellDetail, retourne ``(columns, rows_capped, total_matching)``
    ou ``None`` si rien à exporter.

    Priorité :
    1. Si ``detail.rows`` existe et non-vide → utilisation directe (snapshot
       frontend, déjà fetché par l'utilisateur).
    2. Sinon, si ``detail.match`` existe → reconstruction par filtrage du
       tab source (mêmes règles que le builder JS côté frontend, miroir).
    """
    cached_rows = detail.get("rows")
    cached_cols = detail.get("columns")
    if (
        isinstance(cached_rows, list)
        and cached_rows
        and isinstance(cached_cols, list)
        and cached_cols
    ):
        # Normaliser en array-of-arrays, capper à _MAX_DETAIL_ROWS.
        rows_array = []
        for r in cached_rows[:_MAX_DETAIL_ROWS]:
            if isinstance(r, dict):
                rows_array.append([r.get(c) for c in cached_cols])
            elif isinstance(r, list):
                rows_array.append(list(r))
            else:
                continue
        total = detail.get("row_count")
        if not isinstance(total, int) or total < len(rows_array):
            total = len(rows_array)
        return list(cached_cols), rows_array, total

    match = detail.get("match")
    if not isinstance(match, dict) or not match:
        return None

    match_exclude = detail.get("match_exclude")
    if not isinstance(match_exclude, dict):
        match_exclude = {}

    src_idx = _pick_source_tab(detail, cur_tab_idx, all_tabs, match)
    if src_idx is None:
        return None
    src_tab = all_tabs[src_idx]
    src_cols = src_tab.get("columns") or []
    src_rows = src_tab.get("rows") or []
    if not src_cols or not src_rows:
        return None

    col_pos = {name: i for i, name in enumerate(src_cols)}

    filtered: List[List[Any]] = []
    total_matching = 0
    match_keys = list(match.keys())
    for row in src_rows:
        if not isinstance(row, list):
            continue
        ok = True
        for k in match_keys:
            pos = col_pos.get(k)
            if pos is None:
                ok = False
                break
            row_val = row[pos] if pos < len(row) else None
            if not _match_value(row_val, match[k]):
                ok = False
                break
        if not ok:
            continue
        # match_exclude : aucune clé ne doit matcher.
        for k_ex, v_ex in match_exclude.items():
            pos = col_pos.get(k_ex)
            if pos is None:
                continue
            row_val = row[pos] if pos < len(row) else None
            if _match_excluded(row_val, v_ex):
                ok = False
                break
        if not ok:
            continue
        total_matching += 1
        if len(filtered) < _MAX_DETAIL_ROWS:
            # On copie SEULEMENT les colonnes du source (toutes — donne un
            # vrai breakdown).
            filtered.append([row[i] if i < len(row) else None for i in range(len(src_cols))])

    if not filtered:
        return None
    return list(src_cols), filtered, total_matching


def _pick_source_tab(
    detail: Dict[str, Any],
    cur_tab_idx: int,
    all_tabs: List[Dict[str, Any]],
    match: Dict[str, Any],
) -> Optional[int]:
    """Retourne l'index du tab source à utiliser pour reconstruire les
    rows, ou ``None`` si aucun candidat valide.

    Stratégie :
    1. Si ``detail.source_tab_index`` est valide, dans les bornes, ≠ cur,
       et que ses ``columns`` couvrent les clés de ``match`` → on l'utilise.
    2. Sinon auto-detect : parmi les autres onglets, prendre celui qui est
       le plus AGRÉGÉ sur les match_keys — c.à.d. celui dont les
       match_keys représentent la plus grande fraction de ses colonnes.

       Pourquoi cette heuristique : si match = {exercice, mois, expert,
       statistique} et qu'un classeur contient :

       - Tab "Base SAGE" (247 colonnes brutes — 1 row = 1 facture)
       - Tab "Mois CA" (5 colonnes — 1 row = 1 combo agrégé)

       Le filtre de match cherche des rows AGRÉGÉES qui matchent les 4
       dimensions. Sur Base SAGE, AUCUNE row ne matche exactement
       (chaque row a aussi numFacture, dateFacture, compteClient, etc.
       et n'est pas alignée sur ces 4 dims). Sur Mois CA, les rows
       SONT par construction des combos de ces 4 dims → matchera.

       Métrique : ``fraction = len(match_keys) / len(tab.columns)``
       — plus c'est grand, plus le tab est "fait pour" ces dimensions.
       Tie-break : on prend le tab avec le plus PETIT nombre total
       de colonnes (= le plus agrégé).
    """
    match_keys = set(match.keys())
    n_tabs = len(all_tabs)

    hint = detail.get("source_tab_index")
    if isinstance(hint, int) and 0 <= hint < n_tabs and hint != cur_tab_idx:
        cols = all_tabs[hint].get("columns") or []
        if isinstance(cols, list) and match_keys.issubset(set(cols)):
            rows = all_tabs[hint].get("rows") or []
            if rows:
                return hint

    best_idx: Optional[int] = None
    best_fraction = -1.0
    best_total_cols = 10**9
    for ti in range(n_tabs):
        if ti == cur_tab_idx:
            continue
        cols = all_tabs[ti].get("columns") or []
        if not isinstance(cols, list) or not cols:
            continue
        if not match_keys.issubset(set(cols)):
            continue
        rows = all_tabs[ti].get("rows") or []
        if not rows:
            continue
        total_cols = len(cols)
        fraction = len(match_keys) / total_cols
        # Préfère fraction max ; tie → moins de colonnes = plus agrégé.
        if fraction > best_fraction or (fraction == best_fraction and total_cols < best_total_cols):
            best_fraction = fraction
            best_total_cols = total_cols
            best_idx = ti
    return best_idx


def _match_value(row_val: Any, match_val: Any) -> bool:
    """Compare une valeur de ligne à une valeur de match (scalaire OU liste)."""
    if isinstance(match_val, list):
        if not match_val:
            return True  # liste vide = pas de filtre (comportement backend)
        return any(_loose_eq(row_val, v) for v in match_val)
    if match_val is None:
        return row_val is None
    return _loose_eq(row_val, match_val)


def _match_excluded(row_val: Any, exclude_val: Any) -> bool:
    """True si la valeur de ligne fait partie des exclusions."""
    if isinstance(exclude_val, list):
        return any(_loose_eq(row_val, v) for v in exclude_val)
    if exclude_val is None:
        return False
    return _loose_eq(row_val, exclude_val)


def _loose_eq(a: Any, b: Any) -> bool:
    """Comparaison lâche : strict d'abord, puis numérique (gère "2023"=="2023"
    et "1 005,76"==1005.76), puis string. Miroir de ``_looseEq`` côté JS."""
    if a == b:
        return True
    if a is None or b is None:
        return False
    # Numeric coercion (gère format FR + types mélangés).
    try:
        na = (
            a
            if isinstance(a, (int, float))
            else float(str(a).replace(" ", "").replace(" ", "").replace(" ", "").replace(",", "."))
        )
        nb = (
            b
            if isinstance(b, (int, float))
            else float(str(b).replace(" ", "").replace(" ", "").replace(" ", "").replace(",", "."))
        )
        if na == nb:
            return True
    except (ValueError, TypeError):
        pass
    return str(a) == str(b)


# ── Internals : helpers nom de feuille ──────────────────────────────────────


def _sanitize_sheet_name(name: str) -> str:
    """Caractères interdits XLSX (/, \\, ?, *, [, ], :) → underscore. Cap 31 char.
    Strip aussi les caractères de contrôle (XML 1.0 illégaux)."""
    if not isinstance(name, str):
        name = str(name) if name is not None else ""
    cleaned = _XLSX_ILLEGAL_CHARS_RE.sub("", name)
    cleaned = _INVALID_SHEET_NAME_CHARS.sub("_", cleaned).strip("'")
    if not cleaned:
        cleaned = "Onglet"
    return cleaned[:31]


def _unique_sheet_name(base: str, used: Set[str]) -> str:
    """Retourne un nom unique en suffixant ~2, ~3… si collision après truncation."""
    if base not in used:
        used.add(base)
        return base
    for n in range(2, 1000):
        suffix = f"~{n}"
        room = 31 - len(suffix)
        candidate = base[:room] + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
    fallback = f"D~{len(used)}"[:31]
    used.add(fallback)
    return fallback


# ── Internals : coercion des valeurs cellule ────────────────────────────────


def _safe_ascii_fallback(val: Any) -> str:
    """Dernier recours : représentation ASCII printable d'une valeur qui a
    fait planter openpyxl/lxml malgré la sanitization standard. Garde
    uniquement les chars 0x20-0x7E + tab/newline/CR. Si le résultat est
    vide, retourne ``[binaire]``."""
    if val is None:
        return ""
    try:
        s = str(val)
    except Exception:  # noqa: BLE001
        return "[non exportable]"
    safe = "".join(c for c in s if (0x20 <= ord(c) <= 0x7E) or c in ("\t", "\n", "\r"))
    return safe if safe else "[binaire]"


def _coerce_cell_value(val: Any) -> Any:
    """Coerce une valeur arbitraire vers un type que openpyxl écrit
    proprement. Strip les caractères de contrôle XML-illégaux sur les
    strings (sinon openpyxl lève ``IllegalCharacterError`` — cas
    d'une colonne BLOB Sage castée en string par pyodbc).
    None/numbers/bool/datetimes passent tels quels."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        return _XLSX_ILLEGAL_CHARS_RE.sub("", val)
    # datetime / date / time → openpyxl gère nativement.
    try:
        from datetime import date, datetime, time

        if isinstance(val, (datetime, date, time)):
            return val
    except ImportError:  # pragma: no cover
        pass
    # bytes → décode best-effort + strip illégaux.
    if isinstance(val, (bytes, bytearray)):
        try:
            decoded = val.decode("utf-8", errors="replace")
            return _XLSX_ILLEGAL_CHARS_RE.sub("", decoded)
        except Exception:  # noqa: BLE001
            return ""
    # Fallback : str() puis strip.
    try:
        return _XLSX_ILLEGAL_CHARS_RE.sub("", str(val))
    except Exception:  # noqa: BLE001
        return ""
