"""Générateurs de rapport pour le pipeline mono-step legacy.

D3 phase 4 (cycle 22) : extrait des méthodes ``_generate_csv``,
``_generate_excel``, ``_generate_pdf`` de :class:`AutomationExecutor`.
Ces helpers sont utilisés UNIQUEMENT par le pipeline legacy
``_run_pipeline`` (autos sans steps ni edges) pour produire le fichier
final dans le format demandé.

Chaque générateur a un fallback CSV en cas d'erreur (openpyxl/ReportLab
manquant ou parsing échoué) — le pipeline ne doit pas crasher pour un
problème de format.

Toutes les fonctions sont synchrones — le caller utilise
``asyncio.to_thread`` quand il a besoin de non-bloquant.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from app.services.automation.workbook_export import reserve_unique_output_path
from app.services.export.csv_export import to_csv_bytes
from app.utils.logger import get_logger
from app.utils.output_safety import excel_safe_cell

logger = get_logger(__name__)

# #134 / #18f verdict #42 — marqueur de troncature pour les fichiers CSV/Excel
# envoyés en pièce jointe (parité avec la bannière PDF, pdf_generator). Sans
# lui, un collaborateur calcule des totaux sur un fichier amputé au cap admin
# SANS aucun signal (donnée fausse silencieuse côté destinataire).
TRUNCATION_MARKER = (
    "⚠ RÉSULTATS TRONQUÉS AU CAP ADMIN — DONNÉES PARTIELLES "
    "(les lignes au-delà de la limite ne sont PAS dans ce fichier)"
)


def generate_csv(
    output_path: Path, results: List[Dict[str, Any]], *, truncated: bool = False
) -> Path:
    """Génère un fichier CSV UTF-8 BOM (compat Excel double-clic).

    Délègue au service unifié :func:`app.services.export.csv_export.to_csv_bytes`
    pour la sanitisation OWASP-CSV-Injection et le BOM — replicate du
    fallback ``Aucun résultat`` historique via ``empty_placeholder``.

    Args:
        truncated: #134 — la source a-t-elle été tronquée au cap admin ? Si oui
            on append une ligne-marqueur explicite en fin de fichier (1re colonne)
            pour que le destinataire voie que les données sont partielles.

    Returns:
        Le chemin écrit (identique à ``output_path`` — utile pour les
        fallbacks qui changent le suffixe).
    """
    augmented = results
    if truncated:
        if results:
            # Ligne-marqueur : warning en 1re colonne, vide ailleurs (les colonnes
            # numériques restent vides → non sommées par erreur côté destinataire).
            fieldnames0 = list(results[0].keys())
            marker_row = {fn: "" for fn in fieldnames0}
            marker_row[fieldnames0[0]] = TRUNCATION_MARKER
            augmented = [*results, marker_row]
        else:
            # Empty + truncated : un filtre post-fetch (RLS/anonymisation) a pu
            # vider `results` alors que la source était cappée. Émettre quand
            # même le marqueur — sinon fichier « Aucun résultat » trompeur alors
            # que des lignes existent au-delà du cap (donnée fausse silencieuse).
            augmented = [{"avertissement": TRUNCATION_MARKER}]
    fieldnames = list(augmented[0].keys()) if augmented else None
    payload = to_csv_bytes(augmented, columns=fieldnames, empty_placeholder="Aucun résultat")
    output_path.write_bytes(payload)
    return output_path


def _reserve_and_write_csv_fallback(
    output_path: Path, results: List[Dict[str, Any]], *, truncated: bool = False
) -> Path:
    """Réserve atomiquement le chemin CSV de fallback (O_EXCL, anti-collision
    avec un autre node) PUIS écrit le CSV — en nettoyant le fichier 0-byte
    réservé si l'écriture lève (#64 / A7-F9-résidu).

    Sans ce nettoyage, un échec de ``generate_csv`` APRÈS la réservation (ex
    ``to_csv_bytes`` sur une donnée hostile, ou ``write_bytes`` OSError) laisse
    un orphelin 0-byte dans ``automation_reports/`` jusqu'au TTL cleanup. On
    libère immédiatement. L'exception est re-levée (comportement inchangé : le
    caller décide). SSoT partagé avec ``executor._safe_output_path``.

    ``truncated`` est forwardé pour que le fallback CSV (openpyxl absent)
    conserve le marqueur de troncature au lieu de le perdre silencieusement.
    """
    fallback_path = reserve_unique_output_path(output_path.with_suffix(".csv"))
    try:
        return generate_csv(fallback_path, results, truncated=truncated)
    except Exception:
        # Le writer a levé après la réservation atomique → libérer le 0-byte
        # orphelin tout de suite (pas dans 30 j). ``missing_ok`` : robuste si le
        # fichier a déjà disparu (race cleanup / write partiel supprimé).
        try:
            fallback_path.unlink(missing_ok=True)
        except OSError:
            logger.warning(
                "report_generator: cleanup orphelin fallback CSV échoué (%s)",
                fallback_path,
                exc_info=True,
            )
        raise


def generate_excel(
    output_path: Path, results: List[Dict[str, Any]], *, truncated: bool = False
) -> Path:
    """Génère un fichier Excel .xlsx.

    Fallback CSV si openpyxl indisponible. Le path retourné peut donc
    avoir un suffixe différent de l'entrée — le caller doit utiliser
    la valeur retournée.

    Args:
        truncated: #134 — si la source a été tronquée au cap admin, on écrit
            une ligne-marqueur (rouge gras) en fin de feuille.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résultats"

        if not results:
            ws["A1"] = "Aucun résultat"
            # Empty + truncated : émettre quand même le marqueur (cf. generate_csv).
            if truncated:
                warn_cell = ws.cell(row=2, column=1, value=TRUNCATION_MARKER)
                warn_cell.font = Font(bold=True, color="CC0000")
            wb.save(output_path)
            return output_path

        # Headers — excel_safe_cell : un nom de colonne commençant par =,+,-,@
        # serait sinon évalué comme formule à l'ouverture (CSV/formula-injection
        # OWASP). Cohérent avec le path CSV (to_csv_bytes) et dashboard export.
        headers = list(results[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=excel_safe_cell(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Données — excel_safe_cell neutralise les formules sur les chaînes mais
        # PRÉSERVE les types natifs (int/float/datetime/None) : un nombre reste
        # sommable/triable côté destinataire, on ne corrige pas la sécurité au
        # prix d'une régression d'usage.
        for row_idx, row_data in enumerate(results, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=excel_safe_cell(row_data.get(header)))

        # #134 — ligne-marqueur de troncature (rouge gras), après les données.
        # Notre constante ne commence pas par =/+/-/@ → pas de risque formule.
        if truncated:
            warn_cell = ws.cell(row=len(results) + 2, column=1, value=TRUNCATION_MARKER)
            warn_cell.font = Font(bold=True, color="CC0000")

        # Ajuster largeur colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output_path)
        return output_path

    except ImportError:
        logger.warning("openpyxl non installe, generation CSV a la place")
        # A7-F9 (+ #64) — réserve atomique du path CSV de fallback PUIS écrit,
        # avec nettoyage immédiat du 0-byte orphelin si l'écriture lève.
        return _reserve_and_write_csv_fallback(output_path, results, truncated=truncated)
    except Exception:
        # Catch-all volontaire (cohérent avec generate_pdf) : une donnée
        # hostile — caractère XML-illégal d'un BLOB Sage casté en string →
        # openpyxl IllegalCharacterError, qui n'est PAS un ImportError — ou
        # toute autre erreur openpyxl ne doit JAMAIS faire crasher le run
        # d'automation (sinon : rapport jamais produit, email jamais envoyé).
        # Fallback CSV : robuste à ces données ET déjà protégé contre
        # l'injection de formule (csv_safe_cell via to_csv_bytes).
        logger.error("Erreur generation Excel, fallback CSV", exc_info=True)
        # A7-F9 (+ #64) — réserve atomique du path CSV de fallback PUIS écrit,
        # avec nettoyage immédiat du 0-byte orphelin si l'écriture lève.
        return _reserve_and_write_csv_fallback(output_path, results, truncated=truncated)


def generate_pdf(
    output_path: Path,
    automation: Any,
    results: List[Dict[str, Any]],
    *,
    truncated: bool = False,
) -> Path:
    """Génère un PDF professionnel via :class:`PDFGenerator`.

    Fallback CSV si ReportLab indisponible OU erreur runtime
    (SQLAlchemyError/OSError/ConnectionError/ValueError). Le path
    retourné peut donc avoir un suffixe différent — le caller doit
    utiliser la valeur retournée.

    Args:
        automation: Object avec attributs ``.name``, ``.description``
            (typiquement :class:`Automation`).
        truncated: #133 — si ``True``, la source SQL a été tronquée au cap
            admin. On préfixe la description du rapport d'une bannière pour
            que l'utilisateur sache que les totaux/agrégats portent sur un
            sous-ensemble (sinon données fausses silencieuses).
    """
    try:
        from app.services.reporting.pdf_generator import PDFGenerator

        # Créer le générateur. ``company_name=None`` → lit la valeur
        # configurée via ``branding.get_company_name`` (anti-hardcode).
        pdf_gen = PDFGenerator(
            company_name=None,
            logo_path=None,  # Logo configurable via config.company.logo_path
        )

        # Métadonnées
        from app.services.branding import get_company_name

        metadata = {
            "author": f"{get_company_name()} Automation",
            "subject": f"Rapport automatisé: {automation.name}",
        }

        # #133 — bannière de troncature SOURCE en tête de description (lossless,
        # ne corrompt pas les colonnes du tableau, contrairement à une ligne
        # injectée dans un CSV/Excel). Sans nom de table/colonne (générique).
        description = automation.description or ""
        if truncated:
            banner = (
                "⚠ Données tronquées à la source : le cap de lignes (configuré par "
                "l'administrateur) a été atteint. Les totaux et agrégats de ce rapport "
                "portent sur un sous-ensemble des données, pas sur leur intégralité."
            )
            description = f"{banner}\n\n{description}" if description else banner

        # Générer le PDF
        pdf_gen.generate_from_query_result(
            output_path=output_path,
            title=automation.name,
            results=results,
            metadata=metadata,
            description=description,
        )
        return output_path

    except ImportError:
        logger.warning("ReportLab non installe, generation CSV a la place")
        # A7-F9 (+ #64) — réserve atomique du path CSV de fallback PUIS écrit,
        # avec nettoyage immédiat du 0-byte orphelin si l'écriture lève.
        return _reserve_and_write_csv_fallback(output_path, results, truncated=truncated)
    except Exception:
        # Catch-all volontaire : le pipeline ne doit JAMAIS crasher pour
        # un échec de format. Fallback CSV. Cohérent avec la version
        # historique (capturait SQLAlchemyError, OSError, ConnectionError,
        # ValueError — on étend pour défense en profondeur).
        logger.error("Erreur generation PDF", exc_info=True)
        # A7-F9 (+ #64) — réserve atomique du path CSV de fallback PUIS écrit,
        # avec nettoyage immédiat du 0-byte orphelin si l'écriture lève.
        return _reserve_and_write_csv_fallback(output_path, results, truncated=truncated)


__all__ = (
    "generate_csv",
    "generate_excel",
    "generate_pdf",
)
# Note : ``TRUNCATION_MARKER`` est importable directement (pas dans ``__all__``
# qui ne régit que ``import *`` et que le test garde restreint aux fonctions).
