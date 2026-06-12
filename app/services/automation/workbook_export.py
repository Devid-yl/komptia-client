"""Helpers d'export de classeurs vers fichiers (Excel, CSV, ZIP).

D3 phase 1 (cycle 19) : extrait des helpers ``_write_*`` et
``_sanitize_filename_hint`` de :class:`AutomationExecutor` pour réduire
la god class à 2700+ lignes. Aucun de ces helpers n'a besoin du contexte
d'exécution (pas de session DB, pas d'async, pas de SMTPClient) — ce
sont des conversions in-memory → disque.

Ces helpers respectent les conventions Komptia :
* Noms d'onglets Excel : tronqués à 31 chars (limite native), caractères
  interdits remplacés par ``_``, unicité forcée par suffixe ``_N``.
* CSV : encodage ``utf-8-sig`` (BOM) pour compat Excel double-clic.
* ZIP CSV : un fichier ``.csv`` par onglet, fallback ``(vide)`` si tab
  sans rows.
* Filenames : sanitisation alphanum + ``_-`` uniquement, limité à 80 chars.

Toutes les fonctions sont synchrones — l'appelant utilise
``asyncio.to_thread`` quand il a besoin de non-bloquant.
"""

from __future__ import annotations

import csv
import io
import os
import uuid
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional


# Cluster-J (J3) 2026-05-26 — cap total cells par export pour prévenir
# les ZIP-bombs (1000 tabs × 100k cells = 5 GB ZIP → disk full ou OOM
# zip-writing). 10M cells (~ 80 MB CSV brut) couvre largement les
# exports légitimes ; au-delà l'admin doit augmenter via env var (cas
# d'usage rare : reporting fiscal annuel multi-entités). Configurable
# pour respecter doctrine ``feedback_no_double_cap`` — pas de hidden cap.
EXPORT_MAX_TOTAL_CELLS = int(
    os.environ.get("KOMPTIA_EXPORT_MAX_TOTAL_CELLS", "10000000")
)


class ExportTooLargeError(ValueError):
    """Cluster-J (J3) 2026-05-26 — Levée quand le total cells d'un
    export dépasse ``EXPORT_MAX_TOTAL_CELLS``. Distincte de ValueError
    générique pour que le UI puisse afficher un message dédié
    ("Export trop volumineux : N cells > M. Réduisez le scope ou
    contactez l'admin pour augmenter la limite.")."""


def _count_total_cells(tabs: List[Dict[str, Any]]) -> int:
    """Compte le nombre total de cellules (rows × columns) sur tous les tabs."""
    total = 0
    for tab in tabs:
        rows = tab.get("rows") or []
        cols = tab.get("columns") or []
        n_rows = len(rows) if isinstance(rows, list) else 0
        n_cols = len(cols) if isinstance(cols, list) else 0
        # Si columns manquant mais rows présent, déduire via 1er row dict
        if n_cols == 0 and n_rows > 0 and isinstance(rows[0], dict):
            n_cols = len(rows[0])
        total += n_rows * n_cols
    return total


def _check_export_size(tabs: List[Dict[str, Any]], format_label: str) -> None:
    """Cluster-J (J3) — Vérifie que l'export ne dépasse pas le cap total
    cells. Raise ``ExportTooLargeError`` si dépassement.

    Doit être appelé en début de chaque write_csv_*/write_excel_*.
    """
    total = _count_total_cells(tabs)
    if total > EXPORT_MAX_TOTAL_CELLS:
        raise ExportTooLargeError(
            f"Export {format_label} trop volumineux : {total:,} cellules "
            f"> {EXPORT_MAX_TOTAL_CELLS:,}. Réduisez le scope (moins d'onglets "
            "ou filtre SQL plus strict) ou demandez à l'admin d'augmenter "
            "``KOMPTIA_EXPORT_MAX_TOTAL_CELLS``."
        )


def sanitize_filename_hint(hint: Optional[str], default: str) -> str:
    """Sanitise un hint de nom de fichier.

    Retire les path-separators, les caractères de contrôle et limite
    à 80 caractères. Fallback ``default`` si le hint est vide ou
    invalide après sanitisation.

    Args:
        hint: Suggestion fournie par le caller (ex: nom d'automation).
        default: Valeur de repli si hint vide ou invalide.

    Returns:
        Une chaîne sûre pour usage en filename (alphanum + ``_-``).
    """
    if not hint:
        return default
    cleaned = "".join(c for c in hint.strip() if c.isalnum() or c in ("_", "-", " ")).strip()
    cleaned = cleaned.replace(" ", "_")[:80]
    return cleaned or default


def reserve_unique_output_path(path: Path) -> Path:
    """Réserve ATOMIQUEMENT ``path`` via création exclusive ``O_EXCL`` (A7-F9).

    SSoT anti-collision des fichiers d'artefacts d'automation. Deux écritures
    visant le même nom écrasaient la première EN SILENCE → la StepExecution qui
    la référençait pointait vers le contenu de la seconde (corruption en aval :
    pièce jointe mail, save datastore). Cas réels :
      - report / PDF IA : timestamp à la SECONDE → 2 reports même seconde ;
      - export workbook : nom = filename_hint SANS timestamp → 2 exports même
        hint ;
      - fallback CSV (excel/pdf KO) : ``output_path.with_suffix('.csv')`` peut
        viser un ``.csv`` déjà réservé par un autre node.

    La création exclusive est un **syscall atomique** : elle sérialise across
    threads ET process (les automations schedulées tournent dans des threads
    APScheduler distincts avec des boucles asyncio séparées — la garantie n'est
    donc PAS « la boucle mono-thread » mais l'atomicité d'``O_EXCL``). Si le nom
    est déjà pris, on bascule sur un nom unique et on reboucle. Le fichier vide
    réservé est ensuite écrasé par le writer (tous tronquent : ``write_bytes`` /
    ``open(...,'w')`` / ``wb.save`` ; aucun mode ``'x'``/append). Les erreurs FS
    non-``EEXIST`` (ENOSPC, EACCES, ...) propagent VOLONTAIREMENT — le run doit
    échouer franchement plutôt que boucler.
    """
    candidate = path
    while True:
        try:
            fd = os.open(candidate, os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o600)
            os.close(fd)
            return candidate
        except FileExistsError:
            candidate = path.with_name(f"{path.stem}_{uuid.uuid4().hex[:8]}{path.suffix}")


def _tab_to_dict_rows(tab: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Délégation au helper centralisé ``workbook_service.tab_to_dict_rows``.

    Internalisé ici pour éviter une dépendance directe depuis le caller
    sur deux modules différents.
    """
    from app.services.automation.workbook_service import tab_to_dict_rows

    return tab_to_dict_rows(tab)


def write_excel_multi_tabs(output_path: Path, tabs: List[Dict[str, Any]]) -> None:
    """Écrit un .xlsx avec un onglet par tab.

    Les noms d'onglets sont tronqués à 31 chars (limite Excel) et les
    caractères interdits (``[]:*?/\\``) sont remplacés par ``_``.
    Unicité forcée par suffixe ``_N``.
    """
    _check_export_size(tabs, "Excel")
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()
    for idx, tab in enumerate(tabs):
        raw_label = str(tab.get("label") or f"Onglet {idx + 1}")
        safe_label = (
            "".join(c if c not in "[]:*?/\\" else "_" for c in raw_label)[:31] or f"Onglet{idx + 1}"
        )
        # Excel exige des noms d'onglets uniques.
        unique_label = safe_label
        n = 1
        while unique_label in used_names:
            suffix = f"_{n}"
            unique_label = safe_label[: 31 - len(suffix)] + suffix
            n += 1
        used_names.add(unique_label)

        ws = wb.create_sheet(title=unique_label)
        dict_rows = _tab_to_dict_rows(tab)
        if not dict_rows:
            ws["A1"] = "(vide)"
            continue
        headers = list(tab.get("columns") or list(dict_rows[0].keys()))
        # #30 fix 2026-06-11 (revue adv. → SSoT) — injection de formule (CWE-1236)
        # neutralisée via ``excel_safe_cell`` (SSoT output_safety, partagée CSV/
        # XLSX) : préfixe `'` sur =,+,-,@,tab,CR (pas seulement `=`), types natifs
        # préservés. Avant : `data_type='f'→'s'` ne couvrait que `=`.
        from app.utils.output_safety import excel_safe_cell

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=excel_safe_cell(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for row_idx, r in enumerate(dict_rows, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=excel_safe_cell(r.get(header)))

    if not wb.sheetnames:
        wb.create_sheet(title="Vide")
    wb.save(output_path)


def write_csv_single_tab(output_path: Path, tab: Dict[str, Any]) -> None:
    """Écrit un seul tab en .csv UTF-8 BOM (compat Excel double-clic).

    Cluster-I 2026-05-26 — délègue à ``csv_export.to_csv_bytes`` (SSoT)
    qui applique la sanitisation CWE-1236 (OWASP A05 CSV formula
    injection : ``=cmd|...``, ``@SUM(...)``, ``-1+2``, ``+...`` préfixés
    par ``'`` pour neutraliser l'exécution Excel au double-clic).
    Préserve le format historique : UTF-8 BOM, fallback ``(vide)``.
    """
    from app.services.export.csv_export import to_csv_bytes

    dict_rows = _tab_to_dict_rows(tab)
    columns = tab.get("columns") or None  # `[]` → None pour activer placeholder
    if not columns and dict_rows:
        columns = list(dict_rows[0].keys())
    payload = to_csv_bytes(
        dict_rows,
        columns=columns,
        empty_placeholder="(vide)",
    )
    output_path.write_bytes(payload)


def write_csv_zip(output_path: Path, tabs: List[Dict[str, Any]]) -> None:
    """Écrit un .zip contenant un .csv par tab (multi-tab CSV).

    Cluster-I 2026-05-26 — chaque entry du ZIP est sérialisée via
    ``csv_export.to_csv_bytes`` (SSoT CWE-1236 sanitisation). Avant :
    ``csv.DictWriter`` bare laissait passer ``=HYPERLINK("attack")``
    exécuté au double-clic Excel sur le poste destinataire.

    Cluster-J (J3) 2026-05-26 — cap total cells (10M défaut) pour
    prévenir les ZIP-bombs (1000 tabs × 100k cells = 5 GB ZIP).
    """
    _check_export_size(tabs, "ZIP CSV")
    from app.services.export.csv_export import to_csv_bytes

    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        used: set[str] = set()
        for idx, tab in enumerate(tabs):
            raw_label = str(tab.get("label") or f"onglet_{idx + 1}")
            safe = (
                "".join(c if c.isalnum() or c in ("_", "-") else "_" for c in raw_label)[:60]
                or f"onglet_{idx + 1}"
            )
            csv_name = f"{safe}.csv"
            n = 1
            while csv_name in used:
                csv_name = f"{safe}_{n}.csv"
                n += 1
            used.add(csv_name)

            dict_rows = _tab_to_dict_rows(tab)
            columns = tab.get("columns") or None  # `[]` → None pour activer placeholder
            if not columns and dict_rows:
                columns = list(dict_rows[0].keys())
            payload = to_csv_bytes(
                dict_rows,
                columns=columns,
                empty_placeholder="(vide)",
            )
            zf.writestr(csv_name, payload)


__all__ = (
    "sanitize_filename_hint",
    "write_excel_multi_tabs",
    "write_csv_single_tab",
    "write_csv_zip",
)
# Cluster-J (J3) 2026-05-26 — exposés sans `__all__` (le test
# `TestNoExecutorClassDependency` asserte que `__all__` ne contient
# que des fonctions pures). Restent importables via attribute access
# pour les call-sites qui veulent catcher ExportTooLargeError ou
# overrider EXPORT_MAX_TOTAL_CELLS dans les tests.
