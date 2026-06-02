"""Charge des onglets Excel (.xlsx, .xls) en conservant les cellules fusionnées.

Préserve l'info des merged cells Excel (`ws.merged_cells.ranges`) pour que
la result area puisse les afficher à l'identique.
"""

from __future__ import annotations

import datetime as _dt
import logging
from pathlib import Path
from typing import Any, List, Optional

import tornado.web

logger = logging.getLogger(__name__)

DEFAULT_MAX_ROWS = 50000
_OPENPYXL_MISSING_MSG = (
    "La librairie openpyxl est requise pour l'import Excel. "
    "Installez-la via `pip install openpyxl>=3.1.0`."
)


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        logger.error("openpyxl indisponible: %s", e)
        raise tornado.web.HTTPError(500, _OPENPYXL_MISSING_MSG)


def _cell_value_safe(value: Any) -> Any:
    """Convertit une valeur de cellule en type serializable JSON."""
    if value is None:
        return None
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return value.isoformat()
    if isinstance(value, _dt.timedelta):
        return value.total_seconds()
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def _xlsx_col_letter(idx: int) -> str:
    """Convertit 0-based column index en lettre Excel (A, B, ..., Z, AA, ...)."""
    idx += 1
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return letters


def list_excel_sheets(path: Path) -> List[str]:
    """Retourne la liste des noms d'onglets d'un fichier Excel."""
    _require_openpyxl()
    import openpyxl

    if not path.exists() or not path.is_file():
        raise tornado.web.HTTPError(404, "Fichier Excel introuvable")

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Impossible d'ouvrir le fichier Excel %s: %s", path, e)
        raise tornado.web.HTTPError(400, "Fichier Excel invalide")

    try:
        return list(wb.sheetnames)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def has_merge_on_first_row(merges: List[dict]) -> bool:
    """Vrai si au moins une fusion touche la première ligne (r1 == 0)."""
    for m in merges:
        if m.get("r1", -1) == 0:
            return True
    return False


def _dedupe_columns(columns: List[str]) -> List[str]:
    """Dédoublonne les colonnes en ajoutant suffixe (2), (3), ... si collision."""
    seen: dict[str, int] = {}
    result = []
    for col in columns:
        base = col if col else ""
        if base not in seen:
            seen[base] = 1
            result.append(base)
        else:
            seen[base] += 1
            result.append(f"{base} ({seen[base]})")
    return result


def load_excel_sheet(
    path: Path,
    sheet_name: Optional[str] = None,
    max_rows: int = DEFAULT_MAX_ROWS,
    first_row_as_header: bool = False,
) -> dict:
    """Charge un onglet Excel et renvoie un dict compatible format classeur Komptia.

    Returns:
        {
            "sheet_name": str,         # nom effectif de l'onglet chargé
            "columns": list[str],      # headers (soit ligne 1, soit "Colonne A"...)
            "rows": list[list],        # données (sans la ligne header si appliquée)
            "row_count": int,          # len(rows)
            "truncated": bool,         # True si > max_rows
            "merges": list[dict],      # [{r1, c1, r2, c2}, ...] en coords 0-based
            "first_row_as_header_effective": bool,
        }

    La présence de merges sur la ligne 0 désactive automatiquement
    `first_row_as_header` (sinon les headers seraient incohérents).
    """
    _require_openpyxl()
    import openpyxl

    if not path.exists() or not path.is_file():
        raise tornado.web.HTTPError(404, "Fichier Excel introuvable")

    if max_rows <= 0:
        max_rows = DEFAULT_MAX_ROWS

    # read_only=False is required to access ws.merged_cells.ranges
    # (ReadOnlyWorksheet does not expose it). Memory usage is bounded by
    # _MAX_EXCEL_FILE_SIZE enforced in the handler upstream.
    try:
        wb = openpyxl.load_workbook(str(path), read_only=False, data_only=True)
    except Exception as e:
        logger.warning("Impossible d'ouvrir le fichier Excel %s: %s", path, e)
        raise tornado.web.HTTPError(400, "Fichier Excel invalide")

    try:
        if sheet_name is None:
            ws = wb.active
            if ws is None:
                if not wb.sheetnames:
                    raise tornado.web.HTTPError(400, "Aucun onglet dans le fichier Excel")
                ws = wb[wb.sheetnames[0]]
            effective_sheet_name = ws.title
        else:
            if sheet_name not in wb.sheetnames:
                raise tornado.web.HTTPError(404, f"Onglet '{sheet_name}' introuvable")
            ws = wb[sheet_name]
            effective_sheet_name = sheet_name

        max_col = ws.max_column or 0

        raw_merges: List[dict] = []
        try:
            for mr in ws.merged_cells.ranges:
                r1 = int(mr.min_row) - 1
                r2 = int(mr.max_row) - 1
                c1 = int(mr.min_col) - 1
                c2 = int(mr.max_col) - 1
                if r1 < 0 or c1 < 0 or r2 < r1 or c2 < c1:
                    continue
                raw_merges.append({"r1": r1, "c1": c1, "r2": r2, "c2": c2})
        except Exception as e:
            logger.warning("Lecture merged_cells échouée sur %s: %s", effective_sheet_name, e)
            raw_merges = []

        merge_on_header = has_merge_on_first_row(raw_merges)
        first_row_as_header_effective = bool(first_row_as_header) and not merge_on_header

        raw_rows: List[List[Any]] = []
        truncated = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows + (1 if first_row_as_header_effective else 0):
                truncated = True
                break
            row_values = list(row)
            if len(row_values) < max_col:
                row_values.extend([None] * (max_col - len(row_values)))
            elif len(row_values) > max_col:
                row_values = row_values[:max_col]
            row_values = [_cell_value_safe(v) for v in row_values]
            raw_rows.append(row_values)

        if first_row_as_header_effective and raw_rows:
            header_row = raw_rows[0]
            columns_raw = [
                (
                    str(v).strip()
                    if v is not None and str(v).strip()
                    else f"Colonne {_xlsx_col_letter(i)}"
                )
                for i, v in enumerate(header_row)
            ]
            columns = _dedupe_columns(columns_raw)
            data_rows = raw_rows[1:]
            shifted_merges: List[dict] = []
            for m in raw_merges:
                if m["r1"] == 0:
                    continue
                shifted_merges.append(
                    {
                        "r1": m["r1"] - 1,
                        "c1": m["c1"],
                        "r2": m["r2"] - 1,
                        "c2": m["c2"],
                    }
                )
            merges = shifted_merges
        else:
            columns = [f"Colonne {_xlsx_col_letter(i)}" for i in range(max_col)]
            data_rows = raw_rows
            merges = list(raw_merges)

        effective_max = max_rows
        if len(data_rows) > effective_max:
            data_rows = data_rows[:effective_max]
            truncated = True

        if data_rows:
            last_r = len(data_rows) - 1
            clamped_merges: List[dict] = []
            for m in merges:
                if m["r1"] > last_r:
                    continue
                new_m = dict(m)
                if new_m["r2"] > last_r:
                    new_m["r2"] = last_r
                if new_m["r2"] == new_m["r1"] and new_m["c2"] == new_m["c1"]:
                    continue
                clamped_merges.append(new_m)
            merges = clamped_merges
        else:
            merges = []

        return {
            "sheet_name": effective_sheet_name,
            "columns": columns,
            "rows": data_rows,
            "row_count": len(data_rows),
            "truncated": truncated,
            "merges": merges,
            "first_row_as_header_effective": first_row_as_header_effective,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass
