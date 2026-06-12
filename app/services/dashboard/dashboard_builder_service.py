"""
Service pour le Dashboard Builder — CRUD dashboards + widgets + récupération de données.

Gère la création, modification, suppression des tableaux de bord personnalisables
et l'exécution des sources de données (métriques prédéfinies ou SQL).
"""

import csv
import io
import logging
import re
from datetime import datetime, timedelta
from typing import Any, Optional

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core import clock
from app.utils.output_safety import csv_safe_cell, excel_safe_cell

logger = logging.getLogger(__name__)


# ── Métriques prédéfinies disponibles ──────────────────────────────────────
# Chaque métrique est un callable async(session, user_id, period_days) -> dict
# Le dict retourné contient les données adaptées au type de widget.

AVAILABLE_METRICS = {
    "total_searches": {
        "label": "Recherches totales",
        "description": "Nombre total de recherches Iris",
        "supports": ["kpi", "chart"],
    },
    "success_rate": {
        "label": "Taux de succès",
        "description": "Pourcentage de recherches réussies",
        "supports": ["kpi"],
    },
    "daily_searches": {
        "label": "Recherches par jour",
        "description": "Nombre de recherches par jour sur la période",
        "supports": ["chart"],
    },
    "execution_status": {
        "label": "Statut des exécutions",
        "description": "Répartition succès/échec des automatisations",
        "supports": ["chart"],
    },
    "active_automations": {
        "label": "Automatisations actives",
        "description": "Nombre d'automatisations actuellement actives",
        "supports": ["kpi"],
    },
    "total_reports": {
        "label": "Rapports générés",
        "description": "Nombre total de rapports",
        "supports": ["kpi"],
    },
    "total_contacts": {
        "label": "Contacts",
        "description": "Nombre de contacts enregistrés",
        "supports": ["kpi"],
    },
    "recent_searches": {
        "label": "Dernières recherches",
        "description": "Liste des recherches récentes",
        "supports": ["table"],
    },
    # ``top_users`` et ``avg_response_time`` ont été retirés en 2026-05-09
    # (BLOCKING #4-5 review consolidée). Ces métriques agrégeaient
    # ``SearchHistory`` / ``AIPerformanceLog`` sans filtre ``user_id`` →
    # tout user voyait les stats de tous les autres users. Komptia étant
    # multi-user sans partage cross-user, c'était un leak silencieux. Le
    # monitoring admin global passe par /dashboard (page distincte gatée
    # par rôle), pas par les dashboards configurables /dashboards.
}


async def _scrub_dashboard_data_for_user(
    data: dict,
    user: Any,
) -> dict:
    """**#140** — Scrubbe un dict de dashboard (déjà sérialisé par
    :meth:`Dashboard.to_dict` et enrichi de ``widgets``/``filters``)
    pour remplacer les noms de tables/colonnes interdits par ``[…]``.

    Cas couverts (mode invisible rétroactif) :

    - ``data["name"]`` — titre du dashboard, saisi par l'admin.
    - ``data["description"]`` — description, peut mentionner une table.
    - ``data["template_description"]`` — même rationale (templates).
    - Pour chaque widget de ``data["widgets"]`` :
        - ``widget["title"]`` — affiché en gras au-dessus du widget.
        - ``widget["data_source_config"]`` (selon ``data_source_type``):
            - ``"sql"`` → scrub ``query`` (le SQL est lu dans l'éditeur).
            - ``"static"`` → scrub ``title`` et ``content`` (texte libre).

    **Stratégie** : ``user=None``, admin, ou sans restrictions → no-op
    (court-circuit). Sinon délégation à :func:`scrub_text_for_user`
    pour chaque champ string identifié.

    **Fail-safe** : si une scrub crash, on retourne la valeur originale
    (cf. comportement de :func:`scrub_text_for_user`). On ne casse
    JAMAIS le rendu du dashboard pour un hoquet BDD.

    **CRITICAL — Isolation des dicts imbriqués** : ``Widget.to_dict()``
    retourne ``self.data_source_config`` PAR RÉFÉRENCE (c'est un
    ``Mapped[dict]`` SQLAlchemy partagé avec l'instance ORM en
    mémoire). Sans deep-copy, mes mutations ici corrompraient
    l'instance ORM (et potentiellement la BDD via
    ``MutableDict.changed()`` côté SQLAlchemy 2.0). On force donc une
    deep-copy des dicts imbriqués AVANT mutation. Pas de copie du
    wrapper data — il est déjà éphémère (créé par ``to_dict``).

    **Returns** : le même wrapper dict (mutation in-place + retour pour
    chaînage). Les inner dicts sont REMPLACÉS par des copies.
    """
    if user is None or not isinstance(data, dict):
        return data
    import copy

    from app.services.data_access.error_messages import scrub_text_for_user

    async def _maybe_scrub(text: Any) -> Any:
        if not isinstance(text, str) or not text:
            return text
        try:
            return await scrub_text_for_user(text, user, context_label="dashboard_persisted")
        except Exception:
            return text

    for top_field in ("name", "description", "template_description"):
        if top_field in data:
            data[top_field] = await _maybe_scrub(data.get(top_field))

    widgets = data.get("widgets")
    if isinstance(widgets, list):
        for idx, widget in enumerate(widgets):
            if not isinstance(widget, dict):
                continue
            widget["title"] = await _maybe_scrub(widget.get("title"))
            dsc_original = widget.get("data_source_config")
            if isinstance(dsc_original, dict):
                # Deep-copy avant mutation pour ne PAS corrompre
                # l'instance ORM partagée.
                dsc = copy.deepcopy(dsc_original)
                dst = widget.get("data_source_type")
                if dst == "sql":
                    dsc["query"] = await _maybe_scrub(dsc.get("query"))
                    # Onglets SQL additionnels (feature menu [+] « Requête
                    # SQL ») : leur requête/titre sont aussi lisibles dans la
                    # config côté client → même scrub que la requête principale
                    # (cohérence mode-invisible rétroactif). dsc est déjà
                    # deep-copié, mutation des sous-dicts sûre.
                    extra_tabs = dsc.get("extra_tabs")
                    if isinstance(extra_tabs, list):
                        for _t in extra_tabs:
                            if isinstance(_t, dict):
                                _t["label"] = await _maybe_scrub(_t.get("label"))
                                _t["query"] = await _maybe_scrub(_t.get("query"))
                elif dst == "static":
                    dsc["title"] = await _maybe_scrub(dsc.get("title"))
                    dsc["content"] = await _maybe_scrub(dsc.get("content"))
                # metric : pas de free-text user, rien à scrubber.
                widget["data_source_config"] = dsc
    return data


class DashboardBuilderService:
    """Service pour gérer les dashboards personnalisables."""

    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    # ── CRUD Dashboard ─────────────────────────────────────────────────────

    async def list_dashboards(
        self,
        session: AsyncSession,
        user_id: int,
        user: Any = None,
    ) -> list[dict]:
        """Liste les dashboards de l'utilisateur courant.

        Strict scope owner-only — aucun partage cross-user. Chaque user voit
        UNIQUEMENT ses propres dashboards (cf. tâche #29 : suppression du
        partage cross-user).

        **#140** — Si ``user`` est passé (objet User du caller), scrub
        les champs textuels (``name``, ``description``,
        ``template_description``) pour retirer les noms de tables
        denied. Sans ``user``, fallback comportement legacy (aucun
        scrub) pour les callers backend/scheduler qui n'ont pas de
        contexte utilisateur.
        """
        from app.models.dashboard import Dashboard

        stmt = (
            select(Dashboard)
            .where(
                Dashboard.user_id == user_id,
                Dashboard.is_template.is_(False),
            )
            .order_by(Dashboard.updated_at.desc())
        )
        result = await session.execute(stmt)
        dashboards = result.scalars().all()

        out = [d.to_dict() for d in dashboards]
        if user is not None:
            for item in out:
                await _scrub_dashboard_data_for_user(item, user)
        return out

    async def get_dashboard(
        self,
        session: AsyncSession,
        dashboard_id: int,
        user_id: int,
        user: Any = None,
    ) -> Optional[dict]:
        """Récupère un dashboard avec ses widgets (owner only).

        Retourne ``None`` si le dashboard n'existe pas OU s'il n'appartient
        pas au user courant — aucune distinction entre "introuvable" et
        "interdit" pour ne pas leaker l'existence cross-user (fail-closed).

        **#140** — Si ``user`` est passé, scrub les champs textuels
        sensibles (name/description/widget.title/widget.sql/widget.text)
        pour retirer les noms de tables denied. Le user étant le owner
        du dashboard, ce sont SES anciens widgets (créés avant la pose
        de la règle deny) qui leakent les noms désormais interdits.
        Sans ``user``, fallback comportement legacy (aucun scrub) pour
        les callers backend/scheduler.
        """
        from app.models.dashboard import Dashboard

        stmt = (
            select(Dashboard)
            .options(selectinload(Dashboard.widgets))
            .options(selectinload(Dashboard.filters))
            .where(Dashboard.id == dashboard_id, Dashboard.user_id == user_id)
        )
        result = await session.execute(stmt)
        dashboard = result.scalar_one_or_none()

        if not dashboard:
            return None

        data = dashboard.to_dict()
        data["widgets"] = [w.to_dict() for w in dashboard.widgets]
        data["filters"] = [f.to_dict() for f in dashboard.filters]
        if user is not None:
            await _scrub_dashboard_data_for_user(data, user)
        return data

    async def create_dashboard(
        self, session: AsyncSession, user_id: int, name: str, description: str = ""
    ) -> dict:
        """Crée un nouveau dashboard."""
        from app.models.dashboard import Dashboard

        if not name or not name.strip():
            raise ValueError("Le nom du dashboard est obligatoire.")

        name = name.strip()[:200]

        dashboard = Dashboard(
            name=name,
            description=description.strip()[:1000] if description else None,
            user_id=user_id,
        )
        session.add(dashboard)
        await session.flush()

        data = dashboard.to_dict()
        await session.commit()

        logger.info("Dashboard créé: id=%s, user=%s", data["id"], user_id)
        # Hook auto-scan anonymization (fire-and-forget) — alimente
        # /data/privacy sans attendre "Scanner mes données".
        from app.services.anonymization.auto_scan import schedule_target_rescan

        schedule_target_rescan(user_id, "dashboard", int(data["id"]))
        return data

    async def update_dashboard(
        self,
        session: AsyncSession,
        dashboard_id: int,
        user_id: int,
        updates: dict,
    ) -> Optional[dict]:
        """Met à jour un dashboard (nom, description) — owner only."""
        from app.models.dashboard import Dashboard

        stmt = select(Dashboard).where(Dashboard.id == dashboard_id, Dashboard.user_id == user_id)
        result = await session.execute(stmt)
        dashboard = result.scalar_one_or_none()

        if not dashboard:
            return None

        allowed_fields = {"name", "description"}
        for field, value in updates.items():
            if field not in allowed_fields:
                continue
            if field == "name":
                if not value or not str(value).strip():
                    raise ValueError("Le nom ne peut pas être vide.")
                value = str(value).strip()[:200]
            elif field == "description":
                value = str(value).strip()[:1000] if value else None
            setattr(dashboard, field, value)

        data = dashboard.to_dict()
        await session.commit()

        logger.info("Dashboard mis à jour: id=%s", dashboard_id)
        from app.services.anonymization.auto_scan import schedule_target_rescan

        schedule_target_rescan(user_id, "dashboard", dashboard_id)
        return data

    async def delete_dashboard(
        self, session: AsyncSession, dashboard_id: int, user_id: int
    ) -> bool:
        """Supprime un dashboard (cascade supprime les widgets)."""
        from app.models.dashboard import Dashboard

        stmt = select(Dashboard).where(Dashboard.id == dashboard_id, Dashboard.user_id == user_id)
        result = await session.execute(stmt)
        dashboard = result.scalar_one_or_none()

        if not dashboard:
            return False

        await session.delete(dashboard)
        await session.commit()

        # Les widgets partent en cascade BDD (ondelete=CASCADE) sans hook
        # par-widget → nettoyer les classeurs de widgets du dashboard ici
        # (glob par préfixe, best-effort, quota libéré).
        from app.services.dashboard import widget_workbook_store as wbstore

        try:
            await wbstore.delete_dashboard_workbooks(user_id, dashboard_id)
        except Exception:  # noqa: BLE001
            logger.warning(
                "Nettoyage des classeurs widgets du dashboard %s échoué",
                dashboard_id,
                exc_info=True,
            )

        logger.info("Dashboard supprimé: id=%s", dashboard_id)
        return True

    async def clone_dashboard(
        self, session: AsyncSession, dashboard_id: int, user_id: int
    ) -> Optional[dict]:
        """Clone un dashboard avec tous ses widgets (owner only).

        Retourne ``None`` si le dashboard n'existe pas OU s'il n'appartient
        pas au user courant. Aucun clonage cross-user — chaque user voit
        uniquement ses propres dashboards (cf. tâche #29).
        """
        from app.models.dashboard import Dashboard, DashboardWidget

        # Charger l'original avec widgets — strict ownership inclus dans le WHERE
        stmt = (
            select(Dashboard)
            .options(selectinload(Dashboard.widgets))
            .options(selectinload(Dashboard.filters))
            .where(Dashboard.id == dashboard_id, Dashboard.user_id == user_id)
        )
        result = await session.execute(stmt)
        original = result.scalar_one_or_none()

        if not original:
            return None

        # Capturer les données des widgets avant toute opération
        widget_configs = []
        widget_src_ids = []
        for w in original.widgets:
            widget_src_ids.append(w.id)
            widget_configs.append(
                {
                    "title": w.title,
                    "widget_type": w.widget_type,
                    "chart_type": w.chart_type,
                    "data_source_type": w.data_source_type,
                    "data_source_config": w.data_source_config,
                    "col_span": w.col_span,
                    "position_order": w.position_order,
                    "style_config": w.style_config,
                }
            )

        # Créer le clone
        clone = Dashboard(
            name=f"{original.name} (copie)",
            description=original.description,
            user_id=user_id,
        )
        session.add(clone)
        await session.flush()

        # Cloner les widgets (refs gardées : la copie des classeurs de widgets
        # grille a besoin des NOUVEAUX ids, connus après flush).
        cloned_widgets = []
        for src_id, wc in zip(widget_src_ids, widget_configs):
            widget = DashboardWidget(dashboard_id=clone.id, **wc)
            session.add(widget)
            cloned_widgets.append((widget, src_id))

        # Cloner les filtres
        from app.models.dashboard import DashboardFilter

        for f in original.filters:
            filter_clone = DashboardFilter(
                dashboard_id=clone.id,
                parameter_name=f.parameter_name,
                label=f.label,
                filter_type=f.filter_type,
                values_source=f.values_source,
                values_config=f.values_config,
                default_value=f.default_value,
                position_order=f.position_order,
            )
            session.add(filter_clone)

        # Mode classeur : copier le fichier classeur de chaque widget grille
        # vers les ids du clone. Copie impossible (source disparue, quota) →
        # retirer le pointeur du clone (fallback legacy via le miroir config,
        # JAMAIS un fichier partagé entre deux widgets).
        # Limite connue (revue adv. 2026-06-10) : si le commit du clone échoue
        # APRÈS la copie, les fichiers copiés + leur quota restent orphelins
        # (réconciliables via StorageManager.sync_user_storage) — événement
        # rare (retry DB en amont), accepté.
        await session.flush()
        from app.services.dashboard import widget_workbook_store as wbstore

        for widget, src_id in cloned_widgets:
            cfg = widget.data_source_config
            if not (
                widget.widget_type == "grid" and isinstance(cfg, dict) and cfg.get("workbook_file")
            ):
                continue
            new_rel = await wbstore.copy_workbook(
                user_id, dashboard_id, src_id, clone.id, widget.id
            )
            new_cfg = dict(cfg)
            if new_rel:
                new_cfg["workbook_file"] = new_rel
            else:
                new_cfg.pop("workbook_file", None)
            widget.data_source_config = new_cfg

        data = clone.to_dict()
        await session.commit()

        logger.info("Dashboard cloné: original=%s → clone=%s", dashboard_id, data["id"])
        return data

    # ── CRUD Widgets ───────────────────────────────────────────────────────

    async def add_widget(
        self,
        session: AsyncSession,
        dashboard_id: int,
        user_id: int,
        widget_data: dict,
    ) -> Optional[dict]:
        """Ajoute un widget à un dashboard."""
        from app.models.dashboard import Dashboard, DashboardWidget

        # Vérifier ownership du dashboard
        stmt = select(Dashboard).where(Dashboard.id == dashboard_id, Dashboard.user_id == user_id)
        result = await session.execute(stmt)
        dashboard = result.scalar_one_or_none()
        if not dashboard:
            return None

        # Calculer position_order (prochain index)
        count_stmt = (
            select(func.count())
            .select_from(DashboardWidget)
            .where(DashboardWidget.dashboard_id == dashboard_id)
        )
        count_result = await session.execute(count_stmt)
        next_order = count_result.scalar() or 0

        # Normalisation : chart_type="" est équivalent à None (signal "auto")
        chart_type_raw = widget_data.get("chart_type")
        chart_type = chart_type_raw.strip() if isinstance(chart_type_raw, str) else chart_type_raw
        if chart_type in (None, "", "auto"):
            chart_type = None

        widget = DashboardWidget(
            dashboard_id=dashboard_id,
            title=str(widget_data.get("title", "")).strip()[:200],
            widget_type=widget_data.get("widget_type", "chart"),
            chart_type=chart_type,
            data_source_type=widget_data.get("data_source_type", "metric"),
            data_source_config=widget_data.get("data_source_config"),
            col_span=widget_data.get("col_span", 6),
            position_order=widget_data.get("position_order", next_order),
            style_config=widget_data.get("style_config"),
        )

        errors = widget.validate()
        if errors:
            raise ValueError("; ".join(errors))

        session.add(widget)
        await session.flush()
        data = widget.to_dict()
        await session.commit()

        logger.info("Widget ajouté: id=%s, dashboard=%s", data["id"], dashboard_id)
        from app.services.anonymization.auto_scan import schedule_target_rescan

        schedule_target_rescan(user_id, "dashboard", dashboard_id)
        return data

    async def update_widget(
        self,
        session: AsyncSession,
        widget_id: int,
        dashboard_id: int,
        user_id: int,
        updates: dict,
    ) -> Optional[dict]:
        """Met à jour un widget."""
        from app.models.dashboard import Dashboard, DashboardWidget

        # Vérifier ownership via dashboard
        stmt = (
            select(DashboardWidget)
            .join(Dashboard)
            .where(
                DashboardWidget.id == widget_id,
                DashboardWidget.dashboard_id == dashboard_id,
                Dashboard.user_id == user_id,
            )
        )
        result = await session.execute(stmt)
        widget = result.scalar_one_or_none()

        if not widget:
            return None

        allowed = {
            "title",
            "widget_type",
            "chart_type",
            "data_source_type",
            "data_source_config",
            "col_span",
            "position_order",
            "style_config",
        }
        for field, value in updates.items():
            if field not in allowed:
                continue
            if field == "title":
                value = str(value).strip()[:200]
            elif field == "col_span":
                try:
                    value = int(value) if value else 6
                except (ValueError, TypeError):
                    value = 6
            elif field == "position_order":
                try:
                    value = int(value) if value is not None else 0
                except (ValueError, TypeError):
                    value = 0
            elif field == "chart_type":
                # "" / "auto" → None (signal pour inférence automatique)
                if isinstance(value, str):
                    value = value.strip()
                if value in ("", "auto"):
                    value = None
            elif field == "data_source_config":
                # ``workbook_file`` est GÉRÉ PAR LE SERVEUR (mode classeur du
                # widget grille). La modale d'édition envoie un snapshot de
                # config figé au render de la page (data-widget) : s'il ne
                # porte pas la clé (page chargée avant le 1er « Enregistrer »
                # du widget), un PUT légitime (renommage du titre, édition du
                # SQL) DÉTACHERAIT silencieusement le classeur — feuilles
                # manuelles/mise en forme perdues + retour à l'ancienne
                # requête, avec 200 OK partout. On reporte la clé d'office
                # (revue adv. 2026-06-10, finding critique C0).
                existing_cfg = widget.data_source_config
                if (
                    isinstance(value, dict)
                    and isinstance(existing_cfg, dict)
                    and existing_cfg.get("workbook_file")
                    and "workbook_file" not in value
                ):
                    value = dict(value)
                    value["workbook_file"] = existing_cfg["workbook_file"]
            setattr(widget, field, value)

        errors = widget.validate()
        if errors:
            raise ValueError("; ".join(errors))

        # Mode classeur : répercuter une éventuelle édition du SQL (modale du
        # widget) dans le classeur — sinon le chargement (qui lit le classeur,
        # source de vérité) ignorerait silencieusement la modification.
        if widget.widget_type == "grid" and "data_source_config" in updates:
            await self._sync_workbook_sql_sheets(
                user_id, dashboard_id, widget_id, widget.data_source_config or {}
            )

        data = widget.to_dict()
        await session.commit()

        logger.info("Widget mis à jour: id=%s", widget_id)
        from app.services.anonymization.auto_scan import schedule_target_rescan

        schedule_target_rescan(user_id, "dashboard", dashboard_id)
        return data

    async def set_widget_extra_tabs(
        self,
        session: AsyncSession,
        widget_id: int,
        dashboard_id: int,
        user_id: int,
        extra_tabs: Any,
    ) -> Optional[dict]:
        """Remplace la liste des onglets SQL additionnels d'un widget grille.

        Feature menu [+] « Requête SQL ». **Read-modify-write côté SERVEUR** :
        le client n'envoie QUE la liste d'onglets — jamais la config complète.
        On lit la config actuelle du widget et on n'y remplace que la clé
        ``extra_tabs``. Cela évite tout clobber de la requête principale /
        transformation par une copie cliente périmée (le littéral
        ``json_encode`` du template est figé au render ; un autre onglet
        navigateur a pu modifier la config entre-temps) — anti données fausses
        silencieuses.

        Owner-only (join ``Dashboard.user_id``). Retourne le widget mis à jour,
        ou ``None`` si introuvable / non-propriétaire. Lève ``ValueError``
        (→ 400) si la validation échoue (requête non SELECT, trop d'onglets,
        titre vide, etc.).
        """
        from app.models.dashboard import Dashboard, DashboardWidget

        stmt = (
            select(DashboardWidget)
            .join(Dashboard)
            .where(
                DashboardWidget.id == widget_id,
                DashboardWidget.dashboard_id == dashboard_id,
                Dashboard.user_id == user_id,
            )
        )
        result = await session.execute(stmt)
        widget = result.scalar_one_or_none()
        if not widget:
            return None

        if widget.widget_type != "grid":
            raise ValueError("Les onglets SQL ne concernent que les widgets grille.")

        # Fail-closed : un body sans clé ``extra_tabs`` (ou non-liste, ex.
        # ``{"extra_tabs": null}``) ne doit PAS persister ``extra_tabs: None`` —
        # ``validate()`` laisse passer (le bloc est gardé par ``is not None``) →
        # config polluée. On rejette explicitement, comme ``set_widget_sheets``
        # pour ``sheets``.
        if not isinstance(extra_tabs, list):
            raise ValueError("Les onglets SQL (extra_tabs) doivent être une liste.")
        # Normalise : ne garder que {label, query} (drop des clés client
        # parasites). Les entrées de type invalide sont laissées telles quelles
        # → signalées par validate() avec un message clair plutôt que
        # silencieusement ignorées.
        normalized: Any = [
            {"label": t.get("label"), "query": t.get("query")} if isinstance(t, dict) else t
            for t in extra_tabs
        ]

        # setattr d'un NOUVEAU dict (≠ mutation in-place) pour que SQLAlchemy
        # détecte le changement (colonne JSON simple, pas MutableDict) et bumpe
        # updated_at (onupdate) → invalide le cache widget.
        new_config = dict(widget.data_source_config or {})
        if isinstance(normalized, list) and not normalized:
            new_config.pop("extra_tabs", None)  # liste vide → config propre
        else:
            new_config["extra_tabs"] = normalized
        widget.data_source_config = new_config

        errors = widget.validate()
        if errors:
            raise ValueError("; ".join(errors))

        # Mode classeur : garder le classeur (source de vérité exécutée)
        # aligné avec le miroir modifié via cette API.
        await self._sync_workbook_sql_sheets(user_id, dashboard_id, widget_id, new_config)

        data = widget.to_dict()
        await session.commit()

        logger.info("Onglets SQL widget mis à jour: id=%s", widget_id)
        from app.services.anonymization.auto_scan import schedule_target_rescan

        schedule_target_rescan(user_id, "dashboard", dashboard_id)
        return data

    async def set_widget_sheets(
        self,
        session: AsyncSession,
        widget_id: int,
        dashboard_id: int,
        user_id: int,
        sheets: Any,
    ) -> Optional[dict]:
        """Remplace TOUTES les feuilles SQL d'un widget grille depuis une liste
        ORDONNÉE unique (« widget grille piloté par les feuilles »).

        La feuille **0** est la feuille PRINCIPALE : son ``query`` va dans
        ``data_source_config["query"]`` (son ``label`` est IGNORÉ — l'onglet
        principal porte le titre du widget, pas de champ label en base). Les
        feuilles **1..n** vont dans ``extra_tabs`` (``[{label, query}]``).

        **Read-modify-write côté SERVEUR** (comme :meth:`set_widget_extra_tabs`) :
        on ne remplace QUE ``query`` + ``extra_tabs`` dans la config existante
        (``transformation`` / ``render_spec`` / ``drill_column`` préservés) — le
        client n'envoie JAMAIS la config complète (pas de clobber par une copie
        périmée, anti données fausses silencieuses).

        Owner-only (join ``Dashboard.user_id``). Retourne le widget mis à jour,
        ou ``None`` si introuvable / non-propriétaire. Lève ``ValueError`` (→ 400)
        si la validation échoue (liste vide, requête principale absente, requête
        non SELECT, trop d'onglets, titre vide…).
        """
        from app.models.dashboard import Dashboard, DashboardWidget

        stmt = (
            select(DashboardWidget)
            .join(Dashboard)
            .where(
                DashboardWidget.id == widget_id,
                DashboardWidget.dashboard_id == dashboard_id,
                Dashboard.user_id == user_id,
            )
        )
        result = await session.execute(stmt)
        widget = result.scalar_one_or_none()
        if not widget:
            return None

        if widget.widget_type != "grid":
            raise ValueError("Les feuilles SQL ne concernent que les widgets grille.")

        # Une grille a TOUJOURS au moins sa feuille principale (``query``
        # obligatoire). Liste vide / non-liste = refus EXPLICITE — ne JAMAIS
        # effacer silencieusement la requête principale.
        if not isinstance(sheets, list) or not sheets:
            raise ValueError("Au moins une feuille SQL est requise.")

        primary = sheets[0]
        primary_query = primary.get("query") if isinstance(primary, dict) else None
        # ``not primary_query.strip()`` : une requête BLANCHE (« "   " ») est un
        # str → passerait un simple isinstance, puis ``validate()`` la laisse
        # passer (le bloc SELECT/WITH est gardé par ``query.strip()`` et
        # « query required » est faux car « "   " » est truthy) → on écraserait
        # SILENCIEUSEMENT la requête principale valide existante (200 au save,
        # grille cassée au render). Parité avec la garde des onglets additionnels.
        if not isinstance(primary_query, str) or not primary_query.strip():
            raise ValueError("La feuille principale doit comporter une requête SQL.")

        # Feuilles additionnelles : ne garder que {label, query} (drop des clés
        # client parasites). Entrées invalides laissées telles quelles → message
        # clair via validate() plutôt qu'ignorées silencieusement.
        extras: Any = [
            {"label": t.get("label"), "query": t.get("query")} if isinstance(t, dict) else t
            for t in sheets[1:]
        ]

        # NOUVEAU dict (≠ mutation in-place) → SQLAlchemy détecte le changement
        # (colonne JSON simple) et bumpe updated_at → invalide le cache widget.
        new_config = dict(widget.data_source_config or {})
        new_config["query"] = primary.get("query")
        if extras:
            new_config["extra_tabs"] = extras
        else:
            new_config.pop("extra_tabs", None)  # plus d'onglet additionnel → config propre
        widget.data_source_config = new_config

        errors = widget.validate()
        if errors:
            raise ValueError("; ".join(errors))

        # Mode classeur : garder le classeur (source de vérité exécutée)
        # aligné avec le miroir modifié via cette API.
        await self._sync_workbook_sql_sheets(user_id, dashboard_id, widget_id, new_config)

        data = widget.to_dict()
        await session.commit()

        logger.info(
            "Feuilles SQL widget mises à jour: id=%s (%s feuille(s))", widget_id, len(sheets)
        )
        from app.services.anonymization.auto_scan import schedule_target_rescan

        schedule_target_rescan(user_id, "dashboard", dashboard_id)
        return data

    async def save_widget_workbook(
        self,
        session: AsyncSession,
        widget_id: int,
        dashboard_id: int,
        user_id: int,
        raw_bytes: bytes,
        expected_hash: Optional[str] = None,
    ) -> Optional[dict]:
        """Sauvegarde MANUELLE du classeur d'un widget grille (bouton
        « Enregistrer » du widget — Ctrl+S de la grille embarquée).

        Le payload est un classeur Komptia complet (``serialize()`` de la
        grille, gzip ou JSON brut) : TOUT l'état du widget — feuilles SQL avec
        leur requête, feuilles manuelles, cellules éditées, mise en forme,
        tris/filtres, cellules SQL (cellDetails). Un seul chemin de
        persistance : peu importe QUI a modifié le SQL (éditeur manuel,
        copilot, collage), la sauvegarde capture l'état réel affiché.

        Après sauvegarde, le classeur devient la SOURCE DE VÉRITÉ exécutée au
        chargement (cf. ``_fetch_sql_data`` mode classeur) ;
        ``config.query``/``extra_tabs`` sont mis à jour comme MIROIR dérivé
        (modale d'édition, exports, rétro-compat).

        Owner-only (join ``Dashboard.user_id``). Retourne ``None`` si widget
        introuvable / non-propriétaire. Lève ``ValueError`` (→ 400),
        :class:`~app.services.dashboard.widget_workbook_store.WorkbookConflictError`
        (→ 412) ou
        :class:`~app.services.dashboard.widget_workbook_store.WorkbookQuotaError`
        (→ 413).
        """
        import asyncio as _asyncio
        import json as _json

        from app.models.dashboard import Dashboard, DashboardWidget
        from app.services.ai.sql_validator import check_sql_dangerous
        from app.services.classeur.reader import decode_afz_bytes
        from app.services.dashboard import widget_workbook_store as wbstore

        stmt = (
            select(DashboardWidget)
            .join(Dashboard)
            .where(
                DashboardWidget.id == widget_id,
                DashboardWidget.dashboard_id == dashboard_id,
                Dashboard.user_id == user_id,
            )
        )
        result = await session.execute(stmt)
        widget = result.scalar_one_or_none()
        if not widget:
            return None
        if widget.widget_type != "grid":
            raise ValueError("La sauvegarde de classeur ne concerne que les widgets grille.")

        from app.services.storage_manager import get_storage_quota_bytes_sync

        # Cap de décompression DÉCOUPLÉ du quota disque (revue adv.
        # 2026-06-10 : une bombe gzip de ~500 Ko explosait jusqu'au quota —
        # 500 Mio de texte JSON en RAM, puis json.loads ×N). Erreur explicite
        # (GunzipTooLargeError → 400), jamais de troncature silencieuse.
        decompress_cap = min(get_storage_quota_bytes_sync(), wbstore.MAX_WORKBOOK_JSON_BYTES)
        try:
            data = await _asyncio.to_thread(
                decode_afz_bytes,
                raw_bytes,
                source=f"widget:{widget_id}",
                max_decompressed_bytes=decompress_cap,
            )
        except (_json.JSONDecodeError, UnicodeDecodeError, ValueError) as exc:
            raise ValueError("Classeur illisible (JSON/gzip invalide).") from exc

        if (
            not isinstance(data, dict)
            or data.get("app") != "komptia"
            or not isinstance(data.get("tabs"), list)
        ):
            raise ValueError("Format de classeur invalide.")
        tabs = data["tabs"]
        if not tabs or not all(isinstance(t, dict) for t in tabs):
            raise ValueError("Format de classeur invalide (onglets).")
        if len(tabs) > DashboardWidget.MAX_GRID_WORKBOOK_TABS:
            raise ValueError(
                f"Trop de feuilles : {len(tabs)} "
                f"(maximum {DashboardWidget.MAX_GRID_WORKBOOK_TABS})."
            )

        sql_sheets = wbstore.extract_sql_sheets(data)
        if not sql_sheets:
            raise ValueError("Le classeur doit contenir au moins une feuille SQL.")
        if len(sql_sheets) > 1 + DashboardWidget.MAX_GRID_EXTRA_TABS:
            raise ValueError(
                f"Trop de feuilles SQL : {len(sql_sheets)} "
                f"(maximum {1 + DashboardWidget.MAX_GRID_EXTRA_TABS})."
            )

        # Mêmes gardes que le save legacy (validate + _execute_grid_extra_tabs) :
        # SELECT/WITH, NUL (CWE-158), patterns dangereux, cap longueur — par
        # feuille SQL, avec un message qui NOMME la feuille fautive.
        for sheet in sql_sheets:
            q = sheet["query"]
            label = sheet["label"] or f"feuille {sheet['index'] + 1}"
            if "\x00" in q:
                raise ValueError(f"Caractère NUL interdit dans la requête (« {label} »).")
            s_upper = q.strip().upper()
            if not (s_upper.startswith("SELECT") or s_upper.startswith("WITH")):
                raise ValueError(f"Seules les requêtes SELECT sont autorisées (« {label} »).")
            if check_sql_dangerous(q):
                raise ValueError(f"Seules les requêtes SELECT sont autorisées (« {label} »).")
            if len(q) > DashboardWidget.MAX_GRID_TAB_QUERY_LEN:
                raise ValueError(
                    f"Requête trop longue (« {label} », maximum "
                    f"{DashboardWidget.MAX_GRID_TAB_QUERY_LEN} caractères)."
                )

        # Cellules SQL (cellDetails) : leur SQL est ré-exécuté à la demande
        # via /api/cell-detail/execute (qui re-valide) — defense-in-depth :
        # on refuse dès le save tout SQL d'écriture persisté.
        for ti, tab in enumerate(tabs):
            cell_details = tab.get("cellDetails")
            if not isinstance(cell_details, dict):
                continue
            for cell_key, detail in cell_details.items():
                if not isinstance(detail, dict):
                    continue
                d_sql = detail.get("sql")
                if not isinstance(d_sql, str) or not d_sql.strip():
                    continue
                if "\x00" in d_sql or check_sql_dangerous(d_sql):
                    raise ValueError(
                        f"SQL de cellule non autorisé (feuille {ti + 1}, cellule {cell_key})."
                    )

        # Miroir config construit et VALIDÉ AVANT toute écriture fichier
        # (revue adv. 2026-06-10) : un save refusé (400) ne doit JAMAIS
        # laisser un fichier « refusé » devenir la source de vérité exécutée.
        extras = [
            {
                "label": (s["label"] or f"Requête {i + 2}")[
                    : DashboardWidget.MAX_GRID_TAB_LABEL_LEN
                ],
                "query": s["query"],
            }
            for i, s in enumerate(sql_sheets[1:])
        ]
        new_config = dict(widget.data_source_config or {})
        new_config["query"] = sql_sheets[0]["query"]
        if extras:
            new_config["extra_tabs"] = extras
        else:
            new_config.pop("extra_tabs", None)
        new_config["workbook_file"] = wbstore.workbook_rel_path(dashboard_id, widget_id)
        widget.data_source_config = new_config

        errors = widget.validate()
        if errors:
            raise ValueError("; ".join(errors))

        # Bump EXPLICITE de updated_at (revue adv. 2026-06-10) : un save
        # « mise en forme seule » laisse le miroir IDENTIQUE → SQLAlchemy
        # n'émet aucun UPDATE → onupdate ne tire pas → le cache de résultats
        # (version = updated_at dans la clé) servirait le classeur PRÉ-save
        # pendant tout le TTL, avec un hash périmé (412 fantôme au save
        # suivant). Le bump force l'invalidation dans tous les cas.
        from app.core import clock as _clock

        widget.updated_at = _clock.now()

        # Verrou par widget : check If-Match + écriture ATOMIQUES (sinon
        # TOCTOU — deux PUT porteurs du même hash passent tous deux le check
        # puis s'écrasent en silence, exactement ce que le 412 doit empêcher).
        async with wbstore.widget_lock(user_id, dashboard_id, widget_id):
            # Anti-clobber multi-onglets (axe 22) : If-Match optimiste sur le
            # hash du fichier courant. Premier save (pas de fichier) → skip.
            if expected_hash:
                current = await wbstore.current_hash(user_id, dashboard_id, widget_id)
                if current is not None and current != expected_hash:
                    raise wbstore.WorkbookConflictError(
                        "Ce widget a été modifié dans un autre onglet. "
                        "Rechargez la page avant d'enregistrer."
                    )

            rel_path, file_hash = await wbstore.save_workbook(
                user_id, dashboard_id, widget_id, data
            )

        widget_data = widget.to_dict()
        await session.commit()

        logger.info("Classeur widget sauvegardé : widget=%s (%s)", widget_id, rel_path)
        from app.services.anonymization.auto_scan import schedule_target_rescan

        schedule_target_rescan(user_id, "dashboard", dashboard_id)
        return {"widget": widget_data, "workbook_hash": file_hash}

    async def _sync_workbook_sql_sheets(
        self, user_id: int, dashboard_id: int, widget_id: int, config: Any
    ) -> None:
        """Répercute le miroir config (``query`` + ``extra_tabs``) dans le
        CLASSEUR du widget quand il existe.

        Sans cette sync, éditer le SQL via la modale du widget (ou les
        endpoints extra-tabs) divergerait du classeur — qui est la source de
        vérité exécutée au chargement → l'édition serait silencieusement
        ignorée. Appelée AVANT commit par ``update_widget`` /
        ``set_widget_sheets`` / ``set_widget_extra_tabs``. No-op si pas de
        classeur ou si rien n'a changé. Quota dépassé ou cap de feuilles
        atteint → ``ValueError`` (400).

        Limite connue (documentée, revue adv. 2026-06-10) : le fichier est
        écrit avant le commit du caller — un échec de commit (rare, retry DB
        en amont) laisse le classeur en avance d'une édition sur le miroir ;
        le prochain save/édition réaligne.
        """
        from app.models.dashboard import DashboardWidget
        from app.services.dashboard import widget_workbook_store as wbstore

        wb_ref = config.get("workbook_file") if isinstance(config, dict) else None
        if not isinstance(wb_ref, str) or not wb_ref:
            return

        desired: list[tuple[Optional[str], str]] = [(None, str(config.get("query") or ""))]
        for t in config.get("extra_tabs") or []:
            if isinstance(t, dict):
                desired.append((str(t.get("label") or ""), str(t.get("query") or "")))

        # Verrou par widget : load → patch → save atomique vis-à-vis d'un
        # « Enregistrer » manuel concurrent (sinon ce read-modify-write
        # réécrirait le classeur ENTIER depuis une copie périmée et
        # annulerait silencieusement le save de l'utilisateur).
        async with wbstore.widget_lock(user_id, dashboard_id, widget_id):
            wb = await wbstore.load_workbook(user_id, dashboard_id, widget_id)
            if wb is None:
                # Fichier absent/corrompu : le chargement retombera en mode
                # legacy (miroir config) — ne pas bloquer l'édition.
                return

            wb_sheets = wbstore.extract_sql_sheets(wb)
            current = [
                (None if i == 0 else s["label"], s["query"]) for i, s in enumerate(wb_sheets)
            ]
            if current == desired:
                return

            tabs = wb.get("tabs") or []

            # Appariement par IDENTITÉ de requête d'abord, position en repli
            # (revue adv. 2026-06-10) : un patch purement positionnel
            # ré-attribuait libellé/mise en forme à la MAUVAISE requête quand
            # une feuille SQL intermédiaire était retirée du miroir.
            unused = list(range(len(wb_sheets)))
            assignment: list[Optional[int]] = []
            for _label, q in desired:
                match_pos = next((p for p in unused if wb_sheets[p]["query"] == q), None)
                if match_pos is not None:
                    unused.remove(match_pos)
                assignment.append(match_pos)
            for di in range(len(assignment)):
                if assignment[di] is None and unused:
                    assignment[di] = unused.pop(0)

            # Cap : tout ce que le serveur écrit doit rester ré-enregistrable
            # par le client (sinon le prochain « Enregistrer » serait rejeté
            # « Trop de feuilles » à cause d'un état créé par le serveur).
            n_appends = sum(1 for p in assignment if p is None)
            if len(tabs) + n_appends > DashboardWidget.MAX_GRID_WORKBOOK_TABS:
                raise ValueError(
                    f"Trop de feuilles dans le classeur du widget "
                    f"(maximum {DashboardWidget.MAX_GRID_WORKBOOK_TABS}). "
                    "Supprimez des feuilles avant d'ajouter des requêtes."
                )

            # 1) Patch des feuilles appariées (query + label pour les extras).
            for di, (label, q) in enumerate(desired):
                pos = assignment[di]
                if pos is None:
                    continue
                tab = tabs[wb_sheets[pos]["index"]]
                src = tab.get("externalSource")
                if isinstance(src, dict):
                    src["query"] = q
                tab["sql"] = q
                if di > 0 and label:
                    tab["label"] = label
            # 2) Feuilles SQL nouvelles → feuilles minimales (les données
            # seront remplies par la ré-exécution au prochain chargement).
            for di, (label, q) in enumerate(desired):
                if assignment[di] is not None:
                    continue
                tabs.append(
                    {
                        "label": label or "Requête SQL",
                        "closable": True,
                        "sql": q,
                        "columns": [],
                        "rows": [],
                        "totalRowCount": 0,
                        "isArrayFormat": True,
                        "externalSource": {"type": "sql_query", "query": q},
                    }
                )
            # 3) Feuilles SQL non réclamées par le miroir → suppression
            # (indexes décroissants pour ne pas décaler les restantes).
            for pos in sorted(unused, key=lambda p: -wb_sheets[p]["index"]):
                del tabs[wb_sheets[pos]["index"]]

            try:
                await wbstore.save_workbook(user_id, dashboard_id, widget_id, wb)
            except wbstore.WorkbookQuotaError as exc:
                raise ValueError(str(exc)) from exc

    async def delete_widget(
        self,
        session: AsyncSession,
        widget_id: int,
        dashboard_id: int,
        user_id: int,
    ) -> bool:
        """Supprime un widget."""
        from app.models.dashboard import Dashboard, DashboardWidget

        stmt = (
            select(DashboardWidget)
            .join(Dashboard)
            .where(
                DashboardWidget.id == widget_id,
                DashboardWidget.dashboard_id == dashboard_id,
                Dashboard.user_id == user_id,
            )
        )
        result = await session.execute(stmt)
        widget = result.scalar_one_or_none()

        if not widget:
            return False

        # Capturer AVANT le delete (l'objet est expiré après commit).
        had_workbook = widget.widget_type == "grid" and bool(
            isinstance(widget.data_source_config, dict)
            and widget.data_source_config.get("workbook_file")
        )

        await session.delete(widget)
        await session.commit()

        if had_workbook:
            # Cycle de vie du classeur widget : le fichier meurt avec le
            # widget (quota libéré). Best-effort — un échec ne casse pas la
            # suppression (fichier orphelin caché, écrasé si l'id est réutilisé).
            from app.services.dashboard import widget_workbook_store as wbstore

            try:
                await wbstore.delete_workbook(user_id, dashboard_id, widget_id)
            except Exception:  # noqa: BLE001
                logger.warning(
                    "Suppression du classeur widget %s échouée", widget_id, exc_info=True
                )

        logger.info("Widget supprimé: id=%s", widget_id)
        return True

    async def reorder_widgets(
        self,
        session: AsyncSession,
        dashboard_id: int,
        user_id: int,
        widget_order: list[int],
    ) -> bool:
        """Réordonne les widgets d'un dashboard.

        widget_order: liste d'IDs de widgets dans le nouvel ordre.
        """
        from app.models.dashboard import Dashboard, DashboardWidget

        # Vérifier ownership
        stmt = select(Dashboard).where(Dashboard.id == dashboard_id, Dashboard.user_id == user_id)
        result = await session.execute(stmt)
        if not result.scalar_one_or_none():
            return False

        # Mettre à jour l'ordre
        for position, widget_id in enumerate(widget_order):
            update_stmt = select(DashboardWidget).where(
                DashboardWidget.id == widget_id,
                DashboardWidget.dashboard_id == dashboard_id,
            )
            result = await session.execute(update_stmt)
            widget = result.scalar_one_or_none()
            if widget:
                widget.position_order = position

        await session.commit()
        logger.info("Widgets réordonnés: dashboard=%s", dashboard_id)
        return True

    # ── Trend computation ──────────────────────────────────────────────────

    @staticmethod
    def _build_trend(
        current: float, previous: float, higher_is_better: bool = True
    ) -> Optional[dict]:
        """Compare current vs previous period and return trend indicator.

        Returns None when no meaningful comparison is possible (previous == 0).
        """
        if previous is None or previous == 0:
            return None

        change_pct = round(((current - previous) / previous) * 100, 1)

        if abs(change_pct) < 0.1:
            return {"direction": "flat", "change_pct": 0, "sentiment": "neutral"}

        direction = "up" if change_pct > 0 else "down"
        if higher_is_better:
            sentiment = "positive" if change_pct > 0 else "negative"
        else:
            sentiment = "negative" if change_pct > 0 else "positive"

        return {
            "direction": direction,
            "change_pct": abs(change_pct),
            "sentiment": sentiment,
        }

    # ── Récupération de données pour les widgets ───────────────────────────

    async def get_widget_data(
        self,
        session: AsyncSession,
        widget_id: int,
        dashboard_id: int,
        user_id: int,
        user: Any = None,
    ) -> Optional[dict]:
        """Récupère les données d'un widget selon sa source (owner only)."""
        from app.models.dashboard import Dashboard, DashboardWidget

        stmt = (
            select(DashboardWidget)
            .join(Dashboard)
            .where(
                DashboardWidget.id == widget_id,
                DashboardWidget.dashboard_id == dashboard_id,
                Dashboard.user_id == user_id,
            )
        )
        result = await session.execute(stmt)
        widget = result.scalar_one_or_none()

        if not widget:
            return None

        config = widget.data_source_config or {}

        if widget.data_source_type == "metric":
            return await self._fetch_metric_data(session, config, widget.widget_type, user_id)
        elif widget.data_source_type == "sql":
            # Propage widget_type pour que le service choisisse le bon path
            # (grid → rendu brut + cap admin ; table/chart/kpi → 500 cap +
            # transform). Sans ça, un widget grid récupéré via cette méthode
            # tomberait silencieusement sur le path "table" (cap 500 + pas
            # de metadata row_count/truncated → régression d'affichage).
            return await self._fetch_sql_data(
                config,
                widget_type=widget.widget_type or "table",
                chart_type=widget.chart_type,
                user=user,
            )
        elif widget.data_source_type == "static":
            return self._fetch_static_data(config)
        else:
            return {"error": "Source de données inconnue."}

    async def get_all_widget_data(
        self,
        session: AsyncSession,
        dashboard_id: int,
        user_id: int,
        period_override: Optional[int] = None,
        filter_state: Optional[dict] = None,
        drill_filters: Optional[dict] = None,
        user: Any = None,
        force_refresh: bool = False,
        apply_render_cap: bool = True,
    ) -> dict[int, dict]:
        """Récupère les données de tous les widgets d'un dashboard.

        ``force_refresh=True`` (bouton « rafraîchir » du front, ``?refresh=1``)
        bypasse le cache de résultats et force une exécution Sage fraîche.

        ``apply_render_cap=False`` (chemins EXPORT/EMAIL, revue adv. #18c
        2026-06-10) : le cap de RENDU 500 (Plotly/DOM) n'a aucun sens pour un
        fichier — avant, un export de widget table sur 600 lignes livrait la
        slice de 500 en silence. Les flags ``source_truncated``/``row_count``
        restent posés (honnêteté), seul le slicing est sauté.
        """
        from app.models.dashboard import Dashboard, DashboardWidget

        # Vérifier accès — strict owner only
        dash_stmt = select(Dashboard).where(
            Dashboard.id == dashboard_id, Dashboard.user_id == user_id
        )
        dash_result = await session.execute(dash_stmt)
        dashboard = dash_result.scalar_one_or_none()
        if not dashboard:
            return {}

        # Charger les widgets
        widget_stmt = (
            select(DashboardWidget)
            .where(DashboardWidget.dashboard_id == dashboard_id)
            .order_by(DashboardWidget.position_order)
        )
        widget_result = await session.execute(widget_stmt)
        widgets = widget_result.scalars().all()

        # Load filter definitions if filter_state is active
        filter_definitions = []
        if filter_state:
            from app.models.dashboard import DashboardFilter

            filt_stmt = select(DashboardFilter).where(DashboardFilter.dashboard_id == dashboard_id)
            filt_result = await session.execute(filt_stmt)
            filter_definitions = [f.to_dict() for f in filt_result.scalars().all()]

        # Cache de résultats (incident lenteur 2026-06-08) : un dashboard
        # ré-ouvert ne doit pas re-lancer chaque requête Sage lourde à chaque
        # fois. Keyé par user (isolation cross-user), invalidé par contenu (le
        # SQL est dans la clé) + TTL court. ``force_refresh`` (?refresh=1) bypasse.
        from app.services.dashboard.widget_result_cache import get_widget_result_cache

        result_cache = get_widget_result_cache()
        # Token de version des règles d'accès (RLS) du user : intégré à la clé
        # pour que tout changement de droits admin invalide AUTOMATIQUEMENT le
        # cache (sinon un résultat déjà filtré resterait servi jusqu'au TTL =
        # sur-exposition de données au même user). O(1), aucun accès BDD.
        from app.services.data_access import enforcer as _da_enforcer

        rls_token = _da_enforcer.rules_cache_token(user_id)

        data = {}
        for widget in widgets:
            config = widget.data_source_config or {}

            # Compute effective filters for this widget (respecting exclusions)
            effective_filters = None
            if filter_state and filter_definitions:
                excluded = set(config.get("excluded_filter_params", []))
                effective_filters = {k: v for k, v in filter_state.items() if k not in excluded}
                if not effective_filters:
                    effective_filters = None

            try:
                if widget.data_source_type == "metric":
                    effective_config = dict(config)
                    if period_override is not None:
                        effective_config["period"] = period_override
                    data[widget.id] = await self._fetch_metric_data(
                        session, effective_config, widget.widget_type, user_id
                    )
                elif widget.data_source_type == "sql":
                    # Clé = (user, dashboard, widget, SQL, type, filtres, drill,
                    # période, version). Hit → résultat instantané + méta
                    # ``_cache`` (âge visible côté front). Miss → exécution + set.
                    cache_key = result_cache.make_key(
                        user_id,
                        dashboard_id,
                        widget.id,
                        query=(config.get("query", "") if isinstance(config, dict) else ""),
                        widget_type=widget.widget_type or "table",
                        chart_type=widget.chart_type,
                        period=period_override,
                        filter_state=effective_filters,
                        drill_filters=drill_filters,
                        version=(getattr(widget, "updated_at", None), rls_token),
                    )
                    cached = None if force_refresh else result_cache.get(cache_key)
                    if cached is not None:
                        data[widget.id] = cached
                    else:
                        # Mode classeur (widget grille sauvegardé) : le classeur
                        # est la SOURCE DE VÉRITÉ exécutée — chargé ici (miss
                        # uniquement, pas de lecture disque sur hit cache).
                        # Fichier absent/corrompu → None = fallback legacy
                        # gracieux dans _fetch_sql_data (config.query miroir).
                        workbook = None
                        workbook_hash = None
                        workbook_missing = False
                        if widget.widget_type == "grid" and isinstance(config, dict):
                            wb_ref = config.get("workbook_file")
                            if isinstance(wb_ref, str) and wb_ref:
                                from app.services.dashboard import widget_workbook_store

                                workbook, workbook_hash = (
                                    await widget_workbook_store.load_workbook_with_hash(
                                        user_id, dashboard_id, widget.id
                                    )
                                )
                                workbook_missing = workbook is None
                        # chart_type=None → inférence automatique côté transform
                        sql_result = await self._fetch_sql_data(
                            config,
                            widget_type=widget.widget_type or "table",
                            chart_type=widget.chart_type,
                            filter_state=effective_filters,
                            filter_definitions=filter_definitions,
                            drill_filters=drill_filters,
                            user=user,
                            apply_render_cap=apply_render_cap,
                            workbook=workbook,
                            workbook_hash=workbook_hash,
                            workbook_missing=workbook_missing,
                        )
                        # Un résultat EXPORT (non cappé au rendu) ne doit PAS
                        # peupler le cache de la vue live : il serait servi au
                        # front sans cap (charts non bornés = le freeze que le
                        # cap de rendu prévient).
                        if apply_render_cap:
                            result_cache.set(cache_key, sql_result)
                        data[widget.id] = sql_result
                elif widget.data_source_type == "static":
                    data[widget.id] = self._fetch_static_data(config)
                else:
                    data[widget.id] = {"error": "Source inconnue."}
            except Exception:
                logger.warning("Erreur récupération données widget %s", widget.id, exc_info=True)
                data[widget.id] = {"error": "Erreur lors de la récupération des données."}

        return data

    def _fetch_static_data(self, config: dict) -> dict:
        """Widget statique (widget_type='text', data_source_type='static').

        Pas de fetch externe : le contenu vit dans ``data_source_config``.
        Retourne un shape uniforme ``{type, title, content}`` que le frontend
        rend via ``textContent`` (zéro risque XSS).

        **Sync intentionnel** — pas d'I/O, contrairement à
        :meth:`_fetch_sql_data` et :meth:`_fetch_metric_data`. Si une
        évolution future ajoute du I/O (ex: anti-PII Ollama sur le contenu),
        cette méthode DOIT devenir async ET les 2 call-sites (``get_widget_data``
        et ``get_all_widget_data``) doivent être adaptés avec ``await``.

        **Defense-in-depth** : caps appliqués ici (truncate plutôt que reject),
        en plus de :meth:`DashboardWidget.validate`. Couvre 2 cas :
        - widget legacy créé avant l'introduction des caps (validate
          rejetterait à la création, mais pas à la lecture)
        - row inserté directement en BDD (SQL/script) qui bypasse validate
        - type mismatch dans config (dict/list au lieu de str) — on retourne
          une string vide plutôt que stringifier un repr Python qui
          confondrait l'utilisateur (``{'x': 1}`` affiché littéralement).
        """
        from app.models.dashboard import DashboardWidget

        cfg = config or {}
        raw_title = cfg.get("title")
        raw_content = cfg.get("content")
        # Type-safe : un dict/list dans config (BDD corrompue) → "" plutôt que repr.
        title = raw_title.strip() if isinstance(raw_title, str) else ""
        content = raw_content if isinstance(raw_content, str) else ""
        return {
            "type": "static",
            "title": title[: DashboardWidget.MAX_TEXT_WIDGET_TITLE_LEN],
            "content": content[: DashboardWidget.MAX_TEXT_WIDGET_CONTENT_LEN],
        }

    async def _fetch_metric_data(
        self,
        session: AsyncSession,
        config: dict,
        widget_type: str,
        user_id: int,
    ) -> dict:
        """Récupère les données d'une métrique prédéfinie."""
        metric_name = config.get("metric_name", "")
        try:
            period_days = min(max(int(config.get("period", 7)), 1), 365)
        except (ValueError, TypeError):
            period_days = 7

        if metric_name not in AVAILABLE_METRICS:
            return {"error": f"Métrique inconnue: {metric_name}"}

        try:
            return await self._execute_metric(session, metric_name, user_id, period_days)
        except Exception:
            logger.warning("Erreur métrique %s", metric_name, exc_info=True)
            return {"error": "Erreur lors du calcul de la métrique."}

    async def _execute_metric(
        self,
        session: AsyncSession,
        metric_name: str,
        user_id: int,
        period_days: int,
    ) -> dict:
        """Exécute une métrique prédéfinie et retourne les données formatées."""
        from app.models.automation import Automation
        from app.models.execution import Execution
        from app.models.search_history import SearchHistory

        cutoff = clock.now() - timedelta(days=period_days)

        if metric_name == "total_searches":
            stmt = (
                select(func.count())
                .select_from(SearchHistory)
                .where(
                    SearchHistory.user_id == user_id,
                    SearchHistory.created_at >= cutoff,
                )
            )
            result = await session.execute(stmt)
            count = result.scalar() or 0

            # Previous period for trend
            prev_start = cutoff - timedelta(days=period_days)
            prev_stmt = (
                select(func.count())
                .select_from(SearchHistory)
                .where(
                    SearchHistory.user_id == user_id,
                    SearchHistory.created_at >= prev_start,
                    SearchHistory.created_at < cutoff,
                )
            )
            prev_count = (await session.execute(prev_stmt)).scalar() or 0

            return {
                "type": "kpi",
                "value": count,
                "label": "Recherches",
                "trend": self._build_trend(count, prev_count, higher_is_better=True),
            }

        elif metric_name == "success_rate":
            total_stmt = (
                select(func.count())
                .select_from(SearchHistory)
                .where(
                    SearchHistory.user_id == user_id,
                    SearchHistory.created_at >= cutoff,
                )
            )
            total = (await session.execute(total_stmt)).scalar() or 0

            success_stmt = (
                select(func.count())
                .select_from(SearchHistory)
                .where(
                    SearchHistory.user_id == user_id,
                    SearchHistory.created_at >= cutoff,
                    SearchHistory.success.is_(True),
                )
            )
            success = (await session.execute(success_stmt)).scalar() or 0

            rate = round((success / total) * 100, 1) if total > 0 else 0

            # Previous period for trend
            prev_start = cutoff - timedelta(days=period_days)
            prev_total_stmt = (
                select(func.count())
                .select_from(SearchHistory)
                .where(
                    SearchHistory.user_id == user_id,
                    SearchHistory.created_at >= prev_start,
                    SearchHistory.created_at < cutoff,
                )
            )
            prev_total = (await session.execute(prev_total_stmt)).scalar() or 0

            prev_success_stmt = (
                select(func.count())
                .select_from(SearchHistory)
                .where(
                    SearchHistory.user_id == user_id,
                    SearchHistory.created_at >= prev_start,
                    SearchHistory.created_at < cutoff,
                    SearchHistory.success.is_(True),
                )
            )
            prev_success = (await session.execute(prev_success_stmt)).scalar() or 0
            prev_rate = round((prev_success / prev_total) * 100, 1) if prev_total > 0 else 0

            return {
                "type": "kpi",
                "value": rate,
                "label": "Taux de succès",
                "unit": "%",
                "trend": self._build_trend(rate, prev_rate, higher_is_better=True),
            }

        elif metric_name == "daily_searches":
            # Single GROUP BY query instead of N+1 (was 1 query per day).
            # On regroupe via func.date() (renvoie la chaîne 'YYYY-MM-DD') et NON
            # cast(created_at, Date) : sous SQLite (la BDD locale), CAST(... AS DATE)
            # applique l'affinité NUMERIC à la chaîne ISO → renvoie un ENTIER → le
            # result-processor Date de SQLAlchemy (date.fromisoformat) lève
            # TypeError. L'exception était avalée par le try/except de
            # _fetch_metric_data → la métrique tombait TOUJOURS en erreur (chart
            # vide). func.date() est non-typé côté SQLAlchemy → aucune conversion
            # appliquée → chaîne brute, pas de crash.
            stmt = (
                select(
                    func.date(SearchHistory.created_at).label("day"),
                    func.count().label("cnt"),
                )
                .where(
                    SearchHistory.user_id == user_id,
                    SearchHistory.created_at >= cutoff,
                )
                .group_by(func.date(SearchHistory.created_at))
            )
            result = await session.execute(stmt)
            # Clés normalisées en 'YYYY-MM-DD' (agnostique backend : SQLite renvoie
            # déjà une chaîne ; un backend date-typé renverrait un date → str()).
            counts_by_day = {str(row[0]): row[1] for row in result.all()}

            # Build full date range (fill missing days with 0)
            today = clock.now().date()
            rows = []
            for i in range(period_days - 1, -1, -1):
                day = today - timedelta(days=i)
                rows.append(
                    {
                        "date": day.strftime("%d/%m"),
                        # day.isoformat() == 'YYYY-MM-DD' → match la clé func.date().
                        "count": counts_by_day.get(day.isoformat(), 0),
                    }
                )
            return {
                "type": "chart",
                "labels": [r["date"] for r in rows],
                "datasets": [{"label": "Recherches", "data": [r["count"] for r in rows]}],
            }

        elif metric_name == "execution_status":
            # Filtre fenêtre temporelle (started_at >= cutoff) — cohérent avec
            # total_searches / success_rate / daily_searches. Sans ce filtre, la
            # répartition succès/échec agrégeait TOUTES les exécutions de tout
            # temps, alors que le dashboard applique un period_override à CHAQUE
            # widget métrique (cf. _fetch_dashboard_widgets) → l'utilisateur qui
            # sélectionne « 7 derniers jours » voyait en réalité la distribution
            # depuis toujours (données fausses silencieuses).
            stmt = (
                select(Execution.status, func.count())
                .join(Automation)
                .where(
                    Automation.user_id == user_id,
                    Execution.started_at >= cutoff,
                )
                .group_by(Execution.status)
            )
            result = await session.execute(stmt)
            rows = result.all()
            labels = [r[0] or "inconnu" for r in rows]
            values = [r[1] for r in rows]
            return {"type": "chart", "labels": labels, "datasets": [{"data": values}]}

        elif metric_name == "active_automations":
            stmt = (
                select(func.count())
                .select_from(Automation)
                .where(Automation.user_id == user_id, Automation.is_active.is_(True))
            )
            count = (await session.execute(stmt)).scalar() or 0
            return {"type": "kpi", "value": count, "label": "Automatisations actives"}

        elif metric_name == "total_reports":
            from app.models.user_storage import FileMetadata

            stmt = (
                select(func.count())
                .select_from(FileMetadata)
                .where(FileMetadata.user_id == user_id)
            )
            count = (await session.execute(stmt)).scalar() or 0
            return {"type": "kpi", "value": count, "label": "Fichiers"}

        elif metric_name == "total_contacts":
            from app.services.contacts.contact_service import ContactService

            svc = ContactService()
            stats = await svc.get_stats(session, user_id)
            return {
                "type": "kpi",
                "value": stats.data.get("total_contacts", 0) if stats.success else 0,
                "label": "Contacts",
            }

        elif metric_name == "recent_searches":
            stmt = (
                select(
                    SearchHistory.question,
                    SearchHistory.success,
                    SearchHistory.created_at,
                )
                .where(SearchHistory.user_id == user_id)
                .order_by(SearchHistory.created_at.desc())
                .limit(10)
            )
            result = await session.execute(stmt)
            rows = result.all()
            return {
                "type": "table",
                "columns": ["Question", "Statut", "Date"],
                "rows": [
                    [
                        r[0][:80] if r[0] else "",
                        "OK" if r[1] else "Erreur",
                        # Heure SERVEUR (config.server.timezone) : cellule de table
                        # widget générique sans <time> par cellule → pas de conversion
                        # navigateur possible ici, donc on rend dans la TZ serveur
                        # (= TZ du cabinet) plutôt que l'UTC brut (sinon +Nh).
                        (clock.to_local(r[2]) or r[2]).strftime("%d/%m %H:%M") if r[2] else "",
                    ]
                    for r in rows
                ],
            }

        # ``top_users`` et ``avg_response_time`` retirés (BLOCKING #4-5
        # review). Defense-in-depth : un dashboard legacy stocké en BDD
        # avec un widget pointant vers ces metrics retournera une erreur
        # explicite plutôt qu'un agrégat cross-user. Le frontend doit
        # afficher cette erreur ou cacher le widget — pas de fail-open.
        if metric_name in ("top_users", "avg_response_time"):
            return {
                "error": (
                    f"Métrique '{metric_name}' retirée pour empêcher la fuite "
                    "cross-user. Modifiez le widget ou supprimez-le."
                )
            }

        return {"error": f"Métrique non implémentée: {metric_name}"}

    # ── Auto chart-type inference ───────────────────────────────────────
    # Heuristique déterministe (code > prompt) pour choisir le type de
    # graphique le plus adapté aux données retournées par une requête SQL,
    # quand l'utilisateur n'en a pas spécifié un. Activée uniquement si
    # `chart_type` est None/""/"auto" — les widgets existants avec un
    # chart_type explicite ne sont pas affectés.

    # Mots-clés temporels matchés par "token" (séparés par _ / espace / début / fin)
    # pour éviter les faux positifs du genre "created_by" → temporel.
    # On n'inclut PAS "created"/"updated"/"modified" parce qu'ils apparaissent
    # aussi dans des champs non-temporels (created_by, updated_by = auteurs).
    # La détection par VALEURS ISO couvre created_at/updated_at si nécessaire.
    _TEMPORAL_NAME_TOKENS = frozenset(
        {
            "date",
            "jour",
            "mois",
            "annee",
            "année",
            "ann",
            "year",
            "day",
            "time",
            "timestamp",
            "horodatage",
        }
    )

    @staticmethod
    def _is_numeric_column(rows: list[list], col_idx: int) -> bool:
        """Teste si une colonne est numérique d'après la 1ère valeur non-None.

        Rejette les booléens (même si bool() est un sous-type d'int) et les
        valeurs flottantes non-finies (NaN/±Inf), qui casseraient la
        sérialisation JSON côté réponse HTTP.
        """
        import math as _math

        for row in rows:
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            if isinstance(val, bool):
                return False
            try:
                parsed = float(val)
            except (ValueError, TypeError):
                return False
            return _math.isfinite(parsed)
        return False

    @classmethod
    def _name_has_temporal_token(cls, name: str) -> bool:
        """True si `name` contient un mot-clé temporel comme token séparé.

        Découpe sur non-alphanumériques (snake_case, kebab-case, espaces, …)
        puis teste chaque token contre la liste fermée. Évite "created_by" /
        "year_plan_author" etc. de déclencher à tort.
        """
        import re as _re

        if not name:
            return False
        tokens = _re.split(r"[^a-zA-Zàâäéèêëîïôöùûüç]+", name.lower())
        return any(tok in cls._TEMPORAL_NAME_TOKENS for tok in tokens if tok)

    @classmethod
    def _is_temporal_column(cls, columns: list[str], rows: list[list], col_idx: int) -> bool:
        """Détecte qu'une colonne représente du temps.

        Combine deux signaux :
        - Nom contient un mot-clé temporel (en tant que token séparé)
        - ≥80% des valeurs non-None / non-vides parsent comme ISO-date/datetime
        """
        if col_idx >= len(columns):
            return False
        if cls._name_has_temporal_token(str(columns[col_idx])):
            return True
        # Heuristique par valeurs : tente parse sur les 20 premières valeurs
        # utiles (non-None, non-whitespace-only).
        sample = []
        for row in rows[:60]:
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            if isinstance(val, str) and not val.strip():
                continue
            sample.append(val)
            if len(sample) >= 20:
                break
        if not sample:
            return False
        parsed = 0
        for val in sample:
            if isinstance(val, datetime):
                parsed += 1
                continue
            if not isinstance(val, str):
                continue
            s = val.strip()
            try:
                datetime.fromisoformat(s.replace("Z", "+00:00"))
                parsed += 1
            except ValueError:
                pass
        return (parsed / len(sample)) >= 0.8

    @classmethod
    def _infer_chart_type(cls, columns: list[str], rows: list[list]) -> str:
        """Choisit le meilleur chart_type à partir de la forme des données.

        Retourne une valeur de ``DashboardWidget.VALID_CHART_TYPES``.
        Règles :
        - Col temporelle (token du nom ou valeurs ISO) + col(s) numérique(s) → "line"
        - 1 catégorie + 1 numérique + 2..8 lignes → "pie" (plus lisible)
        - Catégorie + numérique(s) → "bar"
        - Fallback safe → "bar"

        Note : on NE renvoie PAS "scatter" automatiquement. ``_transform_sql_to_chart``
        utilise toujours la 1ère colonne comme labels (X) et les suivantes numériques
        comme séries (Y) — ce schéma n'est pas compatible avec scatter (qui veut
        X *et* Y numériques). L'utilisateur peut toujours choisir scatter manuellement
        via les paramètres avancés ; la détection auto reste côté "bar" pour éviter
        une visualisation trompeuse.
        """
        if not columns or not rows:
            return "bar"

        numeric_cols = [i for i in range(len(columns)) if cls._is_numeric_column(rows, i)]

        # 1ère colonne non-numérique = candidate "catégorie/axe X"
        if numeric_cols and 0 not in numeric_cols:
            if cls._is_temporal_column(columns, rows, 0):
                return "line"
            if len(numeric_cols) == 1 and 2 <= len(rows) <= 8:
                return "pie"
            return "bar"

        return "bar"

    def _transform_sql_to_chart(
        self,
        columns: list[str],
        rows: list[list],
        chart_type: Optional[str] = None,
    ) -> dict:
        """Transforme des résultats SQL tabulaires en format chart Plotly.

        Convention:
        - 1ère colonne → labels (axe X ou catégories)
        - Colonnes numériques restantes → datasets (séries Y)
        - Si aucune colonne numérique trouvée → fallback table

        ``chart_type`` peut être None / "" / "auto" → on infère le type
        depuis la forme des données. Sinon on respecte la valeur explicite.
        Le type effectif est inclus dans le dict retourné pour que le
        frontend puisse fallback dessus si le widget n'a pas de chart_type
        stocké.
        """
        effective_chart_type = (
            chart_type
            if chart_type and chart_type != "auto"
            else self._infer_chart_type(columns, rows)
        )

        if not columns or not rows:
            return {
                "type": "chart",
                "chart_type": effective_chart_type,
                "labels": [],
                "datasets": [],
            }

        labels = [str(row[0]) if row[0] is not None else "" for row in rows]

        # Une seule source de vérité pour "colonne numérique" : le helper.
        numeric_cols = [i for i in range(1, len(columns)) if self._is_numeric_column(rows, i)]

        # Aucune colonne numérique → fallback table
        if not numeric_cols:
            return {
                "type": "table",
                "columns": columns,
                "rows": rows,
            }

        datasets = []
        for col_idx in numeric_cols:
            col_name = columns[col_idx]
            values = []
            for row in rows:
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    try:
                        values.append(float(val))
                    except (ValueError, TypeError):
                        values.append(0)
                else:
                    values.append(0)
            datasets.append({"label": col_name, "data": values})

        return {
            "type": "chart",
            "chart_type": effective_chart_type,
            "labels": labels,
            "datasets": datasets,
        }

    #: Cap de RENDU des widgets non-grid, appliqué APRÈS transformation/
    #: agrégation (#18c — l'ancien cap 500 appliqué AVANT faussait les
    #: agrégats). Ce n'est PAS un cap de données : l'agrégation a vu toutes
    #: les lignes (jusqu'au cap admin) ; on borne seulement ce que Plotly/
    #: le HTML statique doivent dessiner, avec flags honnêtes.
    _WIDGET_RENDER_MAX_ROWS = 500

    def _cap_widget_render(
        self, result: dict, source_truncated: bool, apply_render_cap: bool = True
    ) -> dict:
        """Borne le RENDU d'un résultat de widget non-grid, honnêtement.

        - ``rows`` au-delà de ``_WIDGET_RENDER_MAX_ROWS`` → slice +
          ``truncated: True`` + ``row_count`` (total avant slice) — mêmes
          clés que le widget grid, le front affiche le même badge.
        - ``source_truncated`` (cap admin atteint au FETCH) → propagé tel
          quel : les agrégats sont alors calculés sur des données partielles,
          le front doit l'annoncer distinctement (« calculé sur les N
          premières lignes — cap admin »).
        """
        if not isinstance(result, dict):
            return result
        rows = result.get("rows")
        if isinstance(rows, list):
            result.setdefault("row_count", len(rows))
            if apply_render_cap and len(rows) > self._WIDGET_RENDER_MAX_ROWS:
                result["rows"] = rows[: self._WIDGET_RENDER_MAX_ROWS]
                result["truncated"] = True
        # Charts : la sortie est {labels, datasets}, PAS rows (revue adv.
        # 2026-06-10, finding CRITIQUE) — sans ce cap, un group-by haute
        # cardinalité (ex. GROUP BY numero_facture) envoie N milliers de
        # buckets entiers à Plotly → freeze/OOM navigateur. ``labels`` est
        # l'axe X (nb de buckets) ; chaque dataset porte une valeur par
        # bucket → slicer en parallèle.
        labels = result.get("labels")
        if (
            apply_render_cap
            and isinstance(labels, list)
            and len(labels) > self._WIDGET_RENDER_MAX_ROWS
        ):
            result.setdefault("row_count", len(labels))
            keep = self._WIDGET_RENDER_MAX_ROWS
            result["labels"] = labels[:keep]
            for ds in result.get("datasets") or []:
                if isinstance(ds, dict) and isinstance(ds.get("data"), list):
                    ds["data"] = ds["data"][:keep]
            result["truncated"] = True
        if source_truncated:
            result["source_truncated"] = True
        return result

    def _transform_sql_to_kpi(self, columns: list[str], rows: list[list]) -> dict:
        """Transforme des résultats SQL en format KPI (valeur unique).

        Convention:
        - 1ère colonne numérique de la 1ère ligne → valeur KPI
        - Nom de la colonne → label
        """
        if not columns or not rows:
            return {"type": "kpi", "value": 0, "label": "Aucune donnée"}

        first_row = rows[0]
        # Chercher la première valeur numérique
        for col_idx, col_name in enumerate(columns):
            val = first_row[col_idx] if col_idx < len(first_row) else None
            if val is not None:
                try:
                    numeric_val = float(val)
                    # Afficher en entier si c'est un entier
                    if numeric_val == int(numeric_val):
                        numeric_val = int(numeric_val)
                    return {"type": "kpi", "value": numeric_val, "label": col_name}
                except (ValueError, TypeError):
                    continue

        # Pas de valeur numérique → texte de la première cellule
        val = first_row[0] if first_row else "—"
        return {"type": "kpi", "value": str(val) if val is not None else "—", "label": columns[0]}

    async def _fetch_sql_data(
        self,
        config: dict,
        widget_type: str = "table",
        chart_type: Optional[str] = None,
        filter_state: Optional[dict] = None,
        filter_definitions: Optional[list[dict]] = None,
        drill_filters: Optional[dict] = None,
        user: Any = None,
        apply_render_cap: bool = True,
        workbook: Optional[dict] = None,
        workbook_hash: Optional[str] = None,
        workbook_missing: bool = False,
    ) -> dict:
        """Exécute une requête SQL contre Sage et retourne les résultats.

        Le format de retour dépend de widget_type :
        - "table" → {type: "table", columns, rows}
        - "chart" → {type: "chart", labels, datasets}
        - "kpi"   → {type: "kpi", value, label}
        - "grid"  → {type: "grid", columns, rows, sql, row_count, truncated,
                    execution_time_ms} — copie conforme result area /iris.
                    Pas de transformation (rendu brut via SqlResultGrid).
                    max_rows respecte le cap admin (parité /iris), pas le
                    cap 500 des widgets agrégés.

        **Mode classeur** (``workbook`` non-None, widget grille sauvegardé) :
        le classeur est la SOURCE DE VÉRITÉ — TOUTES ses feuilles SQL
        (``externalSource.type='sql_query'``) sont ré-exécutées fraîches ici
        (la 1ʳᵉ avec l'enveloppe filtres/drill du dashboard, les suivantes
        indépendantes — parité avec le mode legacy), les feuilles snapshot
        (manuelles, drill-down) gardent leurs données sauvegardées, et le
        payload porte ``workbook`` = classeur hydraté que le frontend ouvre
        via ``loadWorkbook()``. ``config["query"]``/``extra_tabs`` ne sont
        alors qu'un miroir (modale d'édition / rétro-compat) — pas exécutés.
        ``workbook_missing=True`` (fichier référencé mais illisible) →
        fallback legacy + indicateur pour le frontend.
        """
        # Mode classeur : substituer les requêtes du classeur AVANT toute
        # validation/exécution — le miroir config peut être périmé, le
        # classeur fait foi (single source of truth).
        wb_sql_sheets: list[dict] = []
        if workbook is not None and widget_type == "grid":
            from app.services.dashboard.widget_workbook_store import extract_sql_sheets

            wb_sql_sheets = extract_sql_sheets(workbook)
            if not wb_sql_sheets:
                # Classeur sans aucune feuille SQL (état anormal — la grille
                # widget a toujours sa feuille principale SQL) : fallback
                # legacy explicite plutôt qu'une grille vide silencieuse.
                logger.warning("Classeur widget sans feuille SQL — fallback config.")
                workbook = None
                workbook_missing = True

        query = (
            wb_sql_sheets[0]["query"]
            if (workbook is not None and wb_sql_sheets)
            else config.get("query", "")
        )
        if not query:
            return {"error": "Requête SQL vide."}

        # CWE-158 : certains drivers ODBC tronquent silencieusement la requête
        # au 1er NUL → la requête exécutée diffère de l'affichée (données
        # fausses silencieuses). Couvre TOUS les widgets SQL (table/chart/kpi/
        # grid), pas seulement les onglets additionnels. Parité avec
        # ``/api/datastore/sql/execute`` (datastore.py).
        if "\x00" in query:
            return {"error": "Caractère NUL interdit dans la requête."}

        # Validate SELECT-only (prevent writes from dashboard widgets).
        # Détection des verbes/patterns dangereux déléguée au validateur SSoT
        # ``check_sql_dangerous`` (app/services/ai/sql_validator) — déjà utilisé
        # par Iris/training_store. Il couvre PLUS que l'ancien blacklist inline :
        # 17 mots-clés (BACKUP/RESTORE/OPENROWSET/WAITFOR/SHUTDOWN…), les
        # préfixes de procédures (sp_/xp_), les patterns composites
        # (SELECT … INTO, BULK INSERT) ET le strip des commentaires SQL.
        # L'ancien blacklist laissait passer SELECT INTO / sp_ — or un user
        # (require_role "user") peut créer un widget SQL et la BDD connectée
        # n'est PAS garantie read-only (règle GÉNÉRICITÉ). Single source of
        # truth → pas de dérive entre ce guard et celui d'Iris.
        from app.services.ai.sql_validator import check_sql_dangerous

        stripped = query.strip().upper()
        if not stripped.startswith("SELECT") and not stripped.startswith("WITH"):
            return {"error": "Seules les requêtes SELECT sont autorisées."}
        if check_sql_dangerous(query):
            return {"error": "Seules les requêtes SELECT sont autorisées."}

        try:
            from app.services.database.query_executor import QueryExecutor

            executor = QueryExecutor()

            # Build combined WHERE clause from filters and drill-down
            effective_query = query
            sql_params = None
            where_parts = []
            all_params = []

            # 1. Regular dashboard filters
            if filter_state and filter_definitions:
                from app.services.dashboard.filter_service import build_sql_filter_clause

                where_clause, filter_params = build_sql_filter_clause(
                    filter_state, filter_definitions
                )
                if where_clause and filter_params:
                    where_parts.append(where_clause)
                    all_params.extend(filter_params)

            # 2. Drill-down filters (from chart click)
            if drill_filters:
                from app.services.dashboard.filter_service import build_drill_down_clause

                drill_clause, drill_params = build_drill_down_clause(drill_filters)
                if drill_clause:
                    where_parts.append(drill_clause)
                    all_params.extend(drill_params)

            if where_parts:
                combined_where = " AND ".join(where_parts)
                effective_query = f"SELECT * FROM ({query}) AS _wq WHERE {combined_where}"
                sql_params = tuple(all_params)

            # execute() returns QueryResult, supports params natively
            from app.services.data_access.enforcer import DataAccessDeniedError

            # Cap lignes : None pour TOUS les types → cap admin
            # ``DatabaseConnection.max_rows`` (doctrine no-double-cap).
            #
            # **#18c (triage caps 2026-06-10)** — l'ancien ``500`` pour les
            # widgets chart/kpi/table était appliqué AVANT
            # ``apply_transformation`` : un KPI « total CA » ou un graphe
            # agrégé était calculé sur les 500 premières lignes de la requête
            # → CHIFFRES FAUX silencieux sur le dashboard. Le souci de rendu
            # front (Plotly/HTML au-delà de ~500 points) est un problème de
            # SORTIE, pas d'entrée : il est géré APRÈS transformation par
            # ``_cap_widget_render`` (slice honnête + flags truncated/
            # row_count, mêmes clés que le widget grid).
            effective_max_rows = None

            # Mode classeur : isoler l'échec de la feuille PRINCIPALE (revue
            # adv. 2026-06-10). Les feuilles manuelles/snapshot du classeur
            # n'ont pas besoin de Sage — Sage down ne doit pas rendre
            # inconsultables des données saisies à la main (le bandeau
            # d'erreur se pose sur la feuille SQL concernée uniquement,
            # parité avec l'isolation par onglet des feuilles SQL extras).
            workbook_grid = workbook is not None and bool(wb_sql_sheets) and widget_type == "grid"
            qr = None
            main_error: Optional[str] = None
            try:
                qr = await executor.execute(
                    effective_query,
                    params=sql_params,
                    max_rows=effective_max_rows,
                    user=user,
                    rls_source="dashboard_widget",
                    require_user=True,
                )
            except DataAccessDeniedError as exc:
                if not workbook_grid:
                    return {"error": exc.user_message, "blocked_by": "data_access_rule"}
                main_error = exc.user_message
            except Exception:
                if not workbook_grid:
                    # Hors mode classeur : comportement historique (erreur
                    # widget globale via le except englobant de la méthode).
                    raise
                logger.warning(
                    "Erreur exécution feuille principale (widget classeur)", exc_info=True
                )
                main_error = "Erreur lors de l'exécution de la requête SQL."

            if main_error is not None and not apply_render_cap:
                # Chemin EXPORT/EMAIL : pas de classeur dans le payload — des
                # clés plates vides SANS erreur produiraient un fichier vide
                # silencieux. Erreur explicite comme avant.
                return {"error": main_error}

            if qr is not None:
                columns = qr.columns or []
                rows_as_dicts = qr.to_dicts()
                rows = (
                    [[row.get(col) for col in columns] for row in rows_as_dicts]
                    if rows_as_dicts
                    else []
                )
            else:
                columns, rows = [], []

            # ── Widget "grid" : retour brut, pas de transformation ni
            # d'aggrégation. Le frontend instancie GridTabManager direct
            # sur ces données (copie conforme de la result area /iris).
            # On expose row_count + truncated pour que l'UI puisse afficher
            # un badge "tronqué" cohérent avec /iris.
            if widget_type == "grid":
                grid_result = {
                    "type": "grid",
                    "columns": columns,
                    "rows": rows,
                    "sql": effective_query,
                    # ``source_sql`` = requête D'ORIGINE (config["query"]), AVANT
                    # tout filtre/période/drill. ``sql`` ci-dessus peut être la
                    # version filtre-wrappée ``SELECT * FROM (<query>) WHERE …``
                    # (cf. enveloppe filtres). Le front utilise ``source_sql``
                    # comme requête de la « feuille » principale (persistance) —
                    # persister la version wrappée corromprait config["query"].
                    "source_sql": query,
                    "row_count": getattr(qr, "row_count", len(rows)),
                    "truncated": bool(getattr(qr, "truncated", False)),
                    "execution_time_ms": getattr(qr, "execution_time_ms", 0),
                }
                # Onglets SQL additionnels (feature menu [+] « Requête SQL »).
                # Chacun est une requête INDÉPENDANTE ré-exécutée ici à chaque
                # affichage — donc toujours fraîche, et survit au refresh sans
                # snapshot ni localStorage (la requête vit dans la config du
                # widget, comme le tab principal). On NE les enveloppe PAS dans
                # les filtres/période du dashboard (décision produit : requêtes
                # ad-hoc autonomes). Le tab principal reste à plat (columns/rows)
                # pour rétro-compat (exports/consommateurs existants) ; les
                # extras vivent sous la clé ``tabs`` (absente si aucun).
                if workbook is not None and wb_sql_sheets:
                    # Mode classeur : les onglets SQL additionnels viennent du
                    # CLASSEUR (source de vérité), pas du miroir config.
                    wb_extras = [
                        {"label": s["label"], "query": s["query"]} for s in wb_sql_sheets[1:]
                    ]
                    extra_results = (
                        await self._execute_grid_extra_tabs(executor, wb_extras, user)
                        if wb_extras
                        else []
                    )
                    if extra_results:
                        grid_result["tabs"] = extra_results
                    main_fresh = {
                        "columns": columns,
                        "rows": rows,
                        "sql": query,  # requête d'ORIGINE (jamais filtre-wrappée)
                        "row_count": grid_result["row_count"],
                        "truncated": grid_result["truncated"],
                    }
                    if main_error is not None:
                        # Feuille principale en échec : vidée + bandeau (via
                        # _hydrate_workbook), le reste du classeur reste
                        # consultable. PAS d'``error`` top-level : il
                        # déclencherait le rendu « erreur seule » côté front
                        # et masquerait les feuilles manuelles.
                        main_fresh["error"] = main_error
                    grid_result["workbook"] = self._hydrate_workbook(
                        workbook, wb_sql_sheets, main_fresh, extra_results
                    )
                    grid_result["workbook_hash"] = workbook_hash
                    grid_result["workbook_mode"] = True
                    # Vue LIVE : les rows vivent UNIQUEMENT dans ``workbook``
                    # (pas de duplication flat+classeur = payload/cache ×2).
                    # Chemins EXPORT/EMAIL (apply_render_cap=False) : clés
                    # plates pleines, ``workbook`` retiré (les exporteurs
                    # consomment columns/rows/tabs comme avant).
                    if apply_render_cap:
                        grid_result["rows"] = []
                        for t in grid_result.get("tabs", []) or []:
                            t["rows"] = []
                    else:
                        grid_result.pop("workbook", None)
                    return grid_result

                extra_tabs = config.get("extra_tabs") if isinstance(config, dict) else None
                if isinstance(extra_tabs, list) and extra_tabs:
                    grid_result["tabs"] = await self._execute_grid_extra_tabs(
                        executor, extra_tabs, user
                    )
                if workbook_missing:
                    # Fichier classeur référencé mais absent/corrompu : on a
                    # servi le miroir config (données justes), mais les
                    # feuilles manuelles/mise en forme sont perdues — le
                    # frontend affiche un avertissement explicite (jamais de
                    # dégradation silencieuse).
                    grid_result["workbook_missing"] = True
                return grid_result

            # ── Pipeline v2 : si le widget a une transformation persistée
            # (décidée par l'Analyst LLM à la création), on la rejoue en
            # Python ici — la même recette, déterministe, à chaque refresh.
            # Les anciens widgets sans transformation passent dans le chemin
            # historique (backward-compat).
            # #18c — la troncature SOURCE (cap admin atteint au fetch) fausse
            # les agrégats : propagée à part (``source_truncated``) car le
            # message utilisateur diffère du cap de RENDU.
            source_truncated = bool(getattr(qr, "truncated", False))

            recipe = config.get("transformation") if isinstance(config, dict) else None
            if recipe:
                try:
                    from app.services.dashboard.widget_planner.transformations import (
                        TransformationError,
                        apply_transformation,
                    )

                    result = apply_transformation(columns, rows, recipe)
                    # Attache le chart_type s'il était stocké dans la render_spec
                    # (utile pour le front quand le widget est de type chart).
                    render_spec = config.get("render_spec") if isinstance(config, dict) else None
                    if (
                        isinstance(render_spec, dict)
                        and result.get("type") == "chart"
                        and render_spec.get("chart_type")
                    ):
                        result["chart_type"] = render_spec["chart_type"]
                    return self._cap_widget_render(result, source_truncated, apply_render_cap)
                except TransformationError as exc:
                    logger.warning(
                        "Widget transformation invalide : %s — fallback sur SQL brut",
                        exc,
                    )
                    # Fallback : on retombe sur le chemin historique ci-dessous

            # Chemin historique (pas de transformation persistée ou fallback)
            if widget_type == "chart" and rows:
                return self._cap_widget_render(
                    self._transform_sql_to_chart(columns, rows, chart_type),
                    source_truncated,
                    apply_render_cap,
                )
            elif widget_type == "kpi" and rows:
                return self._cap_widget_render(
                    self._transform_sql_to_kpi(columns, rows),
                    source_truncated,
                    apply_render_cap,
                )

            return self._cap_widget_render(
                {"type": "table", "columns": columns, "rows": rows},
                source_truncated,
                apply_render_cap,
            )

        except Exception:
            logger.warning("Erreur exécution SQL widget", exc_info=True)
            return {"error": "Erreur lors de l'exécution de la requête SQL."}

    def _hydrate_workbook(
        self,
        workbook: dict,
        sql_sheets: list[dict],
        main_fresh: dict,
        extra_results: list[dict],
    ) -> dict:
        """Fusionne les résultats SQL FRAIS dans le classeur sauvegardé.

        - Feuilles SQL (``sql_sheets``, indexées dans le classeur) : données
          remplacées par l'exécution fraîche. Une feuille en ERREUR est vidée
          + porte ``error`` (jamais de snapshot périmé présenté comme frais).
        - Si les colonnes fraîches diffèrent des colonnes sauvegardées (le SQL
          a changé de forme), l'état d'affichage indexé par colonne (ordre,
          masquage, tri, filtres, merges, cellDetails) est RESET — appliquer
          des index décalés produirait des données fausses silencieuses.
        - Feuilles snapshot (manuelles, drill-down, imports) : intactes.

        Mute ``workbook`` en place (objet local à la requête ; le cache widget
        fait sa propre copie profonde au ``set``).
        """
        tabs = workbook.get("tabs") or []
        fresh_by_index: dict[int, dict] = {}
        if sql_sheets:
            fresh_by_index[sql_sheets[0]["index"]] = main_fresh
        for offset, res in enumerate(extra_results or []):
            if 1 + offset < len(sql_sheets):
                fresh_by_index[sql_sheets[1 + offset]["index"]] = res
        for idx, fresh in fresh_by_index.items():
            if not (0 <= idx < len(tabs)) or not isinstance(tabs[idx], dict):
                continue
            tab = tabs[idx]
            fresh_cols = list(fresh.get("columns") or [])
            error = fresh.get("error")
            if error:
                tab["rows"] = []
                tab["columns"] = fresh_cols or list(tab.get("columns") or [])
                tab["totalRowCount"] = 0
                tab["truncated"] = False
                tab["error"] = error
                tab["isArrayFormat"] = True
                continue
            stored_cols = list(tab.get("columns") or [])
            if stored_cols != fresh_cols:
                for key in (
                    "columnOrder",
                    "hiddenCols",
                    "filters",
                    "merges",
                    "columnMetadata",
                    "cellDetails",
                ):
                    tab.pop(key, None)
                tab["sortColIndex"] = -1
                tab["sortDirection"] = None
            tab["columns"] = fresh_cols
            tab["rows"] = fresh.get("rows") or []
            tab["totalRowCount"] = fresh.get("row_count", len(tab["rows"]))
            tab["truncated"] = bool(fresh.get("truncated", False))
            tab["sql"] = fresh.get("sql") or tab.get("sql") or ""
            # Les rows fraîches du backend sont TOUJOURS des tableaux de
            # tableaux. Le flag sauvegardé peut être ``false`` (feuille SQL
            # enregistrée alors qu'elle retournait 0 ligne) — loadWorkbook
            # l'appliquerait aux rows fraîches et la grille lirait chaque
            # cellule par NOM de colonne sur un tableau → « null » partout
            # avec un compteur de lignes correct (revue adv. 2026-06-10,
            # données fausses silencieuses).
            tab["isArrayFormat"] = True
            tab.pop("error", None)
        return workbook

    async def _execute_grid_extra_tabs(
        self, executor: Any, extra_tabs: list, user: Any
    ) -> list[dict]:
        """Exécute les onglets SQL additionnels d'un widget grid.

        Contrat (cf. ``_fetch_sql_data`` branche grid) :

        - **Mêmes gardes sécurité que la requête principale** : SELECT/WITH +
          ``check_sql_dangerous`` (SSoT validateur Iris) + exécution avec RLS
          par utilisateur (``rls_source="dashboard_widget"``, ``require_user``).
          Chaque onglet est validé INDIVIDUELLEMENT (defense-in-depth : une
          config clonée/legacy/éditée en concurrence peut contenir un onglet
          invalide même si le save a validé).
        - **Requêtes indépendantes** : pas d'enveloppe par les filtres/période
          du dashboard (décision produit — ce sont des requêtes ad-hoc).
        - **Cap admin sur les lignes** (``max_rows=None`` → ``connector.max_rows``),
          parité avec le tab principal.
        - **Exécution séquentielle** : un widget peut avoir N onglets et N
          viewers peuvent rafraîchir en parallèle → on ne sature pas le pool
          de connexions Sage en lançant tout en parallèle.
        - **Succès partiel, jamais de donnée fausse silencieuse** : un onglet
          en échec porte un champ ``error`` explicite (et des colonnes/lignes
          vides) — distinct d'un résultat réellement vide — sans casser les
          autres onglets ni le widget.
        """
        from app.models.dashboard import DashboardWidget
        from app.services.ai.sql_validator import check_sql_dangerous
        from app.services.data_access.enforcer import DataAccessDeniedError

        out: list[dict] = []
        # Cap dur même si la validation au save a laissé passer (clone/import/
        # édition concurrente). Borne le coût (N requêtes × viewers × refresh).
        capped = extra_tabs[: DashboardWidget.MAX_GRID_EXTRA_TABS]
        for tab in capped:
            if not isinstance(tab, dict):
                continue
            label = tab.get("label") or "Requête"
            tab_query = tab.get("query")
            if not isinstance(tab_query, str) or not tab_query.strip():
                out.append(
                    {
                        "label": label,
                        "sql": "",
                        "columns": [],
                        "rows": [],
                        "row_count": 0,
                        "error": "Requête SQL vide.",
                    }
                )
                continue
            # CWE-158 : NUL → troncature silencieuse driver ODBC (la requête
            # exécutée diffère de l'affichée). Defense-in-depth runtime (une
            # config clonée/legacy/curl pourrait en contenir malgré validate()).
            if "\x00" in tab_query:
                out.append(
                    {
                        "label": label,
                        "sql": tab_query,
                        "columns": [],
                        "rows": [],
                        "row_count": 0,
                        "error": "Caractère NUL interdit dans la requête.",
                    }
                )
                continue
            stripped = tab_query.strip().upper()
            if not (
                stripped.startswith("SELECT") or stripped.startswith("WITH")
            ) or check_sql_dangerous(tab_query):
                out.append(
                    {
                        "label": label,
                        "sql": tab_query,
                        "columns": [],
                        "rows": [],
                        "row_count": 0,
                        "error": "Seules les requêtes SELECT sont autorisées.",
                    }
                )
                continue
            try:
                qr = await executor.execute(
                    tab_query,
                    params=None,
                    max_rows=None,  # cap admin (parité grid principal)
                    user=user,
                    rls_source="dashboard_widget",
                    require_user=True,
                )
            except DataAccessDeniedError as exc:
                out.append(
                    {
                        "label": label,
                        "sql": tab_query,
                        "columns": [],
                        "rows": [],
                        "row_count": 0,
                        "error": exc.user_message,
                        "blocked_by": "data_access_rule",
                    }
                )
                continue
            except Exception:
                logger.warning("Erreur exécution onglet SQL widget", exc_info=True)
                out.append(
                    {
                        "label": label,
                        "sql": tab_query,
                        "columns": [],
                        "rows": [],
                        "row_count": 0,
                        "error": "Erreur lors de l'exécution de la requête SQL.",
                    }
                )
                continue
            cols = qr.columns or []
            # qr.rows (dédup-proof) plutôt que to_dicts() — cf. même fix sur le
            # tab principal (to_dicts dédup les colonnes → perte silencieuse de
            # la 2e colonne homonyme). Fix revue adversariale 2026-06-09.
            tab_rows = [list(r) for r in (qr.rows or [])]
            out.append(
                {
                    "label": label,
                    "columns": cols,
                    "rows": tab_rows,
                    "sql": tab_query,
                    "row_count": getattr(qr, "row_count", len(tab_rows)),
                    "truncated": bool(getattr(qr, "truncated", False)),
                    "execution_time_ms": getattr(qr, "execution_time_ms", 0),
                }
            )
        return out

    # ── Export ─────────────────────────────────────────────────────────────

    async def export_dashboard(
        self,
        session: AsyncSession,
        dashboard_id: int,
        user_id: int,
        fmt: str = "csv",
        period_override: Optional[int] = None,
        user: Any = None,
    ) -> Optional[tuple[str, bytes, str]]:
        """Exporte toutes les données d'un dashboard en CSV ou Excel (owner only).

        Returns:
            Tuple (filename, file_bytes, content_type) ou None si non trouvé
            ou si le dashboard n'appartient pas au user courant.
        """
        from app.models.dashboard import Dashboard

        # Vérifier accès — strict owner only
        dash_stmt = select(Dashboard).where(
            Dashboard.id == dashboard_id, Dashboard.user_id == user_id
        )
        dash_result = await session.execute(dash_stmt)
        dashboard = dash_result.scalar_one_or_none()
        if not dashboard:
            return None

        dash_name = dashboard.name

        # Récupérer toutes les données — ``user`` propagé pour l'enforcement
        # RLS data-access des widgets SQL (sinon ``user=None`` → bypass legacy
        # de l'executor → export de données Sage NON filtrées, fuite cross-
        # périmètre). Les widgets metric restent scopés via ``user_id``.
        # ``force_refresh=True`` : un EXPORT doit toujours refléter les données
        # FRAÎCHES (le fichier n'a aucun indicateur de péremption, contrairement
        # à la vue live). Servir du cache dans un export = donnée fausse
        # silencieuse (règle consequences #5).
        all_data = await self.get_all_widget_data(
            session,
            dashboard_id,
            user_id,
            period_override=period_override,
            user=user,
            force_refresh=True,
            # #18c (revue adv. 2026-06-10) — pas de cap de RENDU sur un
            # export : la slice 500 (Plotly/DOM) livrait un fichier partiel
            # en silence. Le cap admin (fetch) reste la seule borne.
            apply_render_cap=False,
        )

        if fmt == "excel":
            return self._export_excel(dash_name, all_data, dashboard_id)
        else:
            return self._export_csv(dash_name, all_data, dashboard_id)

    @staticmethod
    def _safe_excel_sheet_name(wb: Any, base: str) -> str:
        """Nom de feuille Excel valide ET unique.

        Excel impose : longueur <= 31, sans les caractères ``[]:*?/\\``, et
        unicité dans le classeur. ``base`` provient d'un libellé utilisateur
        (titre d'onglet SQL) → on assainit + tronque + déduplique pour ne PAS
        faire crasher openpyxl (qui lève sur un nom invalide ou dupliqué).
        """
        invalid = set("[]:*?/\\")
        cleaned = "".join("_" if c in invalid else c for c in (base or "Feuille"))
        cleaned = cleaned.strip() or "Feuille"
        cleaned = cleaned[:31]
        if cleaned not in wb.sheetnames:
            return cleaned
        i = 2
        while True:
            suffix = f"_{i}"
            candidate = cleaned[: 31 - len(suffix)] + suffix
            if candidate not in wb.sheetnames:
                return candidate
            i += 1

    @staticmethod
    def _excel_autowidth(ws: Any) -> None:
        """Ajuste la largeur des colonnes (cap 50) — même logique que la
        boucle d'auto-width de ``_export_excel``, factorisée pour les feuilles
        d'onglets SQL additionnels (créées hors de la boucle principale)."""
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    def _export_csv(
        self, dash_name: str, all_data: dict, dashboard_id: int
    ) -> tuple[str, bytes, str]:
        """Génère un export CSV multi-widgets."""
        output = io.StringIO()
        writer = csv.writer(output)

        for widget_id_str, data in all_data.items():
            if data.get("error"):
                # Prod-loop task #16 — anti silent data loss : un widget en
                # erreur DOIT être rendu visible dans le fichier export, pas
                # silencieusement skipped (sinon décision business sur données
                # partielles + destinataire email reçoit un fichier tronqué).
                # csv_safe_cell : le message peut contenir du contenu pseudo-
                # utilisateur via exc.user_message (ligne 1263) → neutraliser
                # toute formule (=, +, -, @, \t, \r) avant écriture, sinon
                # OWASP CSV-injection vers les destinataires d'email.
                error_msg = csv_safe_cell(data.get("error") or "Erreur inconnue")
                writer.writerow([f"Widget {widget_id_str} — ÉCHEC DE RÉCUPÉRATION DES DONNÉES"])
                writer.writerow(["Erreur:", error_msg])
                writer.writerow([])  # separator
                continue

            data_type = data.get("type", "")
            widget_title = f"Widget {widget_id_str}"

            # csv_safe_cell sur TOUTES les cellules de données (pas seulement
            # le message d'erreur ci-dessus) : les valeurs viennent de la BDD
            # source (Sage) et peuvent commencer par =/+/-/@/\t/\r → formule
            # exécutée à l'ouverture Excel chez le destinataire de l'envoi
            # planifié (OWASP CSV-injection). Même neutralisation que la
            # branche if error: et que le helper to_csv_bytes (SSoT).
            if data_type == "kpi":
                writer.writerow([widget_title])
                writer.writerow(
                    [
                        csv_safe_cell(data.get("label", "Valeur")),
                        csv_safe_cell(data.get("value", "")),
                    ]
                )
                writer.writerow([])  # separator
            elif data_type == "table" or data_type == "grid":
                # Adversarial review CRIT-2 — 2026-05-26 : sans cette branche
                # 'grid', les widgets grid étaient silencieusement omis de
                # l'export (data_type == 'grid' ne matchait aucun elif →
                # silent data loss, exactement l'anti-pattern documenté par
                # le commentaire du if error: bloc ci-dessus). Le shape de
                # 'grid' est identique à 'table' (columns + rows) → même
                # rendu CSV.
                columns = data.get("columns", [])
                rows = data.get("rows", [])
                writer.writerow([widget_title])
                if columns:
                    writer.writerow([csv_safe_cell(c) for c in columns])
                for row in rows:
                    writer.writerow([csv_safe_cell(cell) for cell in row])
                writer.writerow([])  # separator
                # Onglets SQL additionnels (feature menu [+] « Requête SQL ») :
                # anti silent data loss — chaque onglet exporté en section
                # distincte (même politique que le widget principal et que le
                # bloc ``if error:`` ci-dessus). csv_safe_cell sur toutes les
                # cellules (OWASP CSV-injection vers le destinataire d'email).
                for tab in data.get("tabs") or []:
                    if not isinstance(tab, dict):
                        continue
                    tab_label = tab.get("label") or "Onglet"
                    writer.writerow([f"{widget_title} — {tab_label}"])
                    if tab.get("error"):
                        writer.writerow(["Erreur:", csv_safe_cell(tab.get("error"))])
                    else:
                        tab_columns = tab.get("columns", [])
                        tab_rows = tab.get("rows", [])
                        if tab_columns:
                            writer.writerow([csv_safe_cell(c) for c in tab_columns])
                        for row in tab_rows:
                            writer.writerow([csv_safe_cell(cell) for cell in row])
                    writer.writerow([])  # separator
            elif data_type == "chart":
                labels = data.get("labels", [])
                datasets = data.get("datasets", [])
                writer.writerow([widget_title])
                # Header: Label + dataset names
                header = ["Label"] + [
                    ds.get("label", f"Serie {i + 1}") for i, ds in enumerate(datasets)
                ]
                writer.writerow([csv_safe_cell(h) for h in header])
                for idx, label in enumerate(labels):
                    row = [label]
                    for ds in datasets:
                        ds_data = ds.get("data", [])
                        row.append(ds_data[idx] if idx < len(ds_data) else "")
                    writer.writerow([csv_safe_cell(cell) for cell in row])
                writer.writerow([])  # separator

        content = output.getvalue()
        # BOM for Excel UTF-8 compatibility
        file_bytes = b"\xef\xbb\xbf" + content.encode("utf-8")
        safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in dash_name)[:50]
        filename = f"dashboard_{safe_name}.csv"
        return filename, file_bytes, "text/csv; charset=utf-8"

    def _export_excel(
        self, dash_name: str, all_data: dict, dashboard_id: int
    ) -> tuple[str, bytes, str]:
        """Génère un export Excel multi-widgets (un onglet par widget)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.warning("openpyxl non disponible, fallback CSV")
            return self._export_csv(dash_name, all_data, dashboard_id)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        sheet_idx = 0

        for widget_id_str, data in all_data.items():
            if data.get("error"):
                # Prod-loop task #16 — anti silent data loss : un widget en
                # erreur DOIT créer une feuille ÉCHEC visible (sinon le
                # destinataire d'un envoi planifié reçoit un Excel tronqué
                # sans signal). Cf. _export_csv pour le pendant CSV.
                # csv_safe_cell : openpyxl n'évalue pas les formules au
                # write, mais le destinataire peut re-exporter en CSV ; on
                # neutralise donc les préfixes formule par cohérence.
                error_msg = csv_safe_cell(data.get("error") or "Erreur inconnue")
                error_sheet_name = f"Widget {widget_id_str} ÉCHEC"[:31]
                ws = wb.create_sheet(title=error_sheet_name)
                sheet_idx += 1
                ws.append([f"Widget {widget_id_str} — ÉCHEC DE RÉCUPÉRATION DES DONNÉES"])
                ws.append(["Erreur:", error_msg])
                error_font = Font(bold=True, color="FF0000")
                for cell in ws[1]:
                    cell.font = error_font
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        except (TypeError, AttributeError):
                            pass
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
                continue

            data_type = data.get("type", "")
            sheet_name = f"Widget {widget_id_str}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            sheet_idx += 1

            # excel_safe_cell sur les cellules de données : openpyxl écrit une
            # chaîne ``"=..."`` comme une FORMULE évaluée à l'ouverture →
            # même classe de CSV/formula-injection que le CSV. Variante
            # type-preserving (les nombres/dates restent natifs pour rester
            # sommables/graphables côté destinataire). Cf. output_safety.
            if data_type == "kpi":
                ws.append(
                    [
                        excel_safe_cell(data.get("label", "Valeur")),
                        excel_safe_cell(data.get("value", "")),
                    ]
                )
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
            elif data_type == "table" or data_type == "grid":
                # Voir _export_csv pour le pendant (adversarial CRIT-2).
                columns = data.get("columns", [])
                rows = data.get("rows", [])
                if columns:
                    ws.append([excel_safe_cell(c) for c in columns])
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                for row in rows:
                    ws.append([excel_safe_cell(cell) for cell in row])
                # Onglets SQL additionnels (feature menu [+] « Requête SQL ») :
                # une feuille par onglet (anti silent data loss). Nom assaini +
                # dédupliqué ; auto-width inline (la boucle externe ne traite
                # que la feuille principale ``ws``).
                for tab in data.get("tabs") or []:
                    if not isinstance(tab, dict):
                        continue
                    tab_label = tab.get("label") or "Onglet"
                    ws_tab = wb.create_sheet(
                        title=self._safe_excel_sheet_name(wb, f"{widget_id_str}-{tab_label}")
                    )
                    sheet_idx += 1
                    if tab.get("error"):
                        ws_tab.append([f"Widget {widget_id_str} — {tab_label}"])
                        ws_tab.append(["Erreur:", excel_safe_cell(tab.get("error"))])
                    else:
                        tab_columns = tab.get("columns", [])
                        tab_rows = tab.get("rows", [])
                        if tab_columns:
                            ws_tab.append([excel_safe_cell(c) for c in tab_columns])
                            for cell in ws_tab[1]:
                                cell.font = header_font
                                cell.fill = header_fill
                        for row in tab_rows:
                            ws_tab.append([excel_safe_cell(cell) for cell in row])
                    self._excel_autowidth(ws_tab)
            elif data_type == "chart":
                labels = data.get("labels", [])
                datasets = data.get("datasets", [])
                header = ["Label"] + [
                    ds.get("label", f"Serie {i + 1}") for i, ds in enumerate(datasets)
                ]
                ws.append([excel_safe_cell(h) for h in header])
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                for idx, label in enumerate(labels):
                    row = [label]
                    for ds in datasets:
                        ds_data = ds.get("data", [])
                        row.append(ds_data[idx] if idx < len(ds_data) else "")
                    ws.append([excel_safe_cell(cell) for cell in row])

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        if not wb.sheetnames:
            ws = wb.create_sheet(title="Vide")
            ws.append(["Aucune donnée disponible."])

        buf = io.BytesIO()
        wb.save(buf)
        file_bytes = buf.getvalue()

        safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in dash_name)[:50]
        filename = f"dashboard_{safe_name}.xlsx"
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return filename, file_bytes, content_type

    # ── Utilitaires ────────────────────────────────────────────────────────

    def get_available_metrics(self) -> dict:
        """Retourne la liste des métriques prédéfinies disponibles."""
        return AVAILABLE_METRICS
